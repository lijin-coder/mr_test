
import glob
import os
import xml.dom.minidom as xmldom
import re
from lxml import etree
import time
import math
import openpyxl
import mr_globel as gl
import mr_utils
import shutil

write_info = lambda info :gl.str_info.append(str(info))


def test51_file_integrity():
    mr_utils.test_out_data_item_header("test_51")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)
    # time_sub = mr_utils.get_two_linux_time_sub()
    try:
        #统计预期的文件总数,计算: 测量总时长/测量周期
        startTimestamp = 0
        endTimestamp = mr_utils.get_timestamp_by_str_format('2030-01-01T00:00:00Z', '%Y-%m-%dT%H:%M:%SZ')
        predict_file_num =  float(gl.TEST_CONF['test_total_time'])*3600 / (60 * int(gl.MR_CONF['UploadPeriod']) )
        if gl.MR_CONF['SampleEndTime'] != '0001-01-01T00:00:00Z' and gl.MR_CONF['SampleBeginTime'] != '0001-01-01T00:00:00Z' :
            startTimestamp = mr_utils.get_timestamp_by_str_format(gl.MR_CONF['SampleBeginTime'], '%Y-%m-%dT%H:%M:%SZ') / 1000
            endTimestamp = mr_utils.get_timestamp_by_str_format(gl.MR_CONF['SampleEndTime'], '%Y-%m-%dT%H:%M:%SZ') / 1000
            spec_timestamp = endTimestamp - startTimestamp
            predict_file_num = (spec_timestamp - spec_timestamp%60) / (60 * int(gl.MR_CONF['UploadPeriod']) )
        print (predict_file_num)
        print (startTimestamp)
        print (endTimestamp)
        mr_integrity_flag = 0
        mro_file_num = 0
        mre_file_num = 0
        mrs_file_num = 0
        #MRS文件数量统计
        for time_entity in gl.MR_DICT:
            file_time_stamp = mr_utils.get_timestamp_by_str_format(time_entity, '%Y%m%d%H%M%S') / 1000 - 8 * 60*60

            if file_time_stamp < startTimestamp  or file_time_stamp > endTimestamp :
                print ('hello ss')
                continue
            mro_file_name = gl.MR_DICT[time_entity][0]['MRO']
            mre_file_name = gl.MR_DICT[time_entity][0]['MRE']
            mrs_file_name = gl.MR_DICT[time_entity][0]['MRS']
            if mro_file_name != '':
                if os.path.getsize(mro_file_name) == 0 or mr_utils.MR_xml_file_name_accuracy(mro_file_name) == False :
                    mr_integrity_flag = mr_integrity_flag | (0x1 << 1)
                else :
                    mro_file_num += 1
            if mre_file_name != '':
                if os.path.getsize(mre_file_name) == 0 or mr_utils.MR_xml_file_name_accuracy(mre_file_name) == False:
                    mr_integrity_flag = mr_integrity_flag | (0x1 << 1)
                else:
                    mre_file_num += 1
            if mrs_file_name != '':
                if os.path.getsize(mrs_file_name) == 0 or mr_utils.MR_xml_file_name_accuracy(mrs_file_name) == False:
                    mr_integrity_flag = mr_integrity_flag | (0x1 << 1)
                else:
                    mrs_file_num += 1


        with open(gl.OUT_PATH, 'a') as file_object:
            file_object.write(date_time + " | ")
            file_object.write(gl.TEST_CONF['test_total_time'] + " | ")
            file_object.write(str(len(gl.TEST_CONF['enbid'].strip(',').split(','))) + " | ")
            file_object.write(str(predict_file_num * (len(gl.MR_CONF['MeasureType'].strip(',').split(',')))) + ' | ')
            file_object.write(str(mrs_file_num) + ' | ')
            file_object.write(str(mro_file_num) + ' | ')
            file_object.write(str(mre_file_num) + ' | ')
            file_object.write(str(mr_integrity_flag & (0x1 << 3) == 0 and \
                                  (predict_file_num == mrs_file_num ) \
                                  if re.search(r'MRS',gl.MR_CONF['MeasureType']) != None \
                                  else  (mrs_file_num == 0))  + ' | ')
            file_object.write(str(mr_integrity_flag & (0x1 << 1) == 0 and \
                                  (predict_file_num == mro_file_num  ) \
                                  if re.search('MRO', gl.MR_CONF['MeasureType']) != None \
                                  else mro_file_num == 0) + ' | ')
            file_object.write(str(mr_integrity_flag & (0x1 << 2) == 0 and \
                                  (predict_file_num == mre_file_num )\
                                  if re.search('MRE', gl.MR_CONF['MeasureType']) != None \
                                  else mre_file_num == 0 ) + ' | ')
    except Exception as result:
        raise Exception('-%s- <%s>'%(str(result.__traceback__.tb_lineno),result))


def test52_file_integrity():
    mr_utils.test_out_data_item_header("test_52")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    out_dict = {}
    output_temp_dict = {}
    out_sort_list = ["MR.RSRP", "MR.RSRQ","MR.PowerHeadRoom", "MR.SinrUL" ]

    for time_entity in gl.MR_DICT:
        try:
            mrs_temp_dom = gl.MR_DICT[time_entity][gl.MR_TYPE['MRS']]['xmldom']
            mrs_file_entity = gl.MR_DICT[time_entity][0]['MRS']
            mrs_root = mrs_temp_dom.documentElement
            if out_dict.__contains__(mrs_file_entity) == False:
                out_dict[mrs_file_entity] = {"MR.RSRP": {'num':0, 'list':[]}, "MR.RSRQ":{'num':0, 'list':[]}, "MR.SinrUL":{'num':0, 'list':[]}, "MR.PowerHeadRoom":{'num':0, 'list':[]}}
            for enb_entity in mrs_root.getElementsByTagName('eNB'):
                enbid = int(enb_entity.getAttribute('id'))
                for measurement_entity in enb_entity.getElementsByTagName('measurement'):
                    mrName = measurement_entity.getAttribute('mrName')
                    object_list = measurement_entity.getElementsByTagName('object')
                    for mr_name_entity in out_dict[mrs_file_entity]:
                        if mr_name_entity == mrName:
                            for object_entity in object_list:
                                eci_id = int(object_entity.getAttribute('id').strip(':').split(':')[0])
                                cellid_ret_list = mr_utils.is_cell_id_exist(eci_id)
                                if  cellid_ret_list[0] == False:
                                    mr_utils.out_text_dict_append_list(output_temp_dict, mrs_file_entity, '[%s]-[cellid:%d not in list] '%(mrName, cellid_ret_list[1]))
                                else:
                                    if cellid_ret_list[1] not in out_dict[mrs_file_entity][mr_name_entity]['list']:
                                        out_dict[mrs_file_entity][mr_name_entity]['num'] += 1
                                        out_dict[mrs_file_entity][mr_name_entity]['list'].append(cellid_ret_list[1])
                            break
        except Exception as result:
            raise Exception('-%s- <%s> : %s '%(str(result.__traceback__.tb_lineno),result, gl.MR_DICT[time_entity][0]['MRS']))


    with open(gl.OUT_PATH, 'a') as file_object:
        cell_num = len(gl.TEST_CONF['cellid'].strip(',').split(','))
        file_object.write(date_time + " | ")
        file_object.write(gl.TEST_CONF['test_total_time'] + " | ")
        for file_name in out_dict:
            file_object.write('\n=====> ' + file_name + ': |')
            temp_flag_test = 0
            for mr_name in out_sort_list:
                cell_num_mrs_item = out_dict[file_name][mr_name]['num']
                file_object.write(str(cell_num) + " | ")
                if (gl.MR_CONF['MeasureItems'] == 'all' or re.search(mr_name, gl.MR_CONF['MeasureItems']) != None) and cell_num_mrs_item != cell_num:
                    temp_flag_test = 1
                    mr_utils.out_text_dict_append_list(output_temp_dict, file_name, '[%s]=<cell num not match>:[%d(mrs)]!=[%d(target)]'%(mr_name, cell_num_mrs_item, cell_num))
                if mr_utils.is_mr_item_need_exist(mr_name) == False and cell_num_mrs_item != 0:
                    temp_flag_test = 1
                    mr_utils.out_text_dict_append_list(output_temp_dict, file_name, '[{0}]=<MeasureItems not have {0}>:cellnum=[{1} ->target:0]'.format(mr_name, cell_num_mrs_item))
            file_object.write(str(temp_flag_test == 0))
            if output_temp_dict.__contains__(file_name) == True and len(output_temp_dict[file_name]) != 0:
                file_object.write('\n===> error:\n')
                for i in range(len(output_temp_dict[file_name])):
                    file_object.write('\t\t' + output_temp_dict[file_name][i] + '\n')


def test53_file_integrity():
    mr_utils.test_out_data_item_header("test_53")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)
    file_list_mrs = glob.glob(gl.MR_TEST_PATH + '*MRS*.xml')
    text_consistent = {}
    subfram_dict = {}
    for file_mrs_name in file_list_mrs:
        try:
            subfram_dict[file_mrs_name] = {}
            for cell_id in gl.TEST_CONF['cellid'].strip(',').split(','):
                subfram_dict[file_mrs_name][cell_id] = {}
                for sub_frame_key in gl.MR_CONF['SubFrameNum'].strip(',').split(','):
                    subfram_dict[file_mrs_name][cell_id][sub_frame_key] = 0
                subfram_dict[file_mrs_name][cell_id]['consistent'] = 0
            mrs_dom = xmldom.parse(file_mrs_name)
            mrs_root = mrs_dom.documentElement
            measurement_list = mrs_root.getElementsByTagName('measurement')
            for measurement_entity in measurement_list:
                if measurement_entity.getAttribute('mrName') == 'MR.ReceivedIPower' :
                    object_list = measurement_entity.getElementsByTagName('object')
                    for object_entity in object_list:
                        eci_id = int(object_entity.getAttribute('id').strip(':').split(':')[0])
                        cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)
                        subfram_id = object_entity.getAttribute('id').strip(':').split(':')[2]
                        test_count = 0
                        if subfram_dict[file_mrs_name].__contains__(str(cell_id_ret_list[1])) == False:
                            mr_utils.out_text_dict_append_list(text_consistent, file_mrs_name, '[MR.ReceivedIPower]:<cell_id[%d] not in list>'%(cell_id_ret_list[1]))
                        else:
                            for subfram_key in subfram_dict[file_mrs_name][str(cell_id_ret_list[1])]:
                                if subfram_key == subfram_id:
                                    subfram_dict[file_mrs_name][str(cell_id_ret_list[1])][subfram_key] += 1
                                    test_count = 1
                                    break
                            if test_count == 0:
                                mr_utils.out_text_dict_append_list(text_consistent, file_mrs_name, '[MR.ReceivedIPower]:<surplus rip [cellid:%d]-[%s]>'%(cell_id_ret_list[1],subfram_id))
                                subfram_dict[file_mrs_name][str(cell_id_ret_list[1])]['consistent'] = 1
                    break
            for cell_id_entity in subfram_dict[file_mrs_name]:
                for subfram_entity in gl.MR_CONF['SubFrameNum'].strip(',').split(','):
                    if subfram_dict[file_mrs_name][cell_id_entity][subfram_entity] != 1 and mr_utils.is_mr_item_need_exist('MR.ReceivedIPower') == True:
                        mr_utils.out_text_dict_append_list(text_consistent, file_mrs_name, '[MR.ReceivedIPower]:<confusion rip [cellid:%s]-[%s]>'%(cell_id_entity,subfram_entity))
                        subfram_dict[file_mrs_name][cell_id_entity]['consistent'] = 1
        except Exception as result:
            raise Exception('-%s- <%s> : [%s]'%(str(result.__traceback__.tb_lineno),result, file_mrs_name))
    with open(gl.OUT_PATH, 'a') as file_object:
        file_object.write(date_time + " | ")
        file_object.write(gl.TEST_CONF['test_total_time'] + " | \n")
        for mrs_file in subfram_dict:
            file_object.write('======>%s:'%(mrs_file))
            for cell_id_entity in subfram_dict[mrs_file]:
                file_object.write(' | cellid:[%s]-ripnum:[%d] | %s | \n'%(cell_id_entity, len(subfram_dict[mrs_file][cell_id_entity]) - 1,
                        str(subfram_dict[mrs_file][cell_id_entity]['consistent'] == 0)))
            if text_consistent.__contains__(mrs_file) == True and len(text_consistent[mrs_file]) != 0:
                file_object.write('====>error:\n')
                for i in range(len(text_consistent[mrs_file])):
                    file_object.write(text_consistent[mrs_file][i] + '\n')


def test54_file_integrity():
    mr_utils.test_out_data_item_header("test_54")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    file_list_mrs = glob.glob(gl.MR_TEST_PATH + '*MRS*.xml')
    prbnum_list = []
    if re.search(r',', gl.MR_CONF['PrbNum']) == None:
        for i in range(int(gl.MR_CONF['PrbNum'].strip('....').split('....')[0]), 1 + int(gl.MR_CONF['PrbNum'].strip('....').split('....')[1])):
            prbnum_list.append(str(i))
    else:
        prbnum_list = gl.MR_CONF['PrbNum'].strip(',').split(',')

    ripprb_num = len(gl.MR_CONF['SubFrameNum'].strip(',').split(',')) * len(prbnum_list)
    is_consistence = 0
    text_consistence = {}
    mrs_ripprb_dict = {}

    for time_entity in gl.MR_DICT:
        try:
            file_name = gl.MR_DICT[time_entity][0]['MRS']
            mrs_ripprb_dict[file_name] = {}
            mrs_dom = gl.MR_DICT[time_entity][gl.MR_TYPE['MRS']]['xmldom']
            mrs_root = mrs_dom.documentElement
            measurement_list = mrs_root.getElementsByTagName('measurement')
            for measurement_entity in measurement_list:
                if measurement_entity.getAttribute('mrName') == 'MR.RIPPRB':
                    object_list = measurement_entity.getElementsByTagName('object')
                    for object_entity in object_list:
                        id_str_list = object_entity.getAttribute('id').strip(':').split(':')
                        eci_id = int(id_str_list[0])
                        rip = id_str_list[2]
                        prb = id_str_list[3]
                        cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)

                        if mrs_ripprb_dict[file_name].__contains__(str(cell_id_ret_list[1])) == False:
                            mrs_ripprb_dict[file_name][str(cell_id_ret_list[1])] = []
                        if re.search(rip, gl.MR_CONF['SubFrameNum']) != None and prb in prbnum_list:
                            mrs_ripprb_dict[file_name][str(cell_id_ret_list[1])].append(rip + ':' + prb)
                        else:
                            mr_utils.out_text_dict_append_list(text_consistence, file_name, '<ripprb not match>:[%s:%s]'%(rip,prb))
                    break
        except Exception as result:
            raise Exception('-%s- <%s> : [%s]'%(str(result.__traceback__.tb_lineno),result, gl.MR_DICT[time_entity][0]['MRS']))
    with open(gl.OUT_PATH, 'a') as file_object:
        file_object.write(date_time + " | ")
        file_object.write(gl.TEST_CONF['test_total_time'] + " | \n")
        for file_name in mrs_ripprb_dict:
            file_object.write('====>%s: \n'%(file_name))
            for cell_id_str in mrs_ripprb_dict[file_name]:
                ripnum = len(mrs_ripprb_dict[file_name][cell_id_str])
                file_object.write('\t\t<[cell:%s]:[%d] | %s>\n'%(cell_id_str, ripnum, ripnum == ripprb_num ))
            if text_consistence.__contains__(file_name) == True:
                file_object.write('\terror:\n')
                for i in range(len(text_consistence[file_name])):
                    file_object.write(text_consistence[file_name][i] + '\n')


def test55_file_integrity():
    mr_utils.test_out_data_item_header("test_55")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)
    out_list_mro = ['MR.LteScRSRP', 'MR.LteScRSRQ', 'MR.LteScPHR']
    out_list_mrs = ['MR.RSRP', 'MR.RSRQ', 'MR.PowerHeadRoom']
    for time_list_entity in gl.MR_DICT:
        mro_count_dict = {'MR.LteScRSRP':{}, 'MR.LteScRSRQ':{}, 'MR.LteScPHR':{}}
        mrs_count_dict = {'MR.RSRP':{}, 'MR.RSRQ':{}, 'MR.PowerHeadRoom':{}}
        pos_dict = {'MR.LteScRSRP': {'pos':2}, 'MR.LteScRSRQ': {'pos':3}, 'MR.LteScPHR':{'pos':4}}
        try:
            mro_measurement_list = []
            mrs_measurement_list = []
            mro_file_name = gl.MR_DICT[time_list_entity][0]['MRO']
            mrs_file_name = gl.MR_DICT[time_list_entity][0]['MRS']
            if re.search(r'MRO', gl.MR_CONF['MeasureType']) != None :
                mro_dom = gl.MR_DICT[time_list_entity][gl.MR_TYPE['MRO']]['xmldom']
                mro_root = mro_dom.documentElement
                mro_measurement_list = mro_root.getElementsByTagName('measurement')

            if re.search(r'MRS', gl.MR_CONF['MeasureType']) != None and gl.MR_DICT[time_list_entity][0]['MRS'] != '':
                mrs_dom = gl.MR_DICT[time_list_entity][gl.MR_TYPE['MRS']]['xmldom']
                mrs_root = mrs_dom.documentElement
                mrs_measurement_list = mrs_root.getElementsByTagName('measurement')

            temp_pos = 2
            for mro_measurement_entity in mro_measurement_list:
                smr_list = mro_measurement_entity.getElementsByTagName('smr')
                for smr_entity in smr_list:
                    smr_str = smr_entity.firstChild.data
                    mr_utils.get_mr_item_pos(pos_dict, smr_str)
                    for mro_mr_Name in mro_count_dict:
                        for smr_name_entity in smr_str.strip().split(' '):
                            item_count = 0
                            if mro_mr_Name == smr_name_entity:
                                for object_entity in mro_measurement_entity.getElementsByTagName('object'):
                                    eci_id = int(object_entity.getAttribute('id'))
                                    cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)
                                    object_id = str(cell_id_ret_list[1])
                                    if mro_count_dict[mro_mr_Name].__contains__(object_id) == False:
                                        mro_count_dict[mro_mr_Name][object_id] = 0
                                    value_list = object_entity.getElementsByTagName('v')
                                    for value_entity in value_list:
                                        if value_entity.firstChild.data.strip().split(' ')[pos_dict[mro_mr_Name]['pos']].isdigit() == True :
                                            mro_count_dict[mro_mr_Name][object_id] += 1
                                    item_count += 1
                                    #if mro_count_dict[mro_mr_Name][object_id] != item_count:
                                    #print (mro_mr_Name + " | " + str(mro_count_dict[mro_mr_Name][object_id]))
                                break
                        temp_pos += 1

            for mrs_measurement_entity in mrs_measurement_list:
                for mrs_mr_Name in mrs_count_dict:
                    if mrs_measurement_entity.getAttribute('mrName') == mrs_mr_Name:
                        for object_entity in mrs_measurement_entity.getElementsByTagName('object'):
                            eci_id = int(object_entity.getAttribute('id'))
                            cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)
                            object_id = str(cell_id_ret_list[1])
                            temp_count = 0
                            for value_entity in object_entity.getElementsByTagName('v'):
                                temp_count += mr_utils.add_digital_string(value_entity.firstChild.data)

                            mrs_count_dict[mrs_mr_Name][object_id] = temp_count

            with open(gl.OUT_PATH, 'a') as file_object:
                file_object.write(time_list_entity + ' : \n')
                file_object.write(date_time + " | ")
                file_object.write(gl.TEST_CONF['test_total_time'] + " | ")

                file_object.write('[MRO文件上报采样点数]' + " | ")
                for mr_name in out_list_mro:
                    file_object.write('{')
                    for object_id in mro_count_dict[mr_name]:
                        file_object.write(' [%s]=%s '%(object_id, str(mro_count_dict[mr_name][object_id])) )
                    file_object.write('} | ')
                file_object.write('\n' + " " * (len(date_time) + len(gl.TEST_CONF['test_total_time']) + 6))

                file_object.write('[MRS文件统计采样点数]' + " | ")
                for mr_name in out_list_mrs:
                    file_object.write('{')
                    for object_id in mrs_count_dict[mr_name]:
                        file_object.write(' [%s]=%s '%(object_id, str(mrs_count_dict[mr_name][object_id])))
                    file_object.write('} | ')
                file_object.write('\n' + " " * (len(date_time) + len(gl.TEST_CONF['test_total_time']) + 6))

                file_object.write('[统计采样点数是否完整]' + " | ")
                file_object.write(str(mro_count_dict['MR.LteScRSRP'] == mrs_count_dict['MR.RSRP'] or \
                                      mr_utils.is_mr_item_need_exist('MR.LteScRSRP') == False or \
                                      mr_utils.is_mr_item_need_exist('MR.RSRP') == False) + ' | ')
                file_object.write(str(mro_count_dict['MR.LteScRSRQ'] == mrs_count_dict['MR.RSRQ'] or \
                                      mr_utils.is_mr_item_need_exist('MR.LteScRSRQ') == False or \
                                      mr_utils.is_mr_item_need_exist('MR.RSRQ') == False) + ' | ')
                file_object.write(str(mro_count_dict['MR.LteScPHR'] == mrs_count_dict['MR.PowerHeadRoom'] or \
                                      mr_utils.is_mr_item_need_exist('MR.LteScPHR') == False or \
                                      mr_utils.is_mr_item_need_exist('MR.PowerHeadRoom') == False) + ' | \n')
        except Exception as result:
            raise Exception('-%s- <%s> MRO:[%s] MRS:[%s]'%(str(result.__traceback__.tb_lineno),result, gl.MR_DICT[time_list_entity][0]['MRO'],gl.MR_DICT[time_list_entity][0]['MRS']))


def test56_file_integrity():
    mr_utils.test_out_data_item_header("test_56")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    out_mro_list = ['MR.LteScRIP', 'MR.LteScSinrUL']
    out_mrs_list = ['MR.ReceivedIPower', 'MR.SinrUL']
    for time_list_entity in gl.MR_DICT:
        try:
            mro_count_dict = {'MR.LteScRIP':{}, 'MR.LteScSinrUL':{}}
            mrs_count_dict = {'MR.ReceivedIPower':{}, 'MR.SinrUL':{}}
            sinrul_pos_dict = {'MR.LteScSinrUL':{'pos':5}}
            mro_measurement_list = []
            mrs_measurement_list = []
            mro_file_name = gl.MR_DICT[time_list_entity][0]['MRO']
            mrs_file_name = gl.MR_DICT[time_list_entity][0]['MRS']
            if re.search(r'MRO', gl.MR_CONF['MeasureType']) != None :
                mro_dom = gl.MR_DICT[time_list_entity][gl.MR_TYPE['MRO']]['xmldom']
                mro_root = mro_dom.documentElement
                mro_measurement_list = mro_root.getElementsByTagName('measurement')
            if re.search(r'MRS', gl.MR_CONF['MeasureType']) != None and gl.MR_DICT[time_list_entity][0]['MRS'] != '':
                mrs_dom = gl.MR_DICT[time_list_entity][gl.MR_TYPE['MRS']]['xmldom']
                mrs_root = mrs_dom.documentElement
                mrs_measurement_list = mrs_root.getElementsByTagName('measurement')

            for mr_name in out_mro_list:
                for mro_measurement_entity in mro_measurement_list:
                    for smr_entity in mro_measurement_entity.getElementsByTagName('smr'):
                        if mr_name in smr_entity.firstChild.data.strip().split(' ') :
                            for object_entity in mro_measurement_entity.getElementsByTagName('object'):
                                object_value = object_entity.getAttribute('id')
                                eci_id = int(object_value.strip(':').split(':')[0])
                                cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)
                                object_id = str(cell_id_ret_list[1])
                                object_prb = ""
                                if mr_name == 'MR.LteScRIP':
                                    object_prb = object_value.strip(':').split(':')[2]

                                if mro_count_dict[mr_name].__contains__(object_id) == False :

                                    if mr_name == 'MR.LteScRIP':
                                        mro_count_dict[mr_name][object_id] = {}
                                    else:
                                        mro_count_dict[mr_name][object_id] = 0


                                if mr_name == 'MR.LteScRIP' and mro_count_dict[mr_name][object_id].__contains__(object_prb) == False:
                                    mro_count_dict[mr_name][object_id][object_prb] = 0

                                for value_entity in object_entity.getElementsByTagName('v'):
                                    if mr_name == 'MR.LteScRIP' and value_entity.firstChild.data.isdigit() == True:
                                        mro_count_dict[mr_name][object_id][object_prb] += 1
                                    elif mr_name == 'MR.LteScSinrUL' :
                                        mr_utils.get_mro_pos_list_by_mapping(sinrul_pos_dict, smr_entity.firstChild.data)
                                        if value_entity.firstChild.data.strip().split(' ')[sinrul_pos_dict['MR.LteScSinrUL']['pos']].isdigit() == True:
                                            #TODO:在这里还需要商量一下,如果在MRO中没有统计,那么在MRS中,采样点数目是对不上的
                                            #if int(value_entity.firstChild.data.split(' ')[5]) != 0 :
                                            mro_count_dict[mr_name][object_id] += 1

            for mr_name in out_mrs_list:
                for measurement_entity in mrs_measurement_list:
                    if mr_name == measurement_entity.getAttribute('mrName'):
                        for object_entity in measurement_entity.getElementsByTagName('object'):
                            eci_id = int(object_entity.getAttribute('id').strip(':').split(':')[0])
                            cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)
                            object_id = str(cell_id_ret_list[1])
                            object_prb = ''
                            if mr_name == 'MR.ReceivedIPower':
                                object_prb = object_entity.getAttribute('id').strip(':').split(':')[2]

                            temp_num = 0
                            for value_entity in object_entity.getElementsByTagName('v'):
                                temp_num += mr_utils.add_digital_string(value_entity.firstChild.data)

                            if mr_name == 'MR.ReceivedIPower':
                                if mrs_count_dict[mr_name].__contains__(object_id) == False:
                                        mrs_count_dict[mr_name][object_id] = {}
                                if mrs_count_dict[mr_name][object_id].__contains__(object_prb) == False:
                                    mrs_count_dict[mr_name][object_id][object_prb] = 0
                                mrs_count_dict[mr_name][object_id][object_prb] = temp_num
                            elif mr_name == 'MR.SinrUL':
                                if mrs_count_dict[mr_name].__contains__(object_id) == False:
                                    mrs_count_dict[mr_name][object_id] = 0
                                mrs_count_dict[mr_name][object_id] = temp_num

            with open(gl.OUT_PATH, 'a') as file_object:
                file_object.write(time_list_entity + ' : \n')
                file_object.write(date_time + " | ")
                file_object.write(gl.TEST_CONF['test_total_time'] + " | ")

                file_object.write('[MRO文件上报采样点数]' + " | ")
                for mr_name in out_mro_list:
                    for object_id in mro_count_dict[mr_name]:
                        if mr_name == 'MR.LteScRIP':
                            file_object.write('[' + object_id + ']={' )
                            for object_prb in mro_count_dict[mr_name][object_id]:
                                file_object.write('[' + object_prb + ']=' + str(mro_count_dict[mr_name][object_id][object_prb]) + ' ')
                            file_object.write('} | ')
                        if mr_name == 'MR.LteScSinrUL':
                            file_object.write('[' + object_id + ']=' + str(mro_count_dict[mr_name][object_id]) + ' | ')

                file_object.write('\n' + " " * (len(date_time) + len(gl.TEST_CONF['test_total_time']) + 6))

                file_object.write('[MRS文件统计采样点数]' + " | ")
                for mr_name in out_mrs_list:
                    for object_id in mrs_count_dict[mr_name]:
                        if mr_name == 'MR.ReceivedIPower':
                            file_object.write('[' + object_id + ']={' )
                            for object_prb in mrs_count_dict[mr_name][object_id]:
                                file_object.write('[' + object_prb + ']=' + str(mrs_count_dict[mr_name][object_id][object_prb]) + ' ')
                            file_object.write('} | ')
                        if mr_name == 'MR.SinrUL':
                            file_object.write('[' + object_id + ']=' + str(mrs_count_dict[mr_name][object_id]) + ' | ')

                file_object.write('\n' + " " * (len(date_time) + len(gl.TEST_CONF['test_total_time']) + 6))

                test_flag = 0
                for object_id in mrs_count_dict['MR.ReceivedIPower']:
                    for object_prb in mrs_count_dict['MR.ReceivedIPower'][object_id]:
                        if mro_count_dict['MR.LteScRIP'].__contains__(object_id) == False:
                            continue
                        if mro_count_dict['MR.LteScRIP'][object_id].__contains__(object_prb) == True and mro_count_dict['MR.LteScRIP'][object_id][object_prb] != mrs_count_dict['MR.ReceivedIPower'][object_id][object_prb]:
                            test_flag = 1
                file_object.write('[统计采样点数是否完整]' + " | ")
                file_object.write(str(0 == test_flag) + ' | ')
                file_object.write(str(mro_count_dict['MR.LteScSinrUL'] == mrs_count_dict['MR.SinrUL'] or\
                                      mr_utils.is_mr_item_need_exist('MR.LteScSinrUL') == False or \
                                      mr_utils.is_mr_item_need_exist('MR.SinrUL') == False) + ' | \n')
        except Exception as result:
            raise Exception('-%s- <%s> : MRO:[%s] MRS:[%s]'%(str(result.__traceback__.tb_lineno),result, gl.MR_DICT[time_list_entity][0]['MRO'], gl.MR_DICT[time_list_entity][0]['MRS']))

def test57_file_integrity():
    mr_utils.test_out_data_item_header("test_57")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    sheet_row_dict = {'MR.RSRP':[3,49], 'MR.RSRQ':[51,67],'MR.PowerHeadRoom':[114,176], 'MR.ReceivedIPower':[178,229],'MR.SinrUL':[807,842], 'MR.RIPPRB':[844, 895]}

    rsrp_smr = 'MR.RSRP.00 MR.RSRP.01 MR.RSRP.02 MR.RSRP.03 MR.RSRP.04 MR.RSRP.05 MR.RSRP.06 MR.RSRP.07 MR.RSRP.08 MR.RSRP.09 MR.RSRP.10 MR.RSRP.11 MR.RSRP.12 MR.RSRP.13 MR.RSRP.14 MR.RSRP.15 MR.RSRP.16 MR.RSRP.17 MR.RSRP.18 MR.RSRP.19 MR.RSRP.20 MR.RSRP.21 MR.RSRP.22 MR.RSRP.23 MR.RSRP.24 MR.RSRP.25 MR.RSRP.26 MR.RSRP.27 MR.RSRP.28 MR.RSRP.29 MR.RSRP.30 MR.RSRP.31 MR.RSRP.32 MR.RSRP.33 MR.RSRP.34 MR.RSRP.35 MR.RSRP.36 MR.RSRP.37 MR.RSRP.38 MR.RSRP.39 MR.RSRP.40 MR.RSRP.41 MR.RSRP.42 MR.RSRP.43 MR.RSRP.44 MR.RSRP.45 MR.RSRP.46 MR.RSRP.47 '
    rsrq_smr = 'MR.RSRQ.00 MR.RSRQ.01 MR.RSRQ.02 MR.RSRQ.03 MR.RSRQ.04 MR.RSRQ.05 MR.RSRQ.06 MR.RSRQ.07 MR.RSRQ.08 MR.RSRQ.09 MR.RSRQ.10 MR.RSRQ.11 MR.RSRQ.12 MR.RSRQ.13 MR.RSRQ.14 MR.RSRQ.15 MR.RSRQ.16 MR.RSRQ.17 '
    rip_smr = 'MR.ReceivedIPower.00 MR.ReceivedIPower.01 MR.ReceivedIPower.02 MR.ReceivedIPower.03 MR.ReceivedIPower.04 MR.ReceivedIPower.05 MR.ReceivedIPower.06 MR.ReceivedIPower.07 MR.ReceivedIPower.08 MR.ReceivedIPower.09 MR.ReceivedIPower.10 MR.ReceivedIPower.11 MR.ReceivedIPower.12 MR.ReceivedIPower.13 MR.ReceivedIPower.14 MR.ReceivedIPower.15 MR.ReceivedIPower.16 MR.ReceivedIPower.17 MR.ReceivedIPower.18 MR.ReceivedIPower.19 MR.ReceivedIPower.20 MR.ReceivedIPower.21 MR.ReceivedIPower.22 MR.ReceivedIPower.23 MR.ReceivedIPower.24 MR.ReceivedIPower.25 MR.ReceivedIPower.26 MR.ReceivedIPower.27 MR.ReceivedIPower.28 MR.ReceivedIPower.29 MR.ReceivedIPower.30 MR.ReceivedIPower.31 MR.ReceivedIPower.32 MR.ReceivedIPower.33 MR.ReceivedIPower.34 MR.ReceivedIPower.35 MR.ReceivedIPower.36 MR.ReceivedIPower.37 MR.ReceivedIPower.38 MR.ReceivedIPower.39 MR.ReceivedIPower.40 MR.ReceivedIPower.41 MR.ReceivedIPower.42 MR.ReceivedIPower.43 MR.ReceivedIPower.44 MR.ReceivedIPower.45 MR.ReceivedIPower.46 MR.ReceivedIPower.47 MR.ReceivedIPower.48 MR.ReceivedIPower.49 MR.ReceivedIPower.50 MR.ReceivedIPower.51 MR.ReceivedIPower.52 '
    ripprb_smr = 'MR.RIPPRB.00 MR.RIPPRB.01 MR.RIPPRB.02 MR.RIPPRB.03 MR.RIPPRB.04 MR.RIPPRB.05 MR.RIPPRB.06 MR.RIPPRB.07 MR.RIPPRB.08 MR.RIPPRB.09 MR.RIPPRB.10 MR.RIPPRB.11 MR.RIPPRB.12 MR.RIPPRB.13 MR.RIPPRB.14 MR.RIPPRB.15 MR.RIPPRB.16 MR.RIPPRB.17 MR.RIPPRB.18 MR.RIPPRB.19 MR.RIPPRB.20 MR.RIPPRB.21 MR.RIPPRB.22 MR.RIPPRB.23 MR.RIPPRB.24 MR.RIPPRB.25 MR.RIPPRB.26 MR.RIPPRB.27 MR.RIPPRB.28 MR.RIPPRB.29 MR.RIPPRB.30 MR.RIPPRB.31 MR.RIPPRB.32 MR.RIPPRB.33 MR.RIPPRB.34 MR.RIPPRB.35 MR.RIPPRB.36 MR.RIPPRB.37 MR.RIPPRB.38 MR.RIPPRB.39 MR.RIPPRB.40 MR.RIPPRB.41 MR.RIPPRB.42 MR.RIPPRB.43 MR.RIPPRB.44 MR.RIPPRB.45 MR.RIPPRB.46 MR.RIPPRB.47 MR.RIPPRB.48 MR.RIPPRB.49 MR.RIPPRB.50 MR.RIPPRB.51 MR.RIPPRB.52 '
    phr_smr = 'MR.PowerHeadRoom.00 MR.PowerHeadRoom.01 MR.PowerHeadRoom.02 MR.PowerHeadRoom.03 MR.PowerHeadRoom.04 MR.PowerHeadRoom.05 MR.PowerHeadRoom.06 MR.PowerHeadRoom.07 MR.PowerHeadRoom.08 MR.PowerHeadRoom.09 MR.PowerHeadRoom.10 MR.PowerHeadRoom.11 MR.PowerHeadRoom.12 MR.PowerHeadRoom.13 MR.PowerHeadRoom.14 MR.PowerHeadRoom.15 MR.PowerHeadRoom.16 MR.PowerHeadRoom.17 MR.PowerHeadRoom.18 MR.PowerHeadRoom.19 MR.PowerHeadRoom.20 MR.PowerHeadRoom.21 MR.PowerHeadRoom.22 MR.PowerHeadRoom.23 MR.PowerHeadRoom.24 MR.PowerHeadRoom.25 MR.PowerHeadRoom.26 MR.PowerHeadRoom.27 MR.PowerHeadRoom.28 MR.PowerHeadRoom.29 MR.PowerHeadRoom.30 MR.PowerHeadRoom.31 MR.PowerHeadRoom.32 MR.PowerHeadRoom.33 MR.PowerHeadRoom.34 MR.PowerHeadRoom.35 MR.PowerHeadRoom.36 MR.PowerHeadRoom.37 MR.PowerHeadRoom.38 MR.PowerHeadRoom.39 MR.PowerHeadRoom.40 MR.PowerHeadRoom.41 MR.PowerHeadRoom.42 MR.PowerHeadRoom.43 MR.PowerHeadRoom.44 MR.PowerHeadRoom.45 MR.PowerHeadRoom.46 MR.PowerHeadRoom.47 MR.PowerHeadRoom.48 MR.PowerHeadRoom.49 MR.PowerHeadRoom.50 MR.PowerHeadRoom.51 MR.PowerHeadRoom.52 MR.PowerHeadRoom.53 MR.PowerHeadRoom.54 MR.PowerHeadRoom.55 MR.PowerHeadRoom.56 MR.PowerHeadRoom.57 MR.PowerHeadRoom.58 MR.PowerHeadRoom.59 MR.PowerHeadRoom.60 MR.PowerHeadRoom.61 MR.PowerHeadRoom.62 MR.PowerHeadRoom.63 '
    sinrul_smr = 'MR.SinrUL.00 MR.SinrUL.01 MR.SinrUL.02 MR.SinrUL.03 MR.SinrUL.04 MR.SinrUL.05 MR.SinrUL.06 MR.SinrUL.07 MR.SinrUL.08 MR.SinrUL.09 MR.SinrUL.10 MR.SinrUL.11 MR.SinrUL.12 MR.SinrUL.13 MR.SinrUL.14 MR.SinrUL.15 MR.SinrUL.16 MR.SinrUL.17 MR.SinrUL.18 MR.SinrUL.19 MR.SinrUL.20 MR.SinrUL.21 MR.SinrUL.22 MR.SinrUL.23 MR.SinrUL.24 MR.SinrUL.25 MR.SinrUL.26 MR.SinrUL.27 MR.SinrUL.28 MR.SinrUL.29 MR.SinrUL.30 MR.SinrUL.31 MR.SinrUL.32 MR.SinrUL.33 MR.SinrUL.34 MR.SinrUL.35 MR.SinrUL.36'

    format_standard_dict = {'MR.RSRP':{'smr':rsrp_smr, 'v':48}, 'MR.RSRQ':{'smr':rsrq_smr, 'v':18}, 'MR.ReceivedIPower':{'smr':rip_smr, 'v':53}, 'MR.RIPPRB':{'smr':ripprb_smr,'v':53}, 'MR.SinrUL':{'smr':sinrul_smr, 'v':37}, 'MR.PowerHeadRoom':{'smr':phr_smr, 'v':64}}
    out_text_list = {}
    try:
        out_mrs_flag_dict = {'MR.RSRP':3, 'MR.RSRQ':3, 'MR.ReceivedIPower':3, 'MR.RIPPRB':3, 'MR.SinrUL':3, 'MR.PowerHeadRoom':3 }
        file_list_mrs = glob.glob(gl.MR_TEST_PATH + '*MRS*.xml')
        for time_entity in gl.MR_DICT:
            out_mrs_flag_dict = {'MR.RSRP':3, 'MR.RSRQ':3, 'MR.ReceivedIPower':3, 'MR.RIPPRB':3, 'MR.SinrUL':3, 'MR.PowerHeadRoom':3 }
            file_name = gl.MR_DICT[time_entity][0]['MRS']
            mrs_dom = gl.MR_DICT[time_entity][gl.MR_TYPE['MRS']]['xmldom']
            mrs_root = mrs_dom.documentElement
            for measurement_entity in mrs_root.getElementsByTagName('measurement'):
                for standard_dict_key in format_standard_dict:
                    if standard_dict_key == measurement_entity.getAttribute('mrName'):
                        #mrName匹配成功
                        out_mrs_flag_dict[standard_dict_key] -= 1
                        #smr数据正确是否
                        smr_list = measurement_entity.getElementsByTagName('smr')
                        if len(smr_list) == 1 and smr_list[0].firstChild.data.strip() == format_standard_dict[standard_dict_key]['smr'].strip():
                            out_mrs_flag_dict[standard_dict_key] -= 1
                        #value对应的个数正确是否
                        test_value_num_flag = 0
                        for object_entity in measurement_entity.getElementsByTagName('object'):
                            for value_entity in object_entity.getElementsByTagName('v'):
                                if format_standard_dict[standard_dict_key]['v'] != len(value_entity.firstChild.data.strip().split(' ')):
                                    test_value_num_flag = 1
                        if test_value_num_flag == 0:
                            out_mrs_flag_dict[standard_dict_key] -= 1
            #判断mrs文件输出是否有问题
            for mrs_dict_key in out_mrs_flag_dict:
                if out_mrs_flag_dict[mrs_dict_key] != 0 and mr_utils.is_mr_item_need_exist(mrs_dict_key) == True:
                    mr_utils.out_text_dict_append_list(out_text_list, file_name, '-[%s not match]' % (mrs_dict_key))
                elif out_mrs_flag_dict[mrs_dict_key] != 3 and mr_utils.is_mr_item_need_exist(mrs_dict_key) == False:
                    mr_utils.out_text_dict_append_list(out_text_list, file_name, '-[%s not match]' % (mrs_dict_key))
            if out_text_list.__contains__(file_name) == True:
                break


        full_name = os.path.join(gl.SOURCE_PATH, gl.XLS_NAME)
        if int(gl.TEST_CONF['is_57_out_excel']) == 1:
            excel_workbook = openpyxl.load_workbook(full_name)
            work_sheet = excel_workbook.worksheets[0]
            sheet_rows = work_sheet.max_row

            for mrs_dict_key in out_mrs_flag_dict:
                if out_mrs_flag_dict[mrs_dict_key] != 0:
                    for i in range(sheet_row_dict[mrs_dict_key][0], sheet_row_dict[mrs_dict_key][1]+1):
                        work_sheet.cell(i, 20, 'N')
                else:
                    for i in range(sheet_row_dict[mrs_dict_key][0], sheet_row_dict[mrs_dict_key][1]+1):
                        work_sheet.cell(i, 20, 'Y')
            excel_workbook.save(filename=full_name)

        with open(gl.OUT_PATH, 'a') as file_object:
            file_object.write(date_time + " | ")
            file_object.write(gl.TEST_CONF['test_total_time'] + " | " )
            if int(gl.TEST_CONF['is_57_out_excel']) == 1:
                file_object.write('结果已写入->'  )
                file_object.write(gl.XLS_NAME)

            if len(out_text_list) != 0:
                file_object.write('\nerror:\n')
                for file_name in out_text_list:
                    file_object.write('======> %s:'%(file_name))
                    for i in range(len(out_text_list[file_name])):
                        file_object.write('\t%s\n' % (out_text_list[file_name][i]))
            file_object.write('\n')
    except Exception as result:
        raise Exception('-%s- <%s>'%(str(result.__traceback__.tb_lineno),result))

def test58_file_integrity():
    mr_utils.test_out_data_item_header("test_58")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    mro_file_list = glob.glob(gl.MR_TEST_PATH + '*MRO*.xml')

    cellid = int(gl.TEST_CONF['cellid'])
    ideal_sample_num = int(int(gl.MR_CONF['UploadPeriod']) * 60 * 1000 / int(gl.MR_CONF['SamplePeriod']))

    test_cell_id = int(gl.TEST_CONF['cellid']) | int(gl.TEST_CONF['cellid'])  << 8
    out_mro_list = ['MR.LteScRSRP','MR.LteScRSRQ','MR.LteScPHR','MR.LteScSinrUL']
    out_text_list = {}
    xml_mro_item_dict = {"MR.LteScRSRP":2,'MR.LteNcRSRP':3,'MR.LteScRSRQ':4,'MR.LteNcRSRQ':5,'MR.LteScPHR':6, 'MR.LteScRIP':7, 'MR.LteScSinrUL':8 }
    temp_loop_flag = 0
    is_object_empty = 0

    with open(gl.OUT_PATH, 'a') as file_object:
        file_object.write(date_time + " | ")
        file_object.write(gl.TEST_CONF['test_total_time'] + " | " )

    for time_entity in gl.MR_DICT:
        try:
            temp_test_flag_dict = {'MR.LteScRSRP':0,'MR.LteScRSRQ':0,'MR.LteScPHR':0,'MR.LteScSinrUL':0,'MR.LteScRIP':0, 'MR.LteNcRSRP':0, 'MR.LteNcRSRQ':0}
            mro_file = gl.MR_DICT[time_entity][0]['MRO']
            mro_dom = gl.MR_DICT[time_entity][gl.MR_TYPE['MRO']]['xmldom']
            mro_root = mro_dom.documentElement
            mro_item_dict = {}

            for enb_entity in mro_root.getElementsByTagName('eNB'):
                if mr_utils.is_enb_id_exist(enb_entity.getAttribute('id')) == True:
                    for measurement_entity in enb_entity.getElementsByTagName('measurement'):
                        if len(measurement_entity.getElementsByTagName('object')) == 0:
                            is_object_empty = 1
                            if measurement_entity.getElementsByTagName('smr')[0].firstChild.data != 'MR.LteScRIP' and\
                                mr_utils.is_mr_item_need_exist('MR.LteScRSRP') == False and\
                                mr_utils.is_mr_item_need_exist('MR.LteScRSRQ') == False and\
                                mr_utils.is_mr_item_need_exist('MR.LteScPHR') == False and\
                                mr_utils.is_mr_item_need_exist('MR.LteScSinrUL') == False :
                                mr_utils.out_text_dict_append_list(out_text_list, mro_file, 'L3 data None')
                            if measurement_entity.getElementsByTagName('smr')[0].firstChild.data == 'MR.LteScRIP' and\
                                    mr_utils.is_mr_item_need_exist('MR.LteScRIP') == False:
                                mr_utils.out_text_dict_append_list(out_text_list, mro_file, 'L2 data None')
                            break
                        for smr_entity in measurement_entity.getElementsByTagName('smr'):
                            for object_entity in measurement_entity.getElementsByTagName('object'):
                                eci_id = int(object_entity.getAttribute('id').strip(':').split(':')[0])
                                cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)
                                object_id = str(cell_id_ret_list[1])
                                object_ue_id = object_id + "|" + object_entity.getAttribute('MmeUeS1apId')
                                if cell_id_ret_list[0] == False:
                                    continue
                                if mro_item_dict.__contains__(object_ue_id) == False  :
                                    mro_item_dict[object_ue_id] = {'MR.LteScRSRP':{'pos':2, 'TimeStamp':0, 'range':[0, 97], 'flag':3, 'num':0, 'item_num':{}}, 'MR.LteScRSRQ':{'pos':3, 'TimeStamp':0, 'range':[0,34],'flag':3, 'num':0, 'item_num':{}},
                                'MR.LteScPHR':{'pos':4, 'TimeStamp':0, 'range':[0,63],'flag':3, 'num':0, 'item_num':{}}, 'MR.LteScSinrUL':{'pos':5, 'TimeStamp':0, 'range':[0,36],'flag':3, 'num':0, 'item_num':{}},
                                'MR.LteScRIP':{'pos':0, 'TimeStamp':0, 'range':[0,511], 'flag':3,'num':0, 'prbnum':{'item_num':{}}}, 'MR.LteNcRSRP':{'pos':8, 'TimeStamp':0, 'range':[0,97],'flag':3, 'num':0, 'item_num':{}},
                                'MR.LteNcRSRQ':{'pos':9, 'TimeStamp':0, 'range':[0,34], 'flag':3, 'num':0, 'item_num':{} } }
                                mr_utils.get_mr_item_pos(mro_item_dict[object_ue_id], smr_entity.firstChild.data)
                                for mr_name_entity in mro_item_dict[object_ue_id]:
                                    #mr_Name匹配上, flag-1
                                    if re.search(mr_name_entity, smr_entity.firstChild.data) == None:
                                        continue
                                    if mr_name_entity == smr_entity.firstChild.data.strip().split(' ')[mro_item_dict[object_ue_id][mr_name_entity]['pos']]:
                                        mro_item_dict[object_ue_id][mr_name_entity]['flag'] -= 1
                                        if mr_name_entity != 'MR.LteScRIP':
                                            #判断 时间戳是否满足 MR测量的 SamplePeriod
                                            if mro_item_dict[object_ue_id][mr_name_entity]['TimeStamp'] == 0:
                                                mro_item_dict[object_ue_id][mr_name_entity]['TimeStamp'] = mr_utils.get_timestamp_by_str_format(object_entity.getAttribute('TimeStamp'))
                                            else:
                                                time_spec = mr_utils.get_timestamp_by_str_format(object_entity.getAttribute('TimeStamp')) - mro_item_dict[object_ue_id][mr_name_entity]['TimeStamp']
                                                if time_spec != int(gl.MR_CONF['SamplePeriod']):
                                                    temp_test_flag_dict[mr_name_entity] += 1
                                                    mr_utils.out_text_dict_append_list(out_text_list, mro_file, '[{0}]= << TimeStamp duplicate >> TimeStamp:[{1}]\n'.format(mr_name_entity, object_entity.getAttribute('TimeStamp')) \
                                                            if time_spec == 0 else '[{0}]= << TimeStamp gap inaccurate >> TimeStamp:[{1}]\n'.format(mr_name_entity, object_entity.getAttribute('TimeStamp')))
                                                    temp_loop_flag = 1
                                                #当出现时间间隔错误, 继续往下检索
                                                mro_item_dict[object_ue_id][mr_name_entity]['TimeStamp'] = mr_utils.get_timestamp_by_str_format(object_entity.getAttribute('TimeStamp'))

                                        else:

                                            prbnum =  str(mr_utils.is_cell_id_exist(int(object_entity.getAttribute('id').strip(':').split(':')[0]))[1]) + ':' + object_entity.getAttribute('id').strip(':').split(':')[2]
                                            if mro_item_dict[object_ue_id][mr_name_entity]['prbnum'].__contains__(prbnum) == False:
                                                mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum] = {'num':0, 'TimeStamp':0}
                                            mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['num'] += 1
                                            #TimeStamp 判断
                                            if mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['TimeStamp'] == 0:
                                                mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['TimeStamp'] = mr_utils.get_timestamp_by_str_format(object_entity.getAttribute('TimeStamp'))
                                            else:
                                                time_spec = mr_utils.get_timestamp_by_str_format(object_entity.getAttribute('TimeStamp')) - mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['TimeStamp']
                                                if time_spec != int(gl.MR_CONF['SamplePeriod']):
                                                    temp_test_flag_dict[mr_name_entity] += 1
                                                    if len(out_text_list[mro_file]) == 0 or re.search(object_entity.getAttribute('TimeStamp'), out_text_list[mro_file][len(out_text_list[mro_file]) - 1]) == None:
                                                        mr_utils.out_text_dict_append_list(out_text_list, mro_file, '[{0}]=<TimeStamp duplicate> TimeStamp:{1}\n'.format(mr_name_entity, object_entity.getAttribute('TimeStamp')) \
                                                          if time_spec == 0 else '[{0}]=<TimeStamp gap inaccurate> TimeStamp:{1}\n'.format(mr_name_entity, object_entity.getAttribute('TimeStamp')))
                                                    temp_loop_flag = 1
                                            mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['TimeStamp'] = mr_utils.get_timestamp_by_str_format(object_entity.getAttribute('TimeStamp'))

                                        if temp_test_flag_dict[mr_name_entity] == 0:
                                            mro_item_dict[object_ue_id][mr_name_entity]['flag'] -= 1
                                        value_temp_test = 0
                                        for value_entity in object_entity.getElementsByTagName('v'):
                                            value_num = -10
                                            if value_entity.firstChild.data.strip().split(' ')[mro_item_dict[object_ue_id][mr_name_entity]['pos']].isdigit() == True:
                                                value_num = int(value_entity.firstChild.data.strip().split(' ')[mro_item_dict[object_ue_id][mr_name_entity]['pos']])
                                            else:
                                                value_temp_test = 1
                                                continue
                                            if value_num < mro_item_dict[object_ue_id][mr_name_entity]['range'][0] or value_num > mro_item_dict[object_ue_id][mr_name_entity]['range'][1]:
                                                temp_test_flag_dict[mr_name_entity] += 1
                                                mr_utils.out_text_dict_append_list(out_text_list, mro_file, '[' + mr_name_entity + ']={' + 'value Comfusion:' + str(value_num) + '=[' + str(mro_item_dict[object_ue_id][mr_name_entity]['range']) + ']'  + '}\n')
                                                temp_loop_flag = 1
                                                value_temp_test = 1
                                            else:
                                                value_temp_test = 0

                                        mro_item_dict[object_ue_id][mr_name_entity]['num'] += 1

                                        if temp_test_flag_dict[mr_name_entity] == 0:
                                            mro_item_dict[object_ue_id][mr_name_entity]['flag'] -= 1

            with open(gl.OUT_PATH, 'a') as file_object:
                file_object.write('\n [' + mro_file + ']:\n')
                file_object.write('[' + str(cellid) + ']=[')
                file_object.write(str(ideal_sample_num) + 'or' + str(ideal_sample_num+1) + '] |')
                test_integrity_pqsh = 0
                for object_id in mro_item_dict:
                    if re.search(r'NIL', object_id) != None:
                        continue
                    file_object.write( object_id + ":")
                    for mr_name in out_mro_list:
                        file_object.write(str(mro_item_dict[object_id][mr_name]['num']) + ' | ')
                    if mro_item_dict[object_id][mr_name]['num'] != ideal_sample_num and mro_item_dict[object_id][mr_name]['num'] != ideal_sample_num + 1:
                        test_integrity_pqsh = 1
                    if temp_test_flag_dict[mr_name] != 0:
                        test_integrity_pqsh = 1
                file_object.write(str(test_integrity_pqsh == 0) + ' | \n')

                idea_rip_num = ideal_sample_num * len(gl.MR_CONF['SubFrameNum'].strip(',').split(','))

                file_object.write(date_time + " | ")
                file_object.write(gl.TEST_CONF['test_total_time'] + " | " )
                file_object.write('ideal rip_num=' + str(idea_rip_num) + 'or' + str((ideal_sample_num+1) * len(gl.MR_CONF['SubFrameNum'].strip(',').split(','))) + ' | ')

                sum_all_prbnum_count = 0
                for object_id in mro_item_dict:
                    if re.search(r'NIL', object_id) == None:
                        continue
                    file_object.write('{')
                    for prbnum_id in mro_item_dict[object_id]['MR.LteScRIP']['prbnum']:
                        if mro_item_dict[object_id]['MR.LteScRIP']['prbnum'][prbnum_id].__contains__('num') == True:
                            file_object.write(' [' + prbnum_id + ']=' + str(mro_item_dict[object_id]['MR.LteScRIP']['prbnum'][prbnum_id]['num'])  )
                            sum_all_prbnum_count += mro_item_dict[object_id]['MR.LteScRIP']['prbnum'][prbnum_id]['num']
                    file_object.write('  [total]=' + str(mro_item_dict[object_id]['MR.LteScRIP']['num']) + ' } | ')

                result_div = sum_all_prbnum_count*1.0 / (idea_rip_num) * 1.0

                rip_accuracy_index = result_div if result_div == 1.0 else sum_all_prbnum_count*1.0 / ((ideal_sample_num + 1) * len(gl.MR_CONF['SubFrameNum'].strip(',').split(','))) * 1.0
                file_object.write(str( rip_accuracy_index) + ' | \n')
        except  Exception as result:
            raise Exception('-%s- [%s]:<%s>'%(str(result.__traceback__.tb_lineno),gl.MR_DICT[time_entity][0]['MRO'], result))

    with open(gl.OUT_PATH, 'a') as file_object:
        if int(gl.TEST_CONF['is_58_out_excel']) == 1:
            file_object.write('\n结果已写入->'  )
            file_object.write(gl.XLS_NAME)

        if temp_loop_flag == 1:
            file_object.write('\nerror:\n')

            for mr_file in out_text_list:
                file_object.write('=====>' + mr_file + ':\n')
                for i in range(len(out_text_list[mr_file])):
                    file_object.write('\t' + out_text_list[mr_file][i])
        file_object.write('\n')
    if int(gl.TEST_CONF['is_58_out_excel']) == 1:
        full_name = os.path.join(gl.SOURCE_PATH, gl.XLS_NAME)
        excel_workbook = openpyxl.load_workbook(full_name)
        work_sheet = excel_workbook.worksheets[2]

        for mrs_dict_key in temp_test_flag_dict:
            if temp_test_flag_dict[mrs_dict_key] == 0:
                work_sheet.cell(xml_mro_item_dict[mrs_dict_key], 17, 'Y')
            else:
                work_sheet.cell(xml_mro_item_dict[mrs_dict_key], 17, 'N')
        excel_workbook.save(filename=full_name)



def test59_file_integrity():
    mr_utils.test_out_data_item_header("test_59")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    mre_event_begin_timestamp = mr_utils.get_timestamp_by_str_format(gl.TEST_CONF['mre_begin_time'])
    mre_event_end_timestamp = mr_utils.get_timestamp_by_str_format(gl.TEST_CONF['mre_end_time'])
    range_list = [[0,97], [0,97], [0,34], [0,34], [0, 41589], [0, 503], [0, 41589],  [0, 503], [0,1023], [0,63], [0,7], [0,7]]

    out_mre_list = ['A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'B1', 'B2']

    mre_smr_head = 'MR.LteScRSRP MR.LteNcRSRP MR.LteScRSRQ MR.LteNcRSRQ MR.LteScEarfcn MR.LteScPci MR.LteNcEarfcn MR.LteNcPci MR.GsmNcellBcch MR.GsmNcellCarrierRSSI MR.GsmNcellNcc MR.GsmNcellBcc'
    mre_smr_list = ['MR.LteScRSRP' ,'MR.LteNcRSRP' ,'MR.LteScRSRQ' ,'MR.LteNcRSRQ' ,'MR.LteScEarfcn' ,'MR.LteScPci' ,'MR.LteNcEarfcn' ,'MR.LteNcPci' ,'MR.GsmNcellBcch' ,'MR.GsmNcellCarrierRSSI' ,'MR.GsmNcellNcc' ,'MR.GsmNcellBcc']




    xml_out_dict = {'A1':{'item':[2,5], 'event':2}, 'A2':{'item':[6,9], 'event':6}, 'A3':{'item':[10,13], 'event':10}, 'A4':{'item':[14,17], 'event':14}, 'A5':{'item':[18,25], 'event':18},
                    'A6':{'item':[26,33], 'event':26}, 'B1':{'item':[34,37], 'event':34}, 'B2':{'item':[38, 45], 'event':38}}

    mre_file_list = glob.glob(gl.MR_TEST_PATH + '*MRE*.xml')

    out_mre_text = {}
    mre_conf_dict = {}
    for mre_event in out_mre_list:
        gl.test_mre_event_num_dict[mre_event]['num'] = 0
        gl.test_mre_event_num_dict[mre_event]['time_str'] = []
    for time_entity in gl.MR_DICT:
        try:
            event_deal_list = []
            mre_conf_dict = {'A1':{'flag':0, 'pos':[0,2,4,5], 'range':range_list, 'error_pos_list':[]},
                         'A2':{'flag':0, 'pos':[0,2,4,5], 'range':range_list, 'error_pos_list':[]},
                         'A3':{'flag':0, 'pos':[0,1,2,3,4,5,6,7], 'range':range_list, 'error_pos_list':[]},
                         'A4':{'flag':0, 'pos':[0,1,2,3,4,5,6,7], 'range':range_list, 'error_pos_list':[]},
                         'A5':{'flag':0, 'pos':[0,1,2,3,4,5,6,7], 'range':range_list, 'error_pos_list':[]},
                         'A6':{'flag':0, 'pos':[0,1,2,3,4,5,6,7], 'range':range_list, 'error_pos_list':[]},
                         'B1':{'flag':0, 'pos':[0,2,4,5,8,9,10,11], 'range':range_list, 'error_pos_list':[]},
                         'B2':{'flag':0, 'pos':[0,2,4,5,8,9,10,11], 'range':range_list, 'error_pos_list':[]}}
            for event_type in mre_conf_dict:
                if re.search(event_type, gl.TEST_CONF['event']) != None:
                    event_deal_list.append(event_type)
            temp_mre_test_flag_dict = {'A1':0, 'A2':0, 'A3':0, 'A4':0, 'A5':0, 'A6':0, 'B1':0, 'B2':0}

            mre_file_name = gl.MR_DICT[time_entity][0]['MRE']
            mre_dom = gl.MR_DICT[time_entity][gl.MR_TYPE['MRE']]['xmldom']
            if mre_file_name == '' or mre_dom == None:
                continue
            mre_root = mre_dom.documentElement
            smr_list = mre_root.getElementsByTagName('smr')
            if len(smr_list) == 0:
                continue
            smr_str = smr_list[0].firstChild.data
            if len(smr_list) == 0 :
                temp_flag = 0
                for i in range(len(mre_smr_list)):
                    if mr_utils.is_mr_item_need_exist(mre_smr_list[i]) == True:
                        temp_flag = 1
                        break
                if temp_flag == 1:
                    for event_type in event_deal_list:
                        temp_mre_test_flag_dict[event_type] = 1
                    mr_utils.out_text_dict_append_list(out_mre_text, mre_file_name, 'smr err:' + 'no smr lable' if len(smr_list) == 0 else smr_list[0].firstChild.data)

            if re.search(smr_str, mre_smr_head) == None:
                mr_utils.get_mre_pos_list_by_mapping(mre_conf_dict, smr_str)
                #for event in mre_conf_dict:
                #    print (str(mre_conf_dict[event]['pos']))
            for object_entity in mre_root.getElementsByTagName('object'):
                test_excess_flag = 0
                for event_type in event_deal_list:
                    if event_type == object_entity.getAttribute('EventType'):
                        time_now_str = str(object_entity.getAttribute('TimeStamp'))
                        time_now_timestamp = mr_utils.get_timestamp_by_str_format(time_now_str)
                        if time_now_timestamp >= mre_event_begin_timestamp and time_now_timestamp <= mre_event_end_timestamp:
                            gl.test_mre_event_num_dict[event_type]['num'] += 1
                            gl.test_mre_event_num_dict[event_type]['time_str'].append(time_now_str)

                        test_excess_flag = 1
                        mre_conf_dict[event_type]['flag'] = 1
                        for value_entity in object_entity.getElementsByTagName('v'):
                            for pos in range(12):
                                if pos in mre_conf_dict[event_type]['pos']:
                                    value_num = 0
                                    if value_entity.firstChild.data.strip().split(' ')[pos].isdigit() == True:
                                        value_num = int(value_entity.firstChild.data.strip().split(' ')[pos])

                                    if value_num < mre_conf_dict[event_type]['range'][pos][0] or value_num > mre_conf_dict[event_type]['range'][pos][1]:
                                        temp_mre_test_flag_dict[event_type] += 1
                                        mr_utils.out_text_dict_append_list(out_mre_text, mre_file_name, 'value confusion: << {0} = {1} ->({2},{3} )>> event-[{4}] TimeStamp:{5}\n'.\
                                            format(mre_smr_list[pos],str(value_num),str(mre_conf_dict[event_type]['range'][pos][0]),
                                            str(mre_conf_dict[event_type]['range'][pos][1]), event_type, object_entity.getAttribute('TimeStamp')) )
                                        #print (event_type + '-' + object_entity.getAttribute('TimeStamp'))
                                        if pos not in mre_conf_dict[event_type]['error_pos_list']:
                                            mre_conf_dict[event_type]['error_pos_list'].append(pos)
                                else:
                                    if len(value_entity.firstChild.data.strip().split(' ')) <= pos:
                                        break
                                    if value_entity.firstChild.data.strip().split(' ')[pos] != 'NIL':
                                        temp_mre_test_flag_dict[event_type] += 1
                                        mr_utils.out_text_dict_append_list(out_mre_text, mre_file_name, 'event format confusion(not NIL):[{0}] event:[{1}}] TimeStamp:[{2}]'.format(mre_smr_list[pos], event_type, object_entity.getAttribute('TimeStamp')))
                                        if pos not in mre_conf_dict[event_type]['error_pos_list']:
                                            mre_conf_dict[event_type]['error_pos_list'].append(pos)

                if test_excess_flag == 0:
                    mr_utils.out_text_dict_append_list(out_mre_text, mre_file_name, 'suplus event:{0} - TimeStamp:{1}\n'.format(object_entity.getAttribute('EventType'), object_entity.getAttribute('TimeStamp')))
            with open(gl.OUT_PATH, 'a') as file_object:
                file_object.write('====>%s:\n'%(mre_file_name))
                file_object.write(date_time + " | ")
                file_object.write(gl.TEST_CONF['test_total_time'] + " | " )

                for event_type in out_mre_list:
                    file_object.write('True | ' if event_type in event_deal_list else 'False | ')

                file_object.write(date_time + " | ")
                file_object.write(gl.TEST_CONF['test_total_time'] + " | " )

                for event_type in out_mre_list:
                    if temp_mre_test_flag_dict[event_type] == 0:
                        file_object.write('Y | ' if event_type in event_deal_list else 'N | ' )

                    else:
                        file_object.write('[')
                        for pos in mre_conf_dict[event_type]['error_pos_list']:
                            file_object.write(mre_smr_list[pos] + ' ' )
                        file_object.write('] | ')

                file_object.write('\n\n')
        except Exception as result:
            raise Exception('-%s- [%s] <%s> %s'%(str(result.__traceback__.tb_lineno),gl.MR_DICT[time_entity][0]['MRE'], result, str(result.__traceback__.tb_lineno)))


    if int(gl.TEST_CONF['is_59_out_excel']) == 1:
        full_name = os.path.join(gl.SOURCE_PATH, gl.XLS_NAME)
        excel_workbook = openpyxl.load_workbook(full_name)
        work_sheet = excel_workbook.worksheets[3]

        for event_type in out_mre_list:
            work_sheet.cell(xml_out_dict[event_type]['event'], 9, 'Y' if event_type in event_deal_list else 'N')
            test_count = 0
            for i in range(xml_out_dict[event_type]['item'][0], xml_out_dict[event_type]['item'][1]+1):
                if mre_conf_dict[event_type]['pos'][test_count] not in mre_conf_dict[event_type]['error_pos_list'] and event_type in event_deal_list :
                    work_sheet.cell(i, 8, 'Y')
                else:
                    work_sheet.cell(i, 8, 'N')
                test_count += 1

        excel_workbook.save(filename=full_name)

    with open(gl.OUT_PATH, 'a') as file_object:
        if int(gl.TEST_CONF['is_59_out_excel']) == 1:
            file_object.write('结果已写入: ' + gl.XLS_NAME)
        for mre_file_name in out_mre_text:
            file_object.write('\n======>' + mre_file_name + ':\n')
            for list_entity in out_mre_text[mre_file_name]:
                file_object.write('\t' + list_entity)




def test61_file_accuracy():
    mr_utils.test_out_data_item_header("test_61")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)
    try:
        with open(gl.OUT_PATH, 'a') as file_object:
            file_object.write(date_time + " | ")
            file_object.write(gl.TEST_CONF['test_total_time'] + " | " )
    except Exception as result:
        raise Exception('%s'%(result))

def test62_file_accuracy():
    mr_utils.test_out_data_item_header("test_62")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)
    try:
        with open(gl.OUT_PATH, 'a') as file_object:
            file_object.write(date_time + " | ")
            file_object.write(gl.TEST_CONF['test_total_time'] + " | " )
    except Exception as result:
        raise Exception('%s'%(result))
def test63_file_accuracy():
    mr_utils.test_out_data_item_header("test_63")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)
    try:
        with open(gl.OUT_PATH, 'a') as file_object:
            file_object.write(date_time + " | ")
            file_object.write(gl.TEST_CONF['test_total_time'] + " | " )
    except Exception as result:
        raise Exception('%s'%(result))
def test71_file_accuracy():
    mr_utils.test_out_data_item_header("test_71")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)
    time_sub = 0 if gl.IS_SSH_CONN == 0 else mr_utils.get_two_linux_time_sub()
    test_flag = 0
    out_text_list = {}
    try:
        file_list = os.listdir(gl.MR_TEST_PATH)
        re_search_str = gl.TEST_CONF['standard_LTE'] + '_MR'

        for time_entity in gl.MR_DICT:
            # full_file = os.path.join(gl.MR_TEST_PATH, file)
            mr_file_name = [gl.MR_DICT[time_entity][0]['MRO'], gl.MR_DICT[time_entity][0]['MRE'], gl.MR_DICT[time_entity][0]['MRS']]
            mr_file_dom = [gl.MR_DICT[time_entity][gl.MR_TYPE['MRO']]['xmldom'], gl.MR_DICT[time_entity][gl.MR_TYPE['MRE']]['xmldom'], gl.MR_DICT[time_entity][gl.MR_TYPE['MRS']]['xmldom']]

            for i in range(len(mr_file_name)):
                if mr_utils.MR_xml_file_name_accuracy(mr_file_name[i]) == False :
                    test_flag |= (0x1 << 1)
                    mr_utils.out_text_dict_append_list(out_text_list, mr_file_name[i], 'file_name format err: ->  %s_MR*_%s_%s_%s_YYmmddHHMMSS.xml  ==> %s\n' %( gl.TEST_CONF['standard_LTE'] ,  gl.TEST_CONF['OEM'] , gl.MR_CONF['OmcName'] , gl.TEST_CONF['cellid'], mr_file_name[i]))
                mr_file_root = mr_file_dom[i].documentElement
                start_report_Time = mr_utils.get_timestamp_by_str_format(mr_file_root.getElementsByTagName('fileHeader')[0].getAttribute('reportTime'))
                start_report_Time /= 1000
                #TODO: get two linux

                start_report_Time -= time_sub
                filename = mr_file_name[i].strip('\\').split('\\')[len(mr_file_name[i].strip('\\').split('\\'))-1]
                if gl.MR_REMOTE_FILE_TIME_DIST.__contains__(filename) == False:
                    file_create_time = time.mktime(time.gmtime(os.path.getmtime(mr_file_name[i])))
                else:
                    file_create_time = gl.MR_REMOTE_FILE_TIME_DIST[filename]
                if file_create_time - start_report_Time > int(gl.TEST_CONF['file_delay_time'])*60:
                    test_flag |= (0x1 << 4)
                    mr_utils.out_text_dict_append_list(out_text_list, mr_file_name[i],
                      'file create time:[%s]-[%s] \n' % (str(time.strftime( '%Y-%m-%dT%H:%M:%S',time.localtime(file_create_time))), str(time.strftime( '%Y-%m-%dT%H:%M:%S',time.localtime(start_report_Time)))  ) )
            mro_ret_list = mr_utils.is_mro_correct(mr_file_dom[0])
            mre_ret_list = mr_utils.is_mre_correct(mr_file_dom[1])
            mrs_ret_list = mr_utils.is_mrs_correct(mr_file_dom[2])
            if mro_ret_list[0] == False:
                test_flag |= (0x1 << 2)
                mr_utils.out_text_dict_append_list(out_text_list, mr_file_name[0], 'file format  -> %s\n' % (mro_ret_list[1]))
            if mre_ret_list[0] == False:
                test_flag |= (0x1 << 2)
                mr_utils.out_text_dict_append_list(out_text_list, mr_file_name[1], 'file format  -> %s\n' % (mre_ret_list[1]))
            if mrs_ret_list[0] == False:
                test_flag |= (0x1 << 2)
                mr_utils.out_text_dict_append_list(out_text_list, mr_file_name[2],'file format  -> %s \n' %(mrs_ret_list[1] ))



        with open(gl.OUT_PATH, 'a') as file_object:
            file_object.write(date_time + " | ")
            file_object.write(gl.TEST_CONF['test_total_time'] + " | " )
            if test_flag & (0x1 << 1) != 0:
                file_object.write('N | ')
            else:
                file_object.write('Y | ')

            if test_flag & (0x1 << 2) != 0:
                file_object.write('N | ')
            else:
                file_object.write('Y | ')

            if test_flag & (0x1 << 4) != 0:
                file_object.write('[] | N | \n')
            else:
                file_object.write('[] | Y | \n')
            if len(out_text_list) != 0:
                file_object.write('error:\n')
            for file_name in out_text_list:
                file_object.write('======>' + file_name + ' :\n')
                for i in range(len(out_text_list[file_name])):
                    file_object.write('\t' + out_text_list[file_name][i])
    except Exception as result:
        # print (result)
        raise Exception('<%s> %s'%(str(result.__traceback__.tb_lineno),result))

def test72_file_accuracy():
    mr_utils.test_out_data_item_header("test_72")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    schema_dict = {"MRO":gl.SOURCE_PATH+"mro_schema.xsd", "MRE":gl.SOURCE_PATH+'mre_schema.xsd', "MRS":gl.SOURCE_PATH+'mrs_schema.xsd'}
    test_flag = 0
    out_text_dict = {}
    file_header_list_name = ['reportTime', 'startTime', 'endTime']

    for time_str in gl.MR_DICT:
        try:
            measureItem_list = {}
            mr_utils.get_measureItem_list(measureItem_list)
            for mr_type in gl.MR_DICT[time_str][0]:
                schema_doc  = etree.parse(schema_dict[mr_type])
                schema_ret = etree.XMLSchema(schema_doc)
                mr_file_full_name = gl.MR_DICT[time_str][0][mr_type]
                if mr_file_full_name == ''  :
                    if mr_utils.is_mr_item_need_exist(mr_type) == True:
                        raise Exception('%s lost %s'%(time_str, mr_type))
                    continue

                data = etree.parse(mr_file_full_name)
                if schema_ret.validate(data) == False:
                    test_flag |= (0x1 << 1)
                    mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name,schema_ret.error_log )

                if mr_type != 'MRS':
                    mr_doc = gl.MR_DICT[time_str][gl.MR_TYPE[mr_type]]['xmldom']
                    mr_root = mr_doc.documentElement
                    file_header_list = mr_root.getElementsByTagName('fileHeader')
                    enb_id = int(mr_root.getElementsByTagName('eNB')[0].getAttribute('id'))

                    for list_item_name in file_header_list_name:
                        if mr_utils.is_str_format_time(file_header_list[0].getAttribute(list_item_name), gl.TIME_FORMAT) == False:
                            mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, list_item_name + ' format error')
                            test_flag |= (0x1 << 2)
                    if file_header_list[0].getAttribute('period') != '0':
                        mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, 'upload peroiod error :[0] !=  ' + file_header_list[0].getAttribute('period'))
                        test_flag |= (0x1 << 2)

                    if int(file_header_list[0].getAttribute('jobid')) < 0 or int(file_header_list[0].getAttribute('jobid')) > 4294967295 :
                        mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, 'jobid :[0-4294967295] !=  ' + file_header_list[0].getAttribute('jobid'))
                        test_flag |= (0x1 << 2)

                    for measurement_entity in mr_root.getElementsByTagName('eNB')[0].getElementsByTagName('measurement'):
                        smr_value = measurement_entity.getElementsByTagName('smr')[0].firstChild.data
                        smr_value_list = smr_value.strip().split(' ')

                        for i in range(len(smr_value_list)):
                            temp_flag = 0
                            for mr_item in measureItem_list:
                                if smr_value_list[i] == mr_item:
                                    temp_flag = 1
                                    measureItem_list[mr_item] += 1
                                    break
                            if temp_flag == 0:
                                mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, 'surplus the mr_item : %s'%(smr_value_list[i]))
                        for object_entity in measurement_entity.getElementsByTagName('object'):
                            eci_id = int(object_entity.getAttribute('id').strip(':').split(':')[0])
                            enb_id = eci_id >> 8 & 0xff
                            cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)
                            if cell_id_ret_list[0] == False or mr_utils.is_enb_id_exist(enb_id) == False:
                                mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '[%s] <cell_id(%d) or enb_id(%d) not exist>'%(
                                    smr_value, cell_id_ret_list[1], enb_id))
                                test_flag |= (0x1 << 3)
                            if mr_utils.is_str_format_time(object_entity.getAttribute('TimeStamp'), gl.TIME_FORMAT) == False:
                                mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '[%s] <timestamp format error> %s'%(
                                   smr_value, object_entity.getAttribute('TimeStamp')))
                                test_flag |= (0x1 << 3)
                            if smr_value_list[0] == 'MR.LteScRIP':
                                if  mr_utils.is_eci_correct(int(object_entity.getAttribute('id').strip(':').split(':')[0])) == False or \
                                    re.search(object_entity.getAttribute('id').strip(':').split(':')[2], gl.MR_CONF['SubFrameNum']) == None or\
                                    object_entity.getAttribute('MmeUeS1apId') != 'NIL' or \
                                    object_entity.getAttribute('MmeGroupId') != 'NIL' or \
                                    object_entity.getAttribute('MmeCode') != 'NIL' :
                                        test_flag |= (0x1 << 4)
                                        mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, 'MR.LteScRIP:<rip object_format error> %s ' % (object_entity.getAttribute('TimeStamp')))
                            else:

                                if mr_utils.is_eci_correct(int(object_entity.getAttribute('id').strip(':').split(':')[0])) == False :
                                    mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '[L3] <eci_id not match> %s'% (object_entity.getAttribute('TimeStamp')))
                                    test_flag |= (0x1 << 3)
                            temp_value_flag = 0

                            for value_entity in object_entity.getElementsByTagName('v'):
                                value_str = value_entity.firstChild.data
                                if mr_utils.is_mr_value_correct(smr_value, value_str) == False:
                                    mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, 'value not match')
                                    test_flag |= (0x1 << 4)
                                    temp_value_flag = 1
                            if temp_value_flag == 0:
                                for mr_entity in measureItem_list:
                                    if mr_entity != 'MR.RSRP' or mr_entity != 'MR.RSRQ' or mr_entity != 'MR.ReceivedIPower' or mr_entity != 'MR.RIPPRB' or mr_entity != 'MR.PowerHeadRoom':
                                        measureItem_list[mr_entity] += 1

                else:
                    mr_doc = gl.MR_DICT[time_str][gl.MR_TYPE[mr_type]]['xmldom']
                    mr_root = mr_doc.documentElement
                    file_header_list = mr_root.getElementsByTagName('fileHeader')
                    #得到MRO的earfcn，判断MRS文件中的earfcn是否一致
                    mro_doc = xmldom.parse(gl.MR_DICT[time_str][0]['MRO'])
                    mro_root = mro_doc.documentElement
                    earfcn_value = mro_root.getElementsByTagName('measurement')[0].getElementsByTagName('object')[0].getElementsByTagName('v')[0].firstChild.data.strip().split(' ')[0] if \
                        len(mro_root.getElementsByTagName('measurement')) != 0 and len(mro_root.getElementsByTagName('measurement')[0].getElementsByTagName('object')) != 0 else '-1'

                    #获得子帧分帧
                    prbnum_list = []
                    if re.search(r',', gl.MR_CONF['PrbNum']) == None and '.' in gl.MR_CONF['PrbNum']:
                        for i in range(int(gl.MR_CONF['PrbNum'].strip('....').split('....')[0]), 1 + int(gl.MR_CONF['PrbNum'].strip('....').split('....')[1])):
                            prbnum_list.append(str(i))
                    else:
                        prbnum_list = gl.MR_CONF['PrbNum'].strip(',').split(',')
                    subfram_list = gl.MR_CONF['SubFrameNum'].strip(',').split(',')

                    for list_item_name in file_header_list_name:
                        if mr_utils.is_str_format_time(file_header_list[0].getAttribute(list_item_name), gl.TIME_FORMAT) == False:
                            mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '%s format error'%(list_item_name))
                            test_flag |= (0x1 << 2)

                    if file_header_list[0].getAttribute('period') != gl.MR_CONF['UploadPeriod']:
                        mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, 'upload peroiod error :[0] !=  ' + file_header_list[0].getAttribute('period'))
                        test_flag |= (0x1 << 2)

                    if int(file_header_list[0].getAttribute('jobid')) < 0 or int(file_header_list[0].getAttribute('jobid')) > 4294967295 :
                        mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, 'jobid :[0-4294967295] !=  ' + file_header_list[0].getAttribute('jobid'))
                        test_flag |= (0x1 << 2)

                    for measurement_entity in mr_root.getElementsByTagName('eNB')[0].getElementsByTagName('measurement'):
                        mr_name = measurement_entity.getAttribute('mrName')
                        if measureItem_list.__contains__(mr_name) == True:
                            measureItem_list[mr_name] += 1
                        else:
                            mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '<surplus mrItem> %s'%(mr_name))
                        smr_value = measurement_entity.getElementsByTagName('smr')[0].firstChild.data
                        object_list = measurement_entity.getElementsByTagName('object')

                        if mr_name == "MR.ReceivedIPower":
                            if len(subfram_list)*len(gl.TEST_CONF['cellid'].strip(',').split(',')) != len(object_list):
                                mr_utils.string_to_list(out_text_dict, mr_file_full_name, "MR.ReceivedIPower:object format error:[%s]" % (str(subfram_list[0:len(subfram_list)])))
                                test_flag |= (0x1 << 2)
                            for object_entity in object_list:
                                id_str = object_entity.getAttribute('id')
                                id_list = id_str.strip(':').split(':')
                                if mr_utils.is_eci_correct(int(id_list[0])) == False or (id_list[1] != earfcn_value and earfcn_value != '-1') or id_list[2] not in subfram_list:
                                    mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name,"MR.ReceivedIPower:object format error:[%s]"%(id_str))
                                    test_flag |= (0x1 << 4)

                        elif mr_name == 'MR.RIPPRB':
                            if len(subfram_list)*len(prbnum_list)*len(gl.TEST_CONF['cellid'].strip(',').split(',')) != len(object_list):
                                mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, "MR.RIPPRB:object format error:[%s * %s]" % (str(subfram_list[0:len(subfram_list)]) , str(prbnum_list[0:len(prbnum_list)])))
                                test_flag |= (0x1 << 4)
                            for object_entity in object_list:
                                id_str = object_entity.getAttribute('id')
                                id_list = id_str.strip(':').split(':')
                                if mr_utils.is_eci_correct(int(id_list[0])) == False or (id_list[1] != earfcn_value and earfcn_value != '-1') or id_list[2] not in subfram_list or id_list[2] not in prbnum_list:
                                    mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, "MR.RIPPRB:object format error:[%s]"%(id_str))
                                    test_flag |= (0x1 << 4)
                        else:
                            if len(object_list) != len(gl.TEST_CONF['cellid'].strip(',').split(',')):
                                mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '%s: object format error [num not match] - [%d]'%(mr_name,len(object_list)))
                                test_flag |= (0x1 << 4)
                            for object_entity in object_list:
                                id_str = object_entity.getAttribute('id')
                                id_list = id_str.strip(':').split(':')
                                if mr_utils.is_eci_correct(int(id_list[0])) == False:
                                    mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '%s <object id not match > %s'%(mr_name, id_str))
                                    test_flag |= (0x1 << 4)
                        temp_value_flag = 0

                        for object_entity in object_list:
                            for value_entity in object_entity.getElementsByTagName('v'):
                                value_str = value_entity.firstChild.data
                                if mr_utils.is_mrs_measurement_smr_value_correct(mr_name, smr_value, value_str) == False:
                                    mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '<smr or value not match> \n--->smr:%s \n---> value:%s'%(smr_value, value_str))
                                    test_flag |= (0x1 << 4)
                                    temp_value_flag = 1
                        if temp_value_flag == 0:
                            measureItem_list[mr_name] += 1

        except Exception as result:
            raise Exception('-%s- <%s> MRO:%s MRE:%s MRS:%s'%(str(result.__traceback__.tb_lineno),result, gl.MR_DICT[time_str][0]['MRO'], gl.MR_DICT[time_str][0]['MRE'], gl.MR_DICT[time_str][0]['MRS']))

    with open(gl.OUT_PATH, 'a') as file_object:
        file_object.write(date_time + " | ")
        file_object.write(gl.TEST_CONF['test_total_time'] + " | " )

        file_object.write('N | N | ' if test_flag & (0x1 << 1) != 0 else 'Y | Y | ')
        file_object.write('N | N | ' if test_flag & (0x1 << 2) != 0 else 'Y | Y | ' )
        file_object.write('N | ' if test_flag & (0x1 << 3) != 0 else 'Y | ')
        file_object.write('N | \n' if test_flag & (0x1 << 4) != 0 else 'Y | \n')
        if test_flag != 0:
            for file_name in out_text_dict:
                file_object.write('=====>' + file_name + ' :\n')
                for i in range(len(out_text_dict[file_name])):
                    file_object.write('\t' + str(out_text_dict[file_name][i]) + '\n')

def test73_file_accuracy():
    mr_utils.test_out_data_item_header("test_73")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    mro_file_list = glob.glob(gl.MR_TEST_PATH + '*MRO*.xml')
    max_mro_object_num = math.ceil( (int(gl.MR_CONF['UploadPeriod']) * 60 * 1000) / float(int(gl.MR_CONF['SamplePeriod'])))
    mro_file_flag_dict = {}
    out_text_dict = {}
    for time_entity in gl.MR_DICT:
        try:
            mro_file = gl.MR_DICT[time_entity][0]['MRO']
            mro_dom = gl.MR_DICT[time_entity][gl.MR_TYPE['MRO']]['xmldom']
            mro_root = mro_dom.documentElement
            time_stamp_tmep = 0
            if mro_file_flag_dict.__contains__(mro_file) == False:
                mro_file_flag_dict[mro_file] = {}
            for measurement_entity in mro_root.getElementsByTagName('measurement'):
                if measurement_entity.getElementsByTagName('smr')[0].firstChild.data.strip().split(' ')[0] == 'MR.LteScRIP' :
                    continue
                for object_entity in measurement_entity.getElementsByTagName('object'):
                    ue_mme_name = object_entity.getAttribute('MmeUeS1apId') + '|' + object_entity.getAttribute('MmeGroupId') + '|' + object_entity.getAttribute('MmeCode')
                    if mro_file_flag_dict[mro_file].__contains__(ue_mme_name) == False:
                        mro_file_flag_dict[mro_file][ue_mme_name] = {'mro_object_num':0, 'is_ascend':0}
                    if time_stamp_tmep != 0:
                        time_spec = mr_utils.get_timestamp_by_str_format(object_entity.getAttribute('TimeStamp')) - time_stamp_tmep
                        if time_spec != int(gl.MR_CONF['SamplePeriod']):
                            mro_file_flag_dict[mro_file][ue_mme_name]['is_ascend'] = 1
                            mr_utils.out_text_dict_append_list(out_text_dict, mro_file, "time spec: [%d]-[%d]=[%d] - [%d] "%(mr_utils.get_timestamp_by_str_format(object_entity.getAttribute('TimeStamp')), time_stamp_tmep, time_spec, int(gl.MR_CONF['SamplePeriod'])))
                    time_stamp_tmep = mr_utils.get_timestamp_by_str_format(object_entity.getAttribute('TimeStamp'))
                    mro_file_flag_dict[mro_file][ue_mme_name]['mro_object_num'] += 1
        except Exception as result:
            raise Exception('-%s- %s <%s>'%(str(result.__traceback__.tb_lineno),gl.MR_DICT[time_entity][0]['MRO'], result))

    with open(gl.OUT_PATH, 'a') as file_object:
        file_object.write(date_time + " | ")
        file_object.write(gl.TEST_CONF['test_total_time'] + " | " )
        file_object.write(str(max_mro_object_num) + ' | \n')
        for mro_file in mro_file_flag_dict:
            file_object.write('======>' + mro_file + ": \n\t")
            if len(mro_file_flag_dict[mro_file]) == 0:
                file_object.write("no L3 information\n")
                continue
            for ue_mme_name in mro_file_flag_dict[mro_file]:
                file_object.write( "[" + ue_mme_name + "]: [" + str(mro_file_flag_dict[mro_file][ue_mme_name]['mro_object_num']) + "]-["+ str(mro_file_flag_dict[mro_file][ue_mme_name]['is_ascend'] == 0) + "]  ")
                if mro_file_flag_dict[mro_file][ue_mme_name]['mro_object_num'] > max_mro_object_num or mro_file_flag_dict[mro_file][ue_mme_name]['is_ascend'] != 0:
                    file_object.write(' | N |\n')
                else:
                    file_object.write(' | Y |\n')

def test_add_timestamp_number():
    with open(gl.OUT_PATH, 'a') as file_object:
        file_object.write('\n<-------------------------->\n')
        file_object.write('test the timestamp [l2-l3]\nresult:\n')

    out_text_dict = {}
    out_list = ['l3','l2']
    smr_l3_str = 'MR.LteScEarfcn MR.LteScPci MR.LteScRSRP MR.LteScRSRQ MR.LteScPHR MR.LteScSinrUL MR.LteNcEarfcn MR.LteNcPci MR.LteNcRSRP MR.LteNcRSRQ'
    smr_rip_str = 'MR.LteScRIP'
    for time_entity in gl.MR_DICT:
        try:
            mro_file = gl.MR_DICT[time_entity][0]['MRO']
            mro_doc = gl.MR_DICT[time_entity][gl.MR_TYPE['MRO']]['xmldom']
            mro_root = mro_doc.documentElement
            timestamp_dict = {'l3': {}, 'l2':{}}
            l3_timestamp_temp = 0
            l2_timestamp_temp = 0

            for enb_entity in mro_root.getElementsByTagName('eNB'):
                start_time_stamp = mr_utils.get_timestamp_by_str_format(mro_root.getElementsByTagName('fileHeader')[0].getAttribute('startTime')) + int(gl.MR_CONF['SamplePeriod'])

                end_time_stamp = mr_utils.get_timestamp_by_str_format(mro_root.getElementsByTagName('fileHeader')[0].getAttribute('endTime'))
                measurement_list = enb_entity.getElementsByTagName('measurement')
                if len(measurement_list) != 2 :
                    continue
                l3_smr_list = measurement_list[0].getElementsByTagName('smr')
                l3_object_list = measurement_list[0].getElementsByTagName('object')
                l2_smr_list = measurement_list[1].getElementsByTagName('smr')
                l2_object_list = measurement_list[1].getElementsByTagName('object')
                if len(l3_smr_list) != 1 or len(l3_object_list) == 0 or len(l2_smr_list) != 1 or len(l2_object_list) == 0:
                    continue
                if re.search(smr_l3_str, l3_smr_list[0].firstChild.data) == None:
                    continue
                if re.search(smr_rip_str, l2_smr_list[0].firstChild.data) == None:
                    continue

                for l3_object_entity in l3_object_list:
                    cell_id_ret_list = mr_utils.is_cell_id_exist(int(l3_object_entity.getAttribute('id').strip(':').split(':')[0]))
                    cell_id = str(cell_id_ret_list[1])
                    if timestamp_dict['l3'].__contains__(cell_id) == False:
                        timestamp_dict['l3'][cell_id] = {'num':0, 'time_str':[], 'time_stamp':[]}

                    time_str = l3_object_entity.getAttribute('TimeStamp')
                    time_stamp = mr_utils.get_timestamp_by_str_format(time_str)

                    if l3_timestamp_temp != time_stamp:
                        timestamp_dict['l3'][cell_id]['num'] += 1
                        l3_timestamp_temp = time_stamp
                        timestamp_dict['l3'][cell_id]['time_str'].append(time_str)
                        timestamp_dict['l3'][cell_id]['time_stamp'].append(time_stamp)
                for l2_object_entity in l2_object_list:
                    cell_id_ret_list = mr_utils.is_cell_id_exist(int(l2_object_entity.getAttribute('id').strip(':').split(':')[0]))
                    cell_id = str(cell_id_ret_list[1])
                    if timestamp_dict['l2'].__contains__(cell_id) == False:
                        timestamp_dict['l2'][cell_id] = {'num':0, 'time_str':[], 'time_stamp':[]}
                    time_str = l2_object_entity.getAttribute('TimeStamp')
                    time_stamp = mr_utils.get_timestamp_by_str_format(time_str)

                    if l2_timestamp_temp != time_stamp:
                        timestamp_dict['l2'][cell_id]['num'] += 1
                        l2_timestamp_temp = time_stamp
                        timestamp_dict['l2'][cell_id]['time_str'].append(time_str)
                        timestamp_dict['l2'][cell_id]['time_stamp'].append(time_stamp)

                for cell_id in timestamp_dict['l2'] if len(timestamp_dict['l2']) >= len(timestamp_dict['l3']) else timestamp_dict['l3']:
                    if timestamp_dict['l2'].__contains__(cell_id) == True and timestamp_dict['l3'].__contains__(cell_id) == True:
                        if timestamp_dict['l2'][cell_id]['num'] != timestamp_dict['l3'][cell_id]['num']:
                            mr_utils.out_text_dict_append_list(out_text_dict, mro_file, '[cellid:%s] object_num : [l3]-%d != [l2]-%d \n'%(cell_id, timestamp_dict['l3'][cell_id]['num'], timestamp_dict['l2'][cell_id]['num']))

                for l_entity in out_list:
                    for cell_id in timestamp_dict[l_entity]:
                        if timestamp_dict[l_entity][cell_id]['time_stamp'][0] != start_time_stamp:
                            mr_utils.out_text_dict_append_list(out_text_dict, mro_file, '[%s] <%s startTimeStamp not match>: %s(start) != %s(first) \n'%(cell_id,l_entity,
                                mr_utils.get_time_format_by_timestamp(start_time_stamp),
                                timestamp_dict[l_entity][cell_id]['time_str'][0]))
                        if timestamp_dict[l_entity][cell_id]['time_stamp'][len(timestamp_dict[l_entity][cell_id]['time_stamp']) - 1] != end_time_stamp:
                            mr_utils.out_text_dict_append_list(out_text_dict, mro_file, '[%s] <%s endTimeStamp not match>: %s(end) != %s(last) \n'%(cell_id,l_entity,
                                mro_root.getElementsByTagName('fileHeader')[0].getAttribute('endTime'),
                                timestamp_dict[l_entity][cell_id]['time_str'][len(timestamp_dict[l_entity][cell_id]['time_stamp']) - 1]))
                for cell_id in timestamp_dict['l2'] if len(timestamp_dict['l2']) >= len(timestamp_dict['l3']) else timestamp_dict['l3']:
                    if timestamp_dict['l2'].__contains__(cell_id) == True and timestamp_dict['l3'].__contains__(cell_id) == True:
                        l3_index = 0
                        l2_index = 0
                        suplus_temp = 0
                        for i in range(0, timestamp_dict['l3'][cell_id]['num'] if timestamp_dict['l3'][cell_id]['num'] >= timestamp_dict['l2'][cell_id]['num'] else timestamp_dict['l2'][cell_id]['num'] ):
                            if l2_index > timestamp_dict['l2'][cell_id]['num']-1:
                                l2_index = timestamp_dict['l2'][cell_id]['num']-1
                                suplus_temp += 1
                            if l3_index > timestamp_dict['l3'][cell_id]['num']-1:
                                l3_index = timestamp_dict['l3'][cell_id]['num']-1
                                suplus_temp += 1
                            if suplus_temp > 3:
                                if l3_index == timestamp_dict['l3'][cell_id]['num']-1 and l2_index != timestamp_dict['l2'][cell_id]['num']-1:
                                    mr_utils.out_text_dict_append_list(out_text_dict, mro_file, '[{4}]<l3 lost part of data>: num:[l3={0}]--[l2={1}] TimeStamp:[l3={2}--l2={3}]'.format(timestamp_dict['l3'][cell_id]['num']-1,
                                                                                timestamp_dict['l2'][cell_id]['num']-1, timestamp_dict['l3'][cell_id]['time_str'][l3_index], timestamp_dict['l2'][cell_id]['time_str'][len(timestamp_dict['l2'][cell_id]['time_str']) - 1], cell_id))
                                elif l3_index != timestamp_dict['l3'][cell_id]['num']-1 and l2_index == timestamp_dict['l2'][cell_id]['num']-1:
                                    mr_utils.out_text_dict_append_list(out_text_dict, mro_file,'[{4}]<l2 lost part of data: num>:[l3={0}]--[l2={1}] TimeStamp:[l3={2}--l2={3}]'.format(timestamp_dict['l3'][cell_id]['num']-1,
                                                                                timestamp_dict['l2'][cell_id]['num']-1, timestamp_dict['l3'][cell_id]['time_str'][len(timestamp_dict['l3'][cell_id]['time_str']) - 1], timestamp_dict['l2'][cell_id]['time_str'][l2_index], cell_id))
                                break

                            if timestamp_dict['l3'][cell_id]['time_stamp'][l3_index] == timestamp_dict['l2'][cell_id]['time_stamp'][l2_index]:
                                l2_index += 1
                                l3_index += 1
                                suplus_temp = 0
                            elif timestamp_dict['l3'][cell_id]['time_stamp'][l3_index] < timestamp_dict['l2'][cell_id]['time_stamp'][l2_index]:
                                mr_utils.out_text_dict_append_list(out_text_dict, mro_file, '[%s] <timestamp not match>: [l3]-[%d]-%s < [l2]-[%d]-%s \n'%(cell_id,l3_index,timestamp_dict['l3'][cell_id]['time_str'][l3_index], l2_index,timestamp_dict['l2'][cell_id]['time_str'][l2_index] ))
                                l3_index += 1
                                suplus_temp += 1
                            elif timestamp_dict['l3'][cell_id]['time_stamp'][l3_index] > timestamp_dict['l2'][cell_id]['time_stamp'][l2_index]:
                                mr_utils.out_text_dict_append_list(out_text_dict, mro_file, '[%s] <timestamp not match>: [l3]-[%d]-%s > [l2]-[%d]-%s \n'%(cell_id,l3_index,timestamp_dict['l3'][cell_id]['time_str'][l3_index], l2_index,timestamp_dict['l2'][cell_id]['time_str'][l2_index] ))
                                l2_index += 1
                                suplus_temp += 1
        except Exception as result:
            raise Exception('-%s- %s <%s>'%(str(result.__traceback__.tb_lineno),gl.MR_DICT[time_entity][0]['MRO'], result))

    with open(gl.OUT_PATH, 'a') as file_object:
        file_object.write('test ok\n')
        if len(out_text_dict) != 0:
            file_object.write('error:\n')
        for mro_file_name in out_text_dict:
            file_object.write('=====>  ' + mro_file_name + ':\n')
            for i in range(0, len(out_text_dict[mro_file_name])):
                file_object.write('\t' + out_text_dict[mro_file_name][i] )

def test_add_mro_s_mapping():

    #范围
    rsrp_range_list = []
    rsrq_range_list = []
    phr_range_list = []
    sinrul_range_list = []
    rip_range_list = []

    rsrp_range_list.append(20)
    for i in range(25, 60):
        rsrp_range_list.append(i)
    for i in range(60, 81, 2):
        rsrp_range_list.append(i)
    rsrp_range_list.append(97)
    for i in range(0, 35, 2):
        rsrq_range_list.append(i)
    for i in range(64):
        phr_range_list.append(i)
    for i in range(0, 37):
        sinrul_range_list.append(i)

    for i in range(0, 521, 10):
        rip_range_list.append(i)

    out_mro_list = ['MR.LteScRSRP','MR.LteScRSRQ','MR.LteScPHR','MR.LteScSinrUL']
    out_text_list = {}
    for time_entity in gl.MR_DICT:


        cellid_item_dict = {'MR.LteScRSRP':{'pos':2, 'TimeStamp':0, 'range':[0, 97], 'flag':3, 'num':0}, 'MR.LteScRSRQ':{'pos':3, 'TimeStamp':0, 'range':[0,34],'flag':3, 'num':0},
                               'MR.LteScPHR':{'pos':4, 'TimeStamp':0, 'range':[0,63],'flag':3, 'num':0}, 'MR.LteScSinrUL':{'pos':5, 'TimeStamp':0, 'range':[0,36],'flag':3, 'num':0},
                               'MR.LteScRIP':{'pos':0, 'TimeStamp':0, 'range':[0,511], 'flag':3,'num':0, 'prbnum':{}}, 'MR.LteNcRSRP':{'pos':8, 'TimeStamp':0, 'range':[0,97],'flag':3, 'num':0},
                            'MR.LteNcRSRQ':{'pos':9, 'TimeStamp':0, 'range':[0,34], 'flag':3, 'num':0 } }
        item_range_list = {'MR.LteScRSRP':rsrp_range_list,
                           'MR.LteScRSRQ':rsrq_range_list,
                           'MR.LteScPHR':phr_range_list,
                           'MR.LteScSinrUL':sinrul_range_list,
                           'MR.LteScRIP':rip_range_list}
        try:

            temp_loop_flag = 0
            is_object_empty = 0
            mro_item_dict = {}
            temp_test_flag_dict = {'MR.LteScRSRP':0,'MR.LteScRSRQ':0,'MR.LteScPHR':0,'MR.LteScSinrUL':0,'MR.LteScRIP':0, 'MR.LteNcRSRP':0, 'MR.LteNcRSRQ':0}
            mro_measurement_list = []
            mrs_measurement_list = []
            mro_file_name = ""
            mrs_file_name = ""
            if re.search(r'MRO', gl.MR_CONF['MeasureType']) != None:
                mro_file_name = gl.MR_DICT[time_entity][0]['MRO']
                mro_dom = gl.MR_DICT[time_entity][gl.MR_TYPE['MRO']]['xmldom']
                mro_root = mro_dom.documentElement
                mro_measurement_list = mro_root.getElementsByTagName('measurement')
            if re.search(r'MRS', gl.MR_CONF['MeasureType']) != None:
                mrs_file_name = gl.MR_DICT[time_entity][0]['MRS']
                mrs_dom = gl.MR_DICT[time_entity][gl.MR_TYPE['MRS']]['xmldom']
                mrs_root = mrs_dom.documentElement
                mrs_measurement_list = mrs_root.getElementsByTagName('measurement')

            for measurement_entity in mro_measurement_list:
                if len(measurement_entity.getElementsByTagName('object')) == 0:
                    is_object_empty = 1
                    continue
                for smr_entity in measurement_entity.getElementsByTagName('smr'):
                    for object_entity in measurement_entity.getElementsByTagName('object'):
                        object_ue_id = object_entity.getAttribute('id').strip(':').split(':')[0]
                        if mro_item_dict.__contains__(object_ue_id) == False :
                            mro_item_dict[object_ue_id] = {'MR.LteScRSRP':{'pos':2, 'mrs_item':'MR.RSRP', 'range':[0, 97], 'flag':3, 'num':0, 'item_num':{}}, 'MR.LteScRSRQ':{'pos':3, 'mrs_item':'MR.RSRQ', 'range':[0,34],'flag':3, 'num':0, 'item_num':{}},
                               'MR.LteScPHR':{'pos':4, 'mrs_item':'MR.PowerHeadRoom', 'range':[0,63],'flag':3, 'num':0, 'item_num':{}}, 'MR.LteScSinrUL':{'pos':5, 'mrs_item':'MR.SinrUL', 'range':[0,36],'flag':3, 'num':0, 'item_num':{}},
                               'MR.LteScRIP':{'pos':0, 'mrs_item':'MR.ReceivedIPower', 'range':[0,511], 'flag':3,'num':0, 'prbnum':{'item_num':{}}}, 'MR.LteNcRSRP':{'pos':8, 'mrs_item':0, 'range':[0,97],'flag':3, 'num':0, 'item_num':{}},
                            'MR.LteNcRSRQ':{'pos':9, 'mrs_item':0, 'range':[0,34], 'flag':3, 'num':0, 'item_num':{} } }
                            mr_utils.get_mro_pos_list_by_mapping(mro_item_dict[object_ue_id], smr_entity.firstChild.data)
                        for mr_name_entity in mro_item_dict[object_ue_id]:
                            #mr_Name匹配上, flag-1
                            if re.search(mr_name_entity, smr_entity.firstChild.data) == None or mro_item_dict[object_ue_id][mr_name_entity]['pos'] == -1:
                                continue

                            if mr_name_entity == smr_entity.firstChild.data.strip().split(' ')[mro_item_dict[object_ue_id][mr_name_entity]['pos']]:
                                if mr_name_entity != 'MR.LteScRIP' and mr_name_entity != 'MR.LteNcRSRP' and mr_name_entity != 'MR.LteNcRSRQ':
                                    #TODO: 得到value, 然后对应的加上
                                    if len(object_entity.getElementsByTagName('v')) != 0:
                                        value_str = object_entity.getElementsByTagName('v')[0].firstChild.data.strip().split(' ')[mro_item_dict[object_ue_id][mr_name_entity]['pos']]
                                        if value_str.isdigit() == True:
                                            value = int(value_str)
                                            # print (mr_name_entity)
                                            # print (item_range_list[mr_name_entity])
                                            # print ((len(item_range_list[mr_name_entity])))
                                            for idx in range(len(item_range_list[mr_name_entity])):
                                                if mro_item_dict[object_ue_id][mr_name_entity].__contains__('item_num') == False:
                                                    mro_item_dict[object_ue_id][mr_name_entity]['item_num'] = {}
                                                if mro_item_dict[object_ue_id][mr_name_entity]['item_num'].__contains__(idx) == False:
                                                    mro_item_dict[object_ue_id][mr_name_entity]['item_num'][idx] = 0
                                                if value <= item_range_list[mr_name_entity][idx]:
                                                    mro_item_dict[object_ue_id][mr_name_entity]['item_num'][idx] += 1
                                                    break
                                elif mr_name_entity == 'MR.LteScRIP':
                                    prbnum = object_entity.getAttribute('id').strip(':').split(':')[0] + ':' + object_entity.getAttribute('id').strip(':').split(':')[2]
                                    if mro_item_dict[object_ue_id][mr_name_entity]['prbnum'].__contains__(prbnum) == False:
                                        mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum] = {'num':0, 'TimeStamp':0, 'item_num':{}}
                                    #TODO:获得value, 加上
                                    if len(object_entity.getElementsByTagName('v')) != 0:
                                        value_str = object_entity.getElementsByTagName('v')[0].firstChild.data.strip().split(' ')[mro_item_dict[object_ue_id][mr_name_entity]['pos']]
                                        if value_str.isdigit() == True:
                                            value = int(value_str)
                                            for idx in range(len(item_range_list[mr_name_entity])):
                                                if mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum].__contains__('item_num') == False:
                                                    mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['item_num'] = {}
                                                if mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['item_num'].__contains__(idx) == False:
                                                    mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['item_num'][idx] = 0
                                                if value <= item_range_list[mr_name_entity][idx]:
                                                    mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['item_num'][idx] += 1
                                                    break

            for measurement_entity in mrs_measurement_list:
                for object_ue_id in mro_item_dict:
                    for mr_name in mro_item_dict[object_ue_id]:
                        if mro_item_dict[object_ue_id][mr_name]['mrs_item'] == measurement_entity.getAttribute('mrName'):
                            if mr_name != 'MR.LteScRIP':
                                for object_entity in measurement_entity.getElementsByTagName('object'):
                                    if object_entity.getAttribute('id') == object_ue_id :
                                        value_entity_list = object_entity.getElementsByTagName('v')[0].firstChild.data.strip().split(' ')
                                        for i in range(len(item_range_list[mr_name])):
                                            if (int(value_entity_list[i]) != 0 and mro_item_dict[object_ue_id][mr_name]['item_num'].__contains__(i) == False) or \
                                                    (mro_item_dict[object_ue_id][mr_name]['item_num'].__contains__(i) == True and \
                                                     int(value_entity_list[i]) != mro_item_dict[object_ue_id][mr_name]['item_num'][i]) :
                                                mr_utils.out_text_dict_append_list(out_text_list, mrs_file_name, '[%s]-[%d]=[%d(mro) != %d(mrs)]'%(mr_name, i,
                                                    mro_item_dict[object_ue_id][mr_name]['item_num'][i] if mro_item_dict[object_ue_id][mr_name]['item_num'] else 0,\
                                                    int(value_entity_list[i])))
                            else:
                                for object_entity in measurement_entity.getElementsByTagName('object'):
                                    prbnum = object_entity.getAttribute('id').strip(':').split(':')[0] + ':' + object_entity.getAttribute('id').strip(':').split(':')[2]
                                    if object_entity.getAttribute('id').strip(':').split(':')[0] == object_ue_id :
                                        for prb_entity in mro_item_dict[object_ue_id][mr_name]['prbnum']:
                                            if prb_entity == prbnum:
                                                value_entity_list = object_entity.getElementsByTagName('v')[0].firstChild.data.strip().split(' ')
                                                for i in range(len(item_range_list[mr_name])):
                                                    if (int(value_entity_list[i]) != 0 and mro_item_dict[object_ue_id][mr_name]['prbnum'][prb_entity]['item_num'].__contains__(i) == False) or \
                                                        (mro_item_dict[object_ue_id][mr_name]['prbnum'][prb_entity]['item_num'].__contains__(i) == True and int(value_entity_list[i]) != mro_item_dict[object_ue_id][mr_name]['prbnum'][prb_entity]['item_num'][i]):
                                                        mr_utils.out_text_dict_append_list(out_text_list, mrs_file_name, '[%s]-[prb:%s id:%d]=[%d(mro) != %d(mrs)]'%(mr_name, prb_entity, i,
                                                            mro_item_dict[object_ue_id][mr_name]['prbnum'][prb_entity]['item_num'][i] if mro_item_dict[object_ue_id][mr_name]['prbnum'][prb_entity]['item_num'].__contains__(i) == True else 0, int(value_entity_list[i])))
        except Exception as result:
            raise Exception('-%s- <%s> MRO:%s MRS:%s'%(str(result.__traceback__.tb_lineno),result, gl.MR_DICT[time_entity][0]['MRO'], gl.MR_DICT[time_entity][0]['MRS']))
    with open(gl.OUT_PATH, 'a') as file_object:
        file_object.write('\n [' + ' add mro-mrs num mapping ' + ']:\n')
        temp_flag = 0
        for mrs_file_name in out_text_list:
            temp_flag = 1
            file_object.write('===>' + mrs_file_name + ' :\n')
            for i in range(len(out_text_list[mrs_file_name])):
                file_object.write('\t' + out_text_list[mrs_file_name][i] + '\n')
        if temp_flag == 0:
            file_object.write('test ok')

def test_add_mre_event_num():
    set_event_str = ('A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'B1', 'B2')
    with open(gl.OUT_PATH, 'a') as file_object:
        file_object.write('\n [' + ' add mre event total num ' + ']:\n')
        file_object.write('==> time: %s - %s :\n'%(gl.TEST_CONF['mre_begin_time'], gl.TEST_CONF['mre_end_time']))
        for event_str in set_event_str:
            file_object.write(' '*4 + event_str + ' : ' + str(gl.test_mre_event_num_dict[event_str]['num']) + '\n')
        for event_str in set_event_str:
            file_object.write('====> time_str: ')
            file_object.write(event_str + ' : ' + str(gl.test_mre_event_num_dict[event_str]['time_str']) + '\n')


def mr_test_process():

    mr_utils.conf_xml_parse()

    mr_utils.MR_xml_init()

    mr_utils.mr_out_file_data_head()

    mr_utils.mr_function_process(test51_file_integrity, 'test51')
    mr_utils.mr_function_process(test52_file_integrity, 'test52')
    mr_utils.mr_function_process(test53_file_integrity, 'test53')
    mr_utils.mr_function_process(test54_file_integrity, 'test54')
    mr_utils.mr_function_process(test55_file_integrity, 'test55')
    mr_utils.mr_function_process(test56_file_integrity, 'test56')
    mr_utils.mr_function_process(test57_file_integrity, 'test57')
    mr_utils.mr_function_process(test58_file_integrity, 'test58')
    mr_utils.mr_function_process(test59_file_integrity, 'test59')
    mr_utils.mr_function_process(test71_file_accuracy, 'test71')
    mr_utils.mr_function_process(test72_file_accuracy, 'test72')
    mr_utils.mr_function_process(test73_file_accuracy, 'test73')
    mr_utils.mr_function_process(test_add_timestamp_number, 'test_add_timestamp')
    mr_utils.mr_function_process(test_add_mro_s_mapping, 'test_add_mro_s_value')
    mr_utils.mr_function_process(test_add_mre_event_num, 'test_add_mre_event_num')

    if int(gl.TEST_CONF['is_57_out_excel']) == 1 or int(gl.TEST_CONF['is_58_out_excel']) == 1 or int(gl.TEST_CONF['is_59_out_excel']) == 1:
        src_name = os.path.join(gl.SOURCE_PATH, gl.XLS_NAME)
        dst_name = os.path.join(gl.OUTPUT_PATH, gl.XLS_NAME)
        shutil.copy(src_name, dst_name )


def pm_log_read_process():
    try:
        ssh = mr_utils.ssh_tool("10.110.38.222", 22, "root", "Bingo1993")
        result1 = ssh.cmd_run("cat /var/log/cspl/log_enodeb.dat | grep -E 'oam_mt_pm_report_meas_data.*is upload pm file:<yes>' | awk '{print $1,$2,$18}'")
        pm_log_list = []
        mr_utils.conf_xml_parse()
        #TODO: 在mr哪里，写xml文件的时候，其实就是global里面，没有添加pm，就会导致出错
        time_spec = int(gl.PM_CONF['PeriodicUploadTime'].split(':')[1]) * 60 % int(gl.PM_CONF['PeriodicUploadInterval'])
        for str1 in result1.decode('utf-8').splitlines():
            print (str1)
            pm_log_dict = {}
            # pm_log_dict['str1'] = str1
            pm_log_dict['time_stamp'] = int (mr_utils.get_timestamp_by_str_format(str1.split('[')[0].strip(), format="%Y.%m.%d %H:%M:%S.%f") / 1000)
            pm_log_dict['filename'] = str2 = str1.split('[')[1].split(']')[0]

            gen_time_str = str2.split('_')[2].split('.')[0] + str2.split('_')[2].split('.')[1].split('-')[1].split('+')[0] + '00'
            pm_log_dict['gen_time'] = int (mr_utils.get_timestamp_by_str_format(gen_time_str, format="%Y%m%d%H%M%S") / 1000)
            pm_log_dict['result'] = 'Y' if pm_log_dict['time_stamp'] - pm_log_dict['gen_time'] == time_spec else 'N'
            pm_log_list.append(pm_log_dict)
        # print (pm_log_list)

        with open(gl.OUTPUT_PATH+'data.txt', 'w') as file_object:
            for i in range(len(pm_log_list)):
                file_object.write(pm_log_list[i]['filename'] + '-->  file report offset_time:  ' + pm_log_list[i]['result'] + '\n')


        del ssh
    except Exception as rt:
        print ("%s %d"%(rt, rt.__traceback__.tb_lineno))
def pm_report_offset_process():
    #这一步是通过远程ssh访问服务器，pm上报到两个服务器上面，首先获取两个服务器的时间差。然后是获取基站的服务器的文件名，到远端服务器去找对应的文件，判断时间差值是否是一致的。
    ssh1 = mr_utils.ssh_tool("10.110.38.222", 22, "root", "Bingo1993")
    ssh2 = mr_utils.ssh_tool("10.110.38.214", 22, "root", "Bingo1993")
    # time1 =
    pass

def pm_test_process():
    try:
        pm_log_read_process()
    except Exception as rt:
        print ("%s %d"%(rt, rt.__traceback__.tb_lineno))





