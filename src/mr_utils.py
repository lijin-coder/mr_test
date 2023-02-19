
# encoding: utf-8


import os
import mr_global as gl
import xml.dom.minidom as xmldom
import sys
import re
import time
import datetime
import io
import glob
import csv
from openpyxl import Workbook
# import paramiko

write_info = lambda info:gl.str_info.append(str(info)) 



#======================================================
#    initial 
#======================================================

# function: parse conf.xml, 
#    TEST_CONF ->gl.TEST_CONF : mr test program config
#    MR_CONF ->gl.MR_CONF : refrence in /LTE/user/Area0/cfg/lte_oam_adp/oam_adp_mt_test_config.xml 
#    PM_CONF->gl.PM_CONF :refrence in /LTE/user/Area0/cfg/lte_oam_adp/oam_adp_mt_test_config.xml 
#    TEST_OUT->gl.TEST_OUT : refrence in LTE数字蜂窝移动通信网无线操作维护中心（OMC-R）测量报告测试要求-扩展型皮基站分册-V2.1.0.docx
#    TEST_ITEM_LIST->gl.TEST_ITEM_LIST : same as front
#    Note: when conf.xml changed , then gl.xx_CONF must be changed immediately, or run error here

def conf_xml_parse():
    full_file_name = os.path.join(gl.TEST_PATH, "conf.xml")

    if os.path.exists(full_file_name):
        conf_dom = xmldom.parse(full_file_name)
        conf_root = conf_dom.documentElement

        #解析TEST_CONF, 就是手动输入测试相关的信息,包括测试的总时间, 小区, 邻区id等
        test_conf_entity = conf_root.getElementsByTagName("TEST_CONF")[0]
        for test_conf_dict_key_entity in gl.TEST_CONF:
            gl.TEST_CONF[test_conf_dict_key_entity] = str(test_conf_entity.getElementsByTagName(test_conf_dict_key_entity)[0].firstChild.data)
            #print (test_conf_dict_key_entity + TEST_CONF[test_conf_dict_key_entity])

        #解析MR_CONF, 这个是下发的MR的测试命令
        mr_conf_entity = conf_root.getElementsByTagName("MR_CONF")[0]
        for mr_conf_dict_key_entity in gl.MR_CONF:
            gl.MR_CONF[mr_conf_dict_key_entity] = mr_conf_entity.getElementsByTagName(mr_conf_dict_key_entity)[0].firstChild.data
            #print( mr_conf_dict_key_entity + " - " + MR_CONF[mr_conf_dict_key_entity])

        #解析PM_CONF， 这个是PM的下发命令
        pm_conf_entity = conf_root.getElementsByTagName('PM_CONF')[0]
        for pm_conf_dict_key_entity in gl.PM_CONF:
            gl.PM_CONF[pm_conf_dict_key_entity] = pm_conf_entity.getElementsByTagName(pm_conf_dict_key_entity)[0].firstChild.data
            # print (pm_conf_dict_key_entity + '-' + gl.PM_CONF[pm_conf_dict_key_entity])

        #解析TEST_OUT, 这个是最后输出的项 分别对应测试文档中的每个项
        test_out_entity = conf_root.getElementsByTagName("TEST_OUT")[0]
        for test_out_dict_key_entity in gl.TEST_OUT:
            tmp_list = []
            tmp_list.append( int(test_out_entity.getElementsByTagName(test_out_dict_key_entity)[0].getAttribute("item_num") ) )
            for i in range(tmp_list[0]):
                tmp_list.append(int(test_out_entity.getElementsByTagName(test_out_dict_key_entity)[0].getAttribute("item" + '%d'%(i+1)) ))
            gl.TEST_OUT[test_out_dict_key_entity] = tmp_list
            # print (TEST_OUT[test_out_dict_key_entity])

        #解析ITEM_NAME, 这个是输出项需要对应的测试项,也就是测试文档中每个测试表中的列
        item_name_entity = conf_root.getElementsByTagName("ITEM_NAME")[0]
        item_entity_list = item_name_entity.getElementsByTagName("item")
        gl.TEST_ITEM_LIST.append(len(item_entity_list))
        for i in range(len(item_entity_list)):
            gl.TEST_ITEM_LIST.append(item_entity_list[i].firstChild.data )
            #print (TEST_ITEM_LIST[i+1])
    else:
        create_conf_xml(full_file_name)
        raise Exception('<no conf.xml file> now create %s \npress button <查看> to modify the conf.xml.'%(full_file_name))




#get omc Name, when omc is OMC-R_1_2048, file_list[3] = OMC-R, file_list[4] = 1 ...
def get_filename_omc_name(file_list):
    omcName = file_list[3]
    if  len(file_list) > 6:
        for i in range(4, len(file_list) - 2):
            omcName += '_' + file_list[i]
    return omcName

# the old function
def MR_xml_file_name_accuracy(file_name):
    # filename = file_name.split(gl.file_split_code)[-1]
    file_name_split_list = get_split_value(file_name, '_')
    file_name_split_num = len(file_name_split_list)
    
    if gl.TEST_CONF['standard_LTE'] != file_name_split_list[0]:
        return False
    if re.search(file_name_split_list[1], gl.MR_CONF['MeasureType']) == None:
        return False
    if gl.TEST_CONF['OEM'] != file_name_split_list[2]:
        return False
    if re.search(get_filename_omc_name(file_name_split_list), gl.MR_CONF['OmcName']) == None:
        return False
    if re.search(file_name_split_list[file_name_split_num - 2], gl.TEST_CONF['enbid']) == None:
        return False
    time_stamp_str = file_name_split_list[file_name_split_num - 1].split('.')[0]
    time_array = time.strptime(time_stamp_str, "%Y%m%d%H%M%S")
    timestamp = int(time.mktime(time_array))
    if timestamp % (60) != 0:
        return False
    return True      


# new , check MR name accuracy from MR_dict
def MR_xml_file_name_check(file_dict):
    temp_dict = {'standard_LTE' : gl.TEST_CONF['standard_LTE'], 'type':gl.MR_CONF['MeasureType'], 
                 'OEM':gl.TEST_CONF['OEM'],  'OmcName':gl.MR_CONF['OmcName'], 'enbid':gl.TEST_CONF['enbid']  }
    for item in temp_dict:
        if file_dict[item] not in temp_dict[item]:
            return False, item
    
    time_stamp_str = file_dict['SampleBeginTime']
    time_array = time.strptime(time_stamp_str, "%Y%m%d%H%M%S")
    timestamp = int(time.mktime(time_array))
    if timestamp % (60) != 0:
        return False, 'timestamp'
    return True,'' 
        
#MR文件xml的初始化,读取xml的文件
def MR_xml_init():
    #if mr_test_process run again, information should clean first
    gl.MR_DICT.clear()
    #find the mr_path file
    if gl.TEST_CONF['mr_path'] != '' and os.path.isdir(gl.TEST_CONF['mr_path']):
        gl.MR_TEST_PATH = gl.TEST_CONF['mr_path']
    file_list = os.listdir(gl.MR_TEST_PATH)
    
    #relax restriction, judge and get file when file_name strictly right before
    # re_search_str = gl.TEST_CONF['standard_LTE'] + '_MR'
    re_search_str = '.*_MR[O|E|S].*\.xml$'
    
    for file in file_list:
        full_file = os.path.join(gl.MR_TEST_PATH, file)
        if  os.path.isdir(full_file) == False and re.match(re_search_str, file) != None:
            # MR_FILE : save the file_name information 
            MR_FILE = {}
            # MR_LIST: [0]: 'MRO':'FDD-LTE_MRO_CMCC_NanoCellTestTool_115_20220622101500.xml', 'MRE': .. , 'MRS': ..
            # [1]: MR_FILE -> mro , [2]: MR_FILE -> mre , [3]: MR_FILE -> mrs  
            MR_LIST = [{'MRO':'','MRE':'','MRS':''},{}, {}, {}]
            
            #no need now
            # if MR_xml_file_name_accuracy(file) == False :
            #     continue
            
            
            file_data_list = file.split('_')
            # print ('111-->' + full_file)
            # print ('--->' + full_file)

            MR_FILE['standard_LTE'] = file_data_list[0]
            MR_FILE['type'] = file_data_list[1]
            MR_FILE['OEM'] = file_data_list[2]
            MR_FILE['OmcName'] = get_filename_omc_name(file_data_list)
            MR_FILE['enbid'] = file_data_list[len(file_data_list) - 2]
            MR_FILE['SampleBeginTime'] = file_data_list[len(file_data_list) - 1].split('.')[0]
            #most important is xmldom to deal MR_xml file
            MR_FILE['xmldom'] = xmldom.parse(full_file)
            
            
            #python2可以使用has_key(), 后面用的是__contains__() , in, get
            if gl.MR_DICT.get(MR_FILE['SampleBeginTime']) == None:
                gl.MR_DICT[MR_FILE['SampleBeginTime']] = MR_LIST 
                
                
            #list中0对应的是mro，mre，mrs对应的名字  MR_DICT[ '20211110165200' ] [0]['MRO'] = FDD-LTE_MRO_CMDI_OMC-R_1_2048_226_20211110165200.xml
            gl.MR_DICT[MR_FILE['SampleBeginTime']][0][MR_FILE['type']] = file
            #add MR_FILE
            for mr_list_entity in gl.MR_DICT:
                list_entity = gl.MR_DICT[mr_list_entity]
                if mr_list_entity == MR_FILE['SampleBeginTime']:
                    #list_entity.append(MR_FILE)
                    list_entity[gl.MR_TYPE[MR_FILE['type']]] = MR_FILE
                    break

    return

#PM文件xml的初始化,读取xml的文件
def PM_xml_init():
    #if mr_test_process run again, information should clean first
    gl.PM_DICT.clear()
    #find the pm_path file
    if gl.TEST_CONF['pm_path'] != '' and os.path.isdir(gl.TEST_CONF['pm_path']) :
        gl.PM_TEST_PATH = gl.TEST_CONF['pm_path']
    file_list = os.listdir(gl.PM_TEST_PATH)

    re_search_str = '^A_LTE_.*\.xml$'
    
    for file in file_list:
        full_file = os.path.join(gl.PM_TEST_PATH, file)
        if  os.path.isdir(full_file) == False and re.match(re_search_str, file) != None:            
            file_data = file.split('_')[2]
            time_split = file_data.split('.')
            time_entity = time_split[0] + time_split[1].split('+')[0] + '00'
            
            gl.PM_DICT[time_entity] = xmldom.parse(full_file)
            
        
    return

#======================================================
#    file 
#======================================================

#don't have the conf.xml file, then create it
def create_conf_xml(path):
    with io.open(path, 'w', encoding='utf-8') as file_object:
        file_object.write(gl.CONF_XML_DATA)

# write data.txt, fileheader 
def mr_out_file_data_head():
    # gl.OUT_PATH = gl.OUTPUT_PATH + "data.txt"
    gl.OUT_STR_LIST.append(' '*14 + "<--------------LTE-TEST-RESULT-------------->")
    # with open(gl.OUTPUT_PATH + "data.txt", 'w') as file_object:
    #     file_object.write()

# get file_list by file_filter
get_file_by_path = lambda filename_filger : glob.glob(filename_filger)

#function interface, use TEST_CONF to run function
def mr_function_process(function_call, id_str, info=""):
    if gl.TEST_CONF[id_str] == '1':
        try:
            test_out_data_item_header(id_str)                
            function_call()
        except BaseException as result:
            print (result)
            if not gl.IS_LINUX:
                write_info(result)
            
            
#write every test project's text header            
def test_out_data_item_header(out_type):
    gl.OUT_STR_LIST.append("\n<----------------------->\n\n---------" + out_type + "---------\n")
    if out_type not in gl.TEST_OUT:
        return
    for i in range( int(gl.TEST_OUT[out_type][0])):
        gl.OUT_STR_LIST.append(gl.TEST_ITEM_LIST[gl.TEST_OUT[out_type][i+1]])
        gl.OUT_STR_LIST.append(" | ")
        if gl.TEST_OUT[out_type][i+1] == 42:
            gl.OUT_STR_LIST.append("\n")
    gl.OUT_STR_LIST.append("\n")
        
# write output file text in  output/data.txt
def write_output_file_text():
    # print (gl.OUT_STR_LIST)
    if gl.IS_PY2:
        with io.open(gl.OUT_PATH, 'w', encoding='utf-8') as file_object:
            file_object.writelines( [ x if type(x) != type(str()) else x.decode("utf-8") for x in gl.OUT_STR_LIST ] )       
    else:
        with open(gl.OUT_PATH, 'w') as file_object:
            file_object.writelines(gl.OUT_STR_LIST) 

# add information to dict
def out_text_dict_append_list(out_dict, key, str):
    if out_dict.__contains__(key) == False:
        out_dict[key] = []
    out_dict[key].append(str)
    
   
    
    
# parse UElog.txt 
#TODO: UElog.txt is not certain, should modify by actual file 
# [  { "timestamp":'' ,"MR.LteScRSRP":-1 , "MR.LteScRSRQ":-1, 'nc_info':[ 'MR.LteNcPci': '', 'MR.LteNcRSRP': -1, 'MR.LteNcRSRQ':-1  ]  }  ]
def read_ue_log_filter(file_name, start_time_str, end_time_str):
    ue_info = []
    state_flag = -1
    start_time_stamp = get_timestamp_by_str_format(start_time_str) - int(gl.MR_CONF['SamplePeriod'])
    end_time_stamp = get_timestamp_by_str_format(end_time_str) 
    #ueLog is a text. so use file_read to get data
    
    if os.path.exists(file_name):
        with open(file_name, "rt") as file_object:
            line = file_object.readline()
            while line:
                
                split_list = get_split_value(line, ' ', [' ', '\n'])
                if 'L->MeasurementReport' in line:
                    # mean a report packet, also can move to 'Computer' in line , to append
                    temp_dict = {"timestamp":'' ,"MR.LteScRSRP":-1 , "MR.LteScRSRQ":-1, 'nc_info':[] }
                                #  'MR.LteNcEarfcn':-1,'MR.LteNcPci':-1, 'MR.LteNcRSRP': -1, 'MR.LteNcRSRQ':-1}
                    ue_info.append(temp_dict)
                    state_flag = 0
                if 'Computer' in line:
                    time_str = split_list[2]
                    
                    temp_time_str_ = ue_info[-1]['timestamp'] = str(start_time_str[0:11] + time_str).strip()

                    temp_time_stamp = get_timestamp_by_str_format(temp_time_str_)

                    if temp_time_stamp <= start_time_stamp and  len(ue_info) != 0:
                        ue_info.pop()
                        state_flag = -1
                    if temp_time_stamp > end_time_stamp :
                        break
                    
                if 'measId' in line and state_flag == 0:
                    if str(split_list[-1]).strip('\r').strip('\n').strip('\t') not in gl.TEST_CONF['ue_meas_id'] and len(ue_info) != 0:
                        # print (split_list[-1])
                        ue_info.pop()
                        state_flag = -1
                if 'measResultPCell' in line and state_flag == 0:
                    state_flag = 1
                if state_flag == 1:
                    if 'rsrpResult' in line:
                        # print (ue_info)
                        ue_info[-1]['MR.LteScRSRP'] = int(split_list[1])
                    if 'rsrqResult' in line :
                        ue_info[-1]['MR.LteScRSRQ'] = int(split_list[1])
                if 'measResultNeighCells' in line and state_flag == 1:
                    state_flag = 2
                if state_flag == 2:
                    if 'physCellId' in line:
                        temp_dict = {'MR.LteNcPci':(int(split_list[1])), 'MR.LteNcRSRP': -1, 'MR.LteNcRSRQ':-1}
                        ue_info[-1]['nc_info'].append(temp_dict)
                    if 'rsrpResult' in line:
                        ue_info[-1]['nc_info'][-1]['MR.LteNcRSRP'] = int(split_list[1])
                    if 'rsrqResult' in line :
                        ue_info[-1]['nc_info'][-1]['MR.LteNcRSRQ'] = int(split_list[1])

                line = file_object.readline() 
    return ue_info


# parse phr csv 
def get_phr_data(file_name, start_time_str, end_time_str):
    phr_info = []
    start_time_stamp = get_timestamp_by_str_format(start_time_str) - int(gl.MR_CONF['SamplePeriod'])
    end_time_stamp = get_timestamp_by_str_format(end_time_str)
    with open(file_name) as csvfile:
        csvReader = csv.reader(csvfile)
        temp_i = 0
        for row in csvReader:
            if str(row[5]) == 'Power Headroom':
                continue
            temp_time_stamp = get_timestamp_by_str_format('20' + row[2], "%Y-%m-%d %H:%M:%S.%f")
            if temp_time_stamp < start_time_stamp or temp_time_stamp > end_time_stamp:
                continue
            if temp_i % 4 == 0 :
                temp_dict = {'timestamp':row[2], 'phr':row[5]}
                phr_info.append(temp_dict)
            temp_i += 1
    return phr_info

# create 6.1 out excel
def create_61_out_excel(time_entity, ue_info_dict, MRS_info_dict):
    book = Workbook()
    sheet = book.active
    value_list = ['MR.RSRP', 'MR.RSRQ', 'MR.PowerHeadRoom']
    for i in range(3):
        sheet.cell(1, 4*i + 1, "测量项")
        sheet.cell(1, 4*i + 2, "信令文件记录采样点数量")
        sheet.cell(1, 4*i + 3, "MR采样点数量")
        sheet.cell(1, 4*i + 4, "偏离程度")
        
        for j in range(1, len(ue_info_dict[value_list[i]]) + 1):
            item_value = int(MRS_info_dict[value_list[i]][j - 1])
            sheet.cell(row = j + 1, column= i * 4 + 1 ,value = value_list[i ] + ".%02d"%(j - 1)) 
            sheet.cell(row = j + 1, column= 4*i + 2, value= ue_info_dict[value_list[i]][j - 1])
            sheet.cell(row = j + 1, column= 4*i + 3, value= item_value)
            sheet.cell(row = j + 1, column= 4*i + 4, value= item_value - ue_info_dict[value_list[i]][j - 1])
    book.save(gl.OUTPUT_PATH + time_entity + ".xlsx")


# check mro standard by  self
def is_mro_correct(mro_root):
    smr_target_list = gl.mro_smr_target_list

    if mro_root.nodeName != 'bulkPmMrDataFile':
        return False,'bulkPmMrDataFile'
    fileHeader_list = xml_element(mro_root, 'fileHeader')
    if len(fileHeader_list) != 1:
        return False,'fileHeader'
    eNB_list = xml_element(mro_root, 'eNB')
    if len(eNB_list) != len(gl.TEST_CONF['cellid'].split(',')):
        return False,'eNB_id not match'
    if len(eNB_list) == 0 and (len(xml_element(mro_root, 'measurement')) != 0 or len(xml_element(mro_root, 'smr')) != 0 or len(xml_element(mro_root, 'object')) != 0):
        return False, 'format error enb'
    elif len(eNB_list) == 0:
        return True,'no enb'

    for eNB_entity in eNB_list:
        if xml_attr(eNB_entity, 'id') not in gl.TEST_CONF['cellid']:
            return False, 'eNB_id not match'
        measurement_list = xml_element(eNB_entity, 'measurement')
        if len(measurement_list) == 0 and (len(xml_element(mro_root, 'smr')) != 0 or len(xml_element(mro_root, 'object')) != 0):
            return False, 'format error measurement'
        elif len(measurement_list) == 0:
            return True,'correct'
        for measurement_entity in measurement_list:
            smr_list = xml_element(measurement_entity, 'smr')
            if len(smr_list) != 1 and len(smr_list) != 0:
                return False, 'smr format error'+str(len(smr_list))
            for mr_item in get_split_value(xml_text(smr_list[0])):
                if mr_item not in smr_target_list[1]:
                    if ( mr_item not in smr_target_list[0] and  is_mr_item_need_exist(mr_item) == True) or \
                        ( mr_item in smr_target_list[0] and  is_mr_item_need_exist(mr_item) == False):
                        return False, 'L3 smr not match  <%s>'%(mr_item)
                else:
                    if ( mr_item not in  smr_target_list[1]  and  is_mr_item_need_exist(mr_item) == True) or (mr_item in smr_target_list[1]  and  is_mr_item_need_exist(mr_item) == False):
                        return False, 'L2 smr not match  <%s>'%(mr_item)

    return True,'correct'


def is_mre_correct(mre_root):
    smr_target_str = gl.mre_smr_target_str

    if mre_root.nodeName != 'bulkPmMrDataFile':
        return False, 'bulkPmMrDataFile'
    fileHeader_list = xml_element(mre_root, 'fileHeader')
    if len(fileHeader_list) != 1:
        return False, 'fileHeader'
    eNB_list = xml_element(mre_root, 'eNB')
    if len(eNB_list) != len(gl.TEST_CONF['cellid'].split(',')):
        return False, 'eNB id not match'
    if len(eNB_list) == 0 and (len(xml_element(mre_root, 'measurement')) != 0 or len(xml_element(mre_root, 'smr')) != 0 or len(xml_element(mre_root, 'object')) != 0):
        return False, 'format error'
    elif len(eNB_list) == 0:
        return True,'correct'

    for eNB_entity in eNB_list:
        if xml_attr(eNB_entity, 'id')  not in gl.TEST_CONF['cellid']:
            return False, 'eNB id not match'
        measurement_list = xml_element(eNB_entity, 'measurement')
        if len(measurement_list) == 0 and (len(xml_element(mre_root, 'smr')) != 0 or len(xml_element(mre_root, 'object')) != 0):
            return False, 'format error'
        elif len(measurement_list) == 0:
            return True,'correct'
        for measurement_entity in measurement_list:
            smr_list = xml_element(measurement_entity, 'smr')
            if len(smr_list) != 1 and len(smr_list) != 0:
                return False, 'smr error'
            for mr_item in get_split_value(xml_text(smr_list[0])):
                if (mr_item not in  smr_target_str and is_mr_item_need_exist(mr_item) == True) or ( mr_item in smr_target_str and is_mr_item_need_exist(mr_item) == False):
                    return False, 'smr not match <%s>' % (mr_item)

    return True,'correct'

def is_mrs_correct(mrs_root):
    smr_target_list = gl.mrs_smr_target_dict
        
    if mrs_root.nodeName != 'bulkPmMrDataFile':
        return False, 'bulkPmMrDataFile'
    fileHeader_list = xml_element(mrs_root, 'fileHeader')
    if len(fileHeader_list) != 1:
        return False, 'fileHeader'
    eNB_list = xml_element(mrs_root, 'eNB')
    if len(eNB_list) != len(gl.TEST_CONF['cellid'].split(',')):
        return False, 'eNB_id not match'
    if len(eNB_list) == 0 and (len(xml_element(mrs_root, 'measurement')) != 0 or len(xml_element(mrs_root, 'smr')) != 0 or len(xml_element(mrs_root, 'object')) != 0):
        return False, 'format error'
    elif len(eNB_list) == 0:
        return True,'correct'

    for eNB_entity in eNB_list:
        if xml_attr(eNB_entity, 'id') not in gl.TEST_CONF['cellid']:
            return False, 'eNB_id not match'
        measurement_list = xml_element(eNB_entity, 'measurement')
        if len(measurement_list) == 0 and (len(xml_element(mrs_root, 'smr')) != 0 or len(xml_element(mrs_root, 'object')) != 0):
            return False, 'format error1'
        elif len(measurement_list) == 0:
            return True,'correct'
        for measurement_entity in measurement_list:
            smr_list = xml_element(measurement_entity, 'smr')
            mr_name = xml_attr(measurement_entity, 'mrName')
            if len(smr_list) != 1 and len(smr_list) != 0:
                return False, 'smr error'
            if xml_text(smr_list[0]).strip() != smr_target_list[mr_name].strip() and is_mr_item_need_exist(xml_attr(measurement_entity, 'mrName')) == True:
                return False, 'smr not match <%s>'%( measurement_entity.getAttribute('mrName'))

    return True,'correct'
     
#======================================================
#    time 
#======================================================

#get strftime , current time
def get_time_format(format="%Y:%m:%d %H:%M:%S"):
    temp_time = time.time()
    localtime = time.localtime(temp_time)
    date_time = time.strftime(format, localtime)
    return date_time

# tranport  timestamp  => strftime
def get_time_format_by_timestamp(time_stamp, format="%Y-%m-%dT%H:%M:%S.%f"):
    d = datetime.datetime.fromtimestamp(time_stamp/1000)
    date_time = d.strftime(format)
    return date_time

#notice: timestamp is milliseconds   strftime => timestamp
def get_timestamp_by_str_format(time_str, format="%Y-%m-%dT%H:%M:%S.%f"):
    time_array = datetime.datetime.strptime(time_str, format)
    timestamp = int(time.mktime(time_array.timetuple())*1000.0 + time_array.microsecond/1000.0)
    return timestamp

def get_two_linux_time_sub():
    remote_user = gl.INFORM_DICT[gl.MR_DOWNLOAD_IP]['user']
    remote_passwd = gl.INFORM_DICT[gl.MR_DOWNLOAD_IP]['passwd']
    enb_ip = [key for key in gl.INFORM_DICT if gl.TEST_CONF['enbid'] in key][0]
    print (enb_ip)
    enb_user = gl.INFORM_DICT[enb_ip]['user']
    enb_passwd = gl.INFORM_DICT[enb_ip]['passwd']
    ssh1 = ssh_tool(enb_ip, 22, enb_user, enb_passwd)
    ssh2 = ssh_tool(gl.MR_DOWNLOAD_IP, 22, remote_user, remote_passwd)
    time1 = ssh1.cmd_run("date +%s")
    time2 = ssh2.cmd_run("date +%s")
    del ssh1
    del ssh2
    return int(time1) - int(time2)

# check the str is time_format
def is_str_format_time(str, format):
    try:
        datetime.datetime.strptime(str, format)
        return True
    except:
        return False

#======================================================
#    MR standard check 
#======================================================

# this is check from TEST_CONF cellid, mod by manual
#TODO: need to modify, when eciid = eciid * 256 + cellid,  judge the enbid and cellid both
def is_cell_id_exist(eciid):
    cell_id = eciid & 0xff
    if str(cell_id) in gl.TEST_CONF['cellid'].split(','):
        return True, cell_id
    else:
        return False,cell_id

def is_mr_item_need_exist(mr_name):
    if gl.MR_CONF['MeasureItems'] == 'all':
        return True
    elif mr_name in gl.MR_CONF['MeasureItems'] :
        return True
    return False

    
#======================================================
#    MR get handle from global variable 
#======================================================
# get xml node elements
xml_element     =   lambda node, childname: node.getElementsByTagName(childname)
# get xml attribute 
xml_attr        =   lambda node, attribute: node.getAttribute(attribute)
# get xml document root
xml_doc         =   lambda root : root.documentElement
# get mr_xml dom
xml_dom         =   lambda time_key, type : gl.MR_DICT[time_key][gl.MR_TYPE[type]]['xmldom']
# get xml node text
xml_text        =   lambda node : node.firstChild.data
# get xml file name
get_file_name   =   lambda time_key, type : gl.MR_DICT[time_key][0][type]
# get xml file list
get_file_dict   =   lambda time_key, type : gl.MR_DICT[time_key][gl.MR_TYPE[type]]

debug_info = lambda  result, file = '', function = '', str='': gl.OUT_DEBUG_LIST.append('%s | %s | [%d] -> %s '%(file, function, result.__traceback__.tb_lineno, result) + ' # ' +  str)


#get prb list from mr_conf 
def get_prb_list():
    prbnum_list = []
    list1 = get_split_value(gl.MR_CONF['PrbNum'], ',', ['{', '}', ','])
    for entity in list1:
        if '..' in entity:
            list2 = get_split_value(entity, '..', [' '])
            for i in range(int(list2[0]), int(list2[-1]) + 1):
                prbnum_list.append(str(i))     
        else:
            prbnum_list.append(entity)
    return prbnum_list

# string to strip and split -> a list
def get_split_value(str1, sequence = ' ', list_sq=[' ']):
    for squ in list_sq:
        str1 = str1.strip(squ)
    return str1.split(sequence)

# get the mr_nama's position, modify the mr_name_dict value
def get_mr_item_pos(mr_name_dict, smr_str):
    smr = smr_str.strip()
    smr_list = smr.split(" ")
    if smr not in mr_name_dict:
        mr_name_dict[smr] = []
    for i in smr_list:
        mr_name_dict[smr].append(i)

def get_mr_item_pos_single_smr(mr_name_dict, smr_str):
    smr_list = get_split_value(smr_str)
    for i in range(len(smr_list)):
        mr_name_dict[smr_list[i]] = i

#get the mr_name's position, save in dict pos        
def get_mr_item_pos_dict(mr_name_dict, smr_str):
    smr_list = smr_str.split(' ')
    for mr_name in mr_name_dict:
        temp_flag = 0
        for i in range(len(smr_list)):
            if mr_name == smr_list[i]:
                if mr_name_dict.__contains__(mr_name) == False:
                    mr_name_dict[mr_name] = {'pos':0}
                if mr_name_dict[mr_name].__contains__('pos') == False:
                    mr_name_dict[mr_name]['pos'] = 0
                mr_name_dict[mr_name]['pos'] = i
                temp_flag = 1
                break
        if temp_flag == 0:
            if mr_name_dict.__contains__(mr_name) == False:
                mr_name_dict[mr_name] = {'pos':0}
            if mr_name_dict[mr_name].__contains__('pos') == False:
                mr_name_dict[mr_name]['pos'] = 0
            mr_name_dict[mr_name]['pos'] = 0
    
# get sum of a number string
def add_digital_string(digital_string):
    temp_count = 0
    for number_string in digital_string.split(' '):
        if number_string.isdigit() == True:
            temp_count += int(number_string)
    return temp_count

# get position single
def get_mro_pos_by_smr(mr_name, smr_str):
    smr_list = get_split_value(smr_str)
    for i in range(len(smr_list)):
        if smr_list[i] == mr_name:
            return i
    return -1

# check enbid
def is_enb_id_exist(enb_id):
    if str(enb_id) not in gl.TEST_CONF['enbid']:
        return False
    else:
        return True

# check eciid
#TODO: eci  
def is_eci_correct(eci_id):
    cell_id_list = get_split_value(gl.TEST_CONF['cellid'], ',')
    enb_id_list =   get_split_value(gl.TEST_CONF['enbid'], ',')

    eci_id_list = []
    for i in range(len(cell_id_list)):
        for j in range(len(enb_id_list)):
            #TODO: modify here , when eci changed. business partner test the endid = 1008, cellid = 115, then eci is wrong. 
            temp_eci_id = (int(enb_id_list[j]) << 8) | int(cell_id_list[i])
            eci_id_list.append(temp_eci_id)
    if eci_id in eci_id_list and is_eci_correct_by_MRECGIList(eci_id):
        return True
    return False

#TODO: MRECGIList need to test, now LTE_MR is not test this function carefully, 
def is_eci_correct_by_MRECGIList(eciid):
    if gl.MR_CONF['MRECGIList'] == 'all':
        return True
    ecgi_list = gl.MR_CONF['MRECGIList'].split(',')
    for ecgi_entity in ecgi_list:
        eci = int(ecgi_entity.split('-')[1])
        if int(eciid) == eci:
            return True
    return False

# check the value of mr_item if it's right  in smr_str
def is_mr_value_correct(smr_str, value_str):
    smr_list = get_split_value(smr_str)
    value_list = get_split_value(value_str)
    range_dict = {'MR.LteScRSRP':[0, 97],'MR.LteNcRSRP':[0, 97],'MR.LteScRSRQ':[0, 34],'MR.LteNcRSRQ':[0, 34],'MR.LteScPHR':[0, 63],'MR.LteScRIP':[0, 511],\
                  'MR.LteScSinrUL':[0, 36],'MR.LteScEarfcn':[0, 41589],'MR.LteScPci':[0, 503],'MR.LteNcEarfcn':[0, 41589],\
                         'MR.LteNcPci':[0, 503],'MR.GsmNcellBcch':[0, 1023],'MR.GsmNcellCarrierRSSI':[0, 63],'MR.GsmNcellNcc':[0, 7],'MR.GsmNcellBcc':[0, 7]}
    if len(smr_list) != len(value_list):
        return False
    for i in range(len(smr_list)):
        if range_dict.__contains__(smr_list[i]) == False:
            return False
        if value_list[i].isdigit() == False:
            continue
        if int(value_list[i]) < range_dict[smr_list[i]][0] or int(value_list[i]) > range_dict[smr_list[i]][1]:
            return False
    return True
    
# get mre mr_item position
def get_mre_pos_list_by_mapping(mre_conf_dict, smr_str):
    standard_smr_str_list = ['MR.LteScRSRP','MR.LteNcRSRP', 'MR.LteScRSRQ', 'MR.LteNcRSRQ', 'MR.LteScEarfcn', 'MR.LteScPci', 'MR.LteNcEarfcn', 'MR.LteNcPci', 
                             'MR.GsmNcellBcch', 'MR.GsmNcellCarrierRSSI', 'MR.GsmNcellNcc', 'MR.GsmNcellBcc']
    mre_smr_head = 'MR.LteScRSRP MR.LteNcRSRP MR.LteScRSRQ MR.LteNcRSRQ MR.LteScEarfcn MR.LteScPci MR.LteNcEarfcn MR.LteNcPci MR.GsmNcellBcch MR.GsmNcellCarrierRSSI MR.GsmNcellNcc MR.GsmNcellBcc '
    if smr_str.strip() == mre_smr_head.strip():
        return
    smr_list = smr_str.split(' ')
    for mre_event in mre_conf_dict:
        temp_pos_list = mre_conf_dict[mre_event]['pos'][:]
        mre_conf_dict[mre_event]['pos'].clear()
        for i in range(len(smr_list)):
            for j in range(len(temp_pos_list)):
                if standard_smr_str_list[temp_pos_list[j]] == smr_list[i]:
                    mre_conf_dict[mre_event]['pos'].append(i)
                    break
                
def get_start_end_timestr_from_mr_file(time_stamp_raw):
    mr_start_time_str = time_stamp_raw[0:4] + '-' + time_stamp_raw[4:6] + '-' + time_stamp_raw[6:8] + 'T' + time_stamp_raw[8:10] + ':' + \
                                    time_stamp_raw[10:12] + ':' + time_stamp_raw[12:14] + '.000'
    mr_start_time_stamp = get_timestamp_by_str_format(mr_start_time_str)
    mr_end_time_stamp = mr_start_time_stamp + int(gl.MR_CONF['UploadPeriod']) * 60000
    mr_end_time_str = get_time_format_by_timestamp(mr_end_time_stamp)
    return mr_start_time_str[0:23], mr_start_time_stamp, mr_end_time_str[0:23], mr_end_time_stamp
                
                
def rsrp_count(value, rsrp_statical_list):
    # rsrp_statical_list = [ 0 for x in range(0, 48) ]
    if value == 'NIL':
        return
    value = int(value)
    if value < 0 or value > 97:
        return
    for rsrp_entity in gl.global_rsrp_list:
        if value == rsrp_entity[0]:
            rsrp_statical_list[rsrp_entity[1]] += 1
            break


def rsrq_count(value, rsrq_statical_list):
    # rsrq_statical_list = [ 0 for x in range(0, 18) ]
    if value == 'NIL':
        return
    value = int(value)
    if value < 0 or value > 34:
        return
    for rsrq_entity in gl.global_rsrq_list:
        if value == rsrq_entity[0]:
            rsrq_statical_list[rsrq_entity[1]] += 1
            break
    return rsrq_statical_list

def phr_count(value, phr_statical_list):
    # phr_statical_list = [ 0 for x in range(0, 64) ]
    if value == 'NIL':
        return
    # print (value)
    value = int(value)
    if value < 0 or value > 63:
        return
    phr_statical_list[value] += 1   
    return phr_statical_list

def sinrul_count(value, sinrul_statical_list):
    # sinrul_statical_list = [ 0 for x in range(0, 37) ]
    if value == 'NIL':
        return
    value = int(value)
    if value < 0 or value > 36:
        return
    sinrul_statical_list[value] += 1
    return sinrul_statical_list

def rip_count(value, id, rip_info_dict):
    # rip_info_dict = {}
    if value == 'NIL':
        return
    value = int(value)
    if value < 0 or value > 520:
        return
    if id not in rip_info_dict:
        return
    if value != 0:
        value = (value + 9) / 10
    rip_info_dict[id][int(value)] += 1
    
def get_map_pci_earfcn():
    map_dict = {}
    map_string = gl.TEST_CONF['nc_pci_earfcn']
    map_item_list = get_split_value(map_string, ',')
    for item in map_item_list:
        temp_list = get_split_value(item, '-')
        map_dict[temp_list[0]] = temp_list[1]
    return map_dict


def get_measureItem_list(mr_dict):
    measure_item_list = ['MR.RSRP','MR.RSRQ','MR.ReceivedIPower','MR.PowerHeadRoom','MR.SinrUL','MR.RIPPRB','MR.RsrpRsrq','MR.RipRsrp','MR.RipRsrq','MR.SinrULRip',\
        'MR.LteScRSRP','MR.LteNcRSRP','MR.LteScRSRQ','MR.LteNcRSRQ','MR.LteScPHR','MR.LteScRIP','MR.LteScSinrUL','MR.LteScEarfcn','MR.LteScPci','MR.LteNcEarfcn',\
                         'MR.LteNcPci','MR.GsmNcellBcch','MR.GsmNcellCarrierRSSI','MR.GsmNcellNcc','MR.GsmNcellBcc']
    if gl.MR_CONF['MeasureItems'] == 'all':
        for i in range(len(measure_item_list)):
            mr_dict[measure_item_list[i]] = 0
    else:
        meas_list = get_split_value(gl.MR_CONF['MeasureItems'], ',')
        for i in range(len(meas_list)):
            for j in range(len(measure_item_list)):
                if meas_list[i] == measure_item_list[j]:
                    mr_dict[meas_list[i]] = 0
                    
def is_mrs_measurement_smr_value_correct(mr_name, smr_str, value_str):
    smr_standard_dict = gl.mrs_smr_target_dict
    for mr_entity in smr_standard_dict:
        if mr_entity == mr_name:
            smr_standard_list = get_split_value(smr_standard_dict[mr_entity])
            smr_list = get_split_value(smr_str)
            value_list = get_split_value(value_str)
            if len(smr_standard_list) != len(smr_list):
                return False
            if len(smr_list) != len(value_list):
                return False

def show_dict(dict1, name=''):
    print ('\n==================>   %s   <==================\n'%(name) )
    for time_entity in dict1:
        print (" %s   :   %s"%(time_entity, dict1[time_entity]))
    print ('\n')
        

#======================================================
#    else tool
#======================================================


class ssh_tool:
    def __init__(self,remote_ip, port=22, username="root", passwd=""):
        # self.ssh = paramiko.SSHClient()
        # self.ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        self.ssh.connect(remote_ip, port, username, passwd, timeout=3)
    def __del__(self):
        self.ssh.close()
    def cmd_run(self, cmd=""):
        stdin, stdout, stderr = self.ssh.exec_command(cmd)
        return stdout.read()
