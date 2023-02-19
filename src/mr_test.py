
# encoding: utf-8

import mr_utils 
import mr_global as gl
import re
import os
import openpyxl
import math
import time
import sys
from lxml import etree

# 
# _ooOoo_
# o8888888o
# 88" . "88
# (| -_- |)
#  O\ = /O
# ___/`---'\____
# .   ' \\| |// `.
# / \\||| : |||// \
# / _||||| -:- |||||- \
# | | \\\ - /// | |
# | \_| ''\---/'' | |
# \ .-\__ `-` ___/-. /
# ___`. .' /--.--\ `. . __
# ."" '< `.___\_<|>_/___.' >'"".
# | | : `- \`.;`\ _ /`;.`/ - ` : | |
# \ \ `-. \_ __\ /__ _/ .-` / /
# ======`-.____`-.___\_____/___.-`____.-'======
# `=---='
#          .............................................
#           佛曰：bug泛滥，我已瘫痪！
# =============================================================

# 
#TODO: when you want to add the test function by yourself, add it to the end
# don't forget add the function in mr_user






#========================================================================================
#             test 51
# check the number of MRO,MRE,MRS file, if they are equal to predict number
#========================================================================================

def test51_file_integrity():
    # get test_51 file_header, the tag
    # mr_utils.test_out_data_item_header("test_51")
    # get current time

    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    # time_sub = mr_utils.get_two_linux_time_sub()
    try:
        #统计预期的文件总数,计算: 测量总时长/测量周期
        #start and end time : initialization 
        startTimestamp = 0
        endTimestamp = mr_utils.get_timestamp_by_str_format(gl.test51_end_time_str, '%Y-%m-%dT%H:%M:%SZ')
        
        # in total time , predict the mr file number
        predict_file_num =  float(gl.TEST_CONF['test_total_time'])*3600 / (60 * int(gl.MR_CONF['UploadPeriod']) )
        
        # not the special situation, compute the duration between sampleBegin and EndTime. then get predict_file_num
        if gl.MR_CONF['SampleEndTime'] != '0001-01-01T00:00:00Z' and gl.MR_CONF['SampleBeginTime'] != '0001-01-01T00:00:00Z' :
            startTimestamp = mr_utils.get_timestamp_by_str_format(gl.MR_CONF['SampleBeginTime'], '%Y-%m-%dT%H:%M:%SZ') / 1000
            endTimestamp = mr_utils.get_timestamp_by_str_format(gl.MR_CONF['SampleEndTime'], '%Y-%m-%dT%H:%M:%SZ') / 1000
            
            spec_timestamp = endTimestamp - startTimestamp
            predict_file_num = (spec_timestamp - spec_timestamp%60) / (60 * int(gl.MR_CONF['UploadPeriod']) )
        
        mr_integrity_flag = 0
        mro_file_num = 0
        mre_file_num = 0
        mrs_file_num = 0
        #MRS文件数量统计

        for time_entity in gl.MR_DICT:
            file_time_stamp = mr_utils.get_timestamp_by_str_format(time_entity, '%Y%m%d%H%M%S') / 1000 - 8 * 60*60

            # not in the duration , not count
            if file_time_stamp < startTimestamp  or file_time_stamp > endTimestamp :
                continue
            
            mro_file_name = gl.MR_DICT[time_entity][0]['MRO']
            mre_file_name = gl.MR_DICT[time_entity][0]['MRE']
            mrs_file_name = gl.MR_DICT[time_entity][0]['MRS']
            if mro_file_name != '':
                # ignore, this in test71
                # if os.path.getsize(mro_file_name) == 0 or mr_utils.MR_xml_file_name_accuracy(mro_file_name) == False :
                #     mr_integrity_flag = mr_integrity_flag | (0x1 << 1)
                # else :
                    mro_file_num += 1
            if mre_file_name != '':
                # ignore, this in test71
                # if os.path.getsize(mre_file_name) == 0 or mr_utils.MR_xml_file_name_accuracy(mre_file_name) == False:
                #     mr_integrity_flag = mr_integrity_flag | (0x1 << 1)
                # else:
                    mre_file_num += 1
            if mrs_file_name != '':
                # ignore, this in test71
                # if os.path.getsize(mrs_file_name) == 0 or mr_utils.MR_xml_file_name_accuracy(mrs_file_name) == False:
                #     mr_integrity_flag = mr_integrity_flag | (0x1 << 1)
                # else:
                    mrs_file_num += 1

        #==========================================================================
        #   write output 
        #==========================================================================
        gl.OUT_STR_LIST.append(date_time + " | ")
        gl.OUT_STR_LIST.append(gl.TEST_CONF['test_total_time'] + " | ")
        gl.OUT_STR_LIST.append(str(len(gl.TEST_CONF['enbid'].strip(',').split(','))) + " | ")
        gl.OUT_STR_LIST.append(str(predict_file_num * (len(gl.MR_CONF['MeasureType'].strip(',').split(',')))) + ' | ')
        gl.OUT_STR_LIST.append(str(mrs_file_num) + ' | ')
        gl.OUT_STR_LIST.append(str(mro_file_num) + ' | ')
        gl.OUT_STR_LIST.append(str(mre_file_num) + ' | ')
        gl.OUT_STR_LIST.append(str(  # mr_integrity_flag & (0x1 << 3) == 0 and \
                                    (predict_file_num == mrs_file_num ) \
                                    if re.search(r'MRS',gl.MR_CONF['MeasureType']) != None \
                                    else  (mrs_file_num == 0))  + ' | ')
        gl.OUT_STR_LIST.append(str(  # mr_integrity_flag & (0x1 << 1) == 0 and \
                                    (predict_file_num == mro_file_num  ) \
                                    if re.search('MRO', gl.MR_CONF['MeasureType']) != None \
                                    else mro_file_num == 0) + ' | ')
        gl.OUT_STR_LIST.append(str(  # mr_integrity_flag & (0x1 << 2) == 0 and \
                                    (predict_file_num == mre_file_num )\
                                    if re.search('MRE', gl.MR_CONF['MeasureType']) != None \
                                    else mre_file_num == 0 ) + ' | \n')
    except Exception as result:
        # this is exception debug
        raise Exception('[%s]   %s |  %s'%(__name__ ,str(result.__traceback__.tb_lineno) if not gl.IS_PY2  else '',result))
    

    


#========================================================================================#
#             test 52
# check the number of cell  <==>  actual cell number
#========================================================================================#

def test52_file_integrity():
    # mr_utils.test_out_data_item_header("test_52")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    out_dict = {}
    output_temp_dict = {}
    out_sort_list = ["MR.RSRP", "MR.RSRQ","MR.PowerHeadRoom", "MR.SinrUL" ]
    for time_entity in gl.MR_DICT:
        try:
            # get MRS filename
            mrs_file_entity = mr_utils.get_file_name(time_entity, 'MRS')
            if mrs_file_entity == '':
                continue
            
            if out_dict.__contains__(mrs_file_entity) == False :
                # every mr_name has "num" -> cell number, "list" -> cell id list
                out_dict[mrs_file_entity] = {"MR.RSRP":     {'num':0, 'list':[]     }, 
                                             "MR.RSRQ":     {'num':0, 'list':[]     }, 
                                             "MR.SinrUL":   {'num':0, 'list':[]     }, 
                                             "MR.PowerHeadRoom":{'num':0, 'list':[]}}
            
            # from MR_DICT get xmldom handle
            # mrs_temp_dom = gl.MR_DICT[time_entity][gl.MR_TYPE['MRS']]['xmldom']
            mrs_temp_dom = mr_utils.xml_dom(time_entity, 'MRS')
            
            # get xml documentElement
            mrs_root = mr_utils.xml_doc(mrs_temp_dom)
            
            for enb_entity in mr_utils.xml_element(mrs_root, 'eNB'):
                #TODO: enbid may use in the future, multi enodeb situation, now is no need
                enbid = int(mr_utils.xml_attr(enb_entity, 'id'))
                # analyze every mr_item
                for measurement_entity in mr_utils.xml_element(enb_entity, 'measurement'):
                    mrName = mr_utils.xml_attr(measurement_entity,'mrName')
                    object_list = mr_utils.xml_element(measurement_entity, 'object')
                    for mr_name_entity in out_dict[mrs_file_entity]:
                        if mr_name_entity == mrName:
                            for object_entity in object_list:
                                eci_id = int(mr_utils.xml_attr(object_entity, 'id').strip(':').split(':')[0])
                                #TODO: must check the eciid, standard eci = enbid*256 + cellid, but now there has some trouble, cause eci= short (enbid<<8 + cellid)
                                cellid_ret_list = mr_utils.is_cell_id_exist(eci_id)
                                
                                if  cellid_ret_list[0] == False:
                                    mr_utils.out_text_dict_append_list(output_temp_dict, mrs_file_entity, '[%s]-[cellid:%d not in list] '%(mrName, cellid_ret_list[1]))
                                else:
                                    if cellid_ret_list[1] not in out_dict[mrs_file_entity][mr_name_entity]['list']:
                                        out_dict[mrs_file_entity][mr_name_entity]['num'] += 1
                                        out_dict[mrs_file_entity][mr_name_entity]['list'].append(cellid_ret_list[1])
                            break
        except Exception as result:
            raise Exception('[%s]   %s |  %s'%(__name__ ,str(result.__traceback__.tb_lineno) if not gl.IS_PY2  else '',result))


    #==========================================================================
    #   write output 
    #==========================================================================
    cell_num = len(gl.TEST_CONF['cellid'].strip(',').split(','))
    gl.OUT_STR_LIST.append(date_time + " | ")
    gl.OUT_STR_LIST.append(gl.TEST_CONF['test_total_time'] + " | ")
    for file_name in out_dict:
        gl.OUT_STR_LIST.append('\n=====> ' + file_name + ': |')
        temp_flag_test = 0
        for mr_name in out_sort_list:
            cell_num_mrs_item = out_dict[file_name][mr_name]['num']
            gl.OUT_STR_LIST.append(str(cell_num) + " | ")
            if mr_utils.is_mr_item_need_exist(mr_name) and cell_num_mrs_item != cell_num:
                temp_flag_test = 1
                mr_utils.out_text_dict_append_list(output_temp_dict, file_name, '[%s]=<cell num not match>:[%d(mrs)]!=[%d(target)]'%(mr_name, cell_num_mrs_item, cell_num))

            if mr_utils.is_mr_item_need_exist(mr_name) == False and cell_num_mrs_item != 0:
                temp_flag_test = 1
                mr_utils.out_text_dict_append_list(output_temp_dict, file_name, '[{0}]=<MeasureItems not have {0}>:cellnum=[{1} ->target:0]'.format(mr_name, cell_num_mrs_item))

        gl.OUT_STR_LIST.append(str(temp_flag_test == 0))
        if output_temp_dict.__contains__(file_name) == True and len(output_temp_dict[file_name]) != 0:
            gl.OUT_STR_LIST.append('\n===> error:\n')
            for i in range(len(output_temp_dict[file_name])):
                gl.OUT_STR_LIST.append('\t\t' + output_temp_dict[file_name][i] + '\n')
        gl.OUT_STR_LIST.append("\n")
        
        
        
        
#========================================================================================#
#             test 53
# check the number of rip  <==>  actual rip number
#========================================================================================#

def test53_file_integrity():
    # mr_utils.test_out_data_item_header("test_53")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)
    
    # file_list_mrs = mr_utils.get_file_by_filter(gl.MR_TEST_PATH + '*MRS*.xml')
    text_consistent = {}
    subfram_dict = {}
    
    for time_entity in gl.MR_DICT:
    # for file_mrs_name in file_list_mrs:
        try:
            file_mrs_name = mr_utils.get_file_name(time_entity, 'MRS')
            if file_mrs_name == "":
                continue
            
            # initial subframe_dict, include { file_mrs_name : { cellid: { sub_frame: 0,  'consistent' : 0  }   }    }
            subfram_dict[file_mrs_name] = {}
            for cell_id in mr_utils.get_split_value(gl.TEST_CONF['cellid'], ',', [',', '}', '{']) :
                subfram_dict[file_mrs_name][cell_id] = {} 
                for sub_frame_key in mr_utils.get_split_value(gl.MR_CONF['SubFrameNum'], ',', [',', '}', '{']) :
                    subfram_dict[file_mrs_name][cell_id][sub_frame_key] = 0
                subfram_dict[file_mrs_name][cell_id]['consistent'] = 0
                
            # initial xml handle
            mrs_dom = mr_utils.xml_dom(time_entity, 'MRS')
            mrs_root = mr_utils.xml_doc(mrs_dom)
            
            # measurement analyse
            for measurement_entity in mr_utils.xml_element(mrs_root, 'measurement'):
                #TODO: should judge the ripprb though
                # mr_name -> rip
                if mr_utils.xml_attr(measurement_entity, 'mrName') == 'MR.ReceivedIPower' :
                    for object_entity in mr_utils.xml_element(measurement_entity, 'object'):
                        #TODO: eciid 
                        eci_id = int(  mr_utils.xml_attr(object_entity, 'id').strip(':').split(':')[0]   )
                        cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)
                        
                        subfram_id = object_entity.getAttribute('id').strip(':').split(':')[2]
                        test_count = 0
                        if str(cell_id_ret_list[1]) not in subfram_dict[file_mrs_name]:
                            mr_utils.out_text_dict_append_list(text_consistent, file_mrs_name, '[MR.ReceivedIPower]:<cell_id[%d] not in list>'%(cell_id_ret_list[1]))
                        else:
                            for subfram_key in subfram_dict[file_mrs_name][str(cell_id_ret_list[1])]:
                                if subfram_key == subfram_id:
                                    subfram_dict[file_mrs_name][str(cell_id_ret_list[1])][subfram_key] += 1
                                    test_count = 1
                                    break
                            if test_count == 0:
                                mr_utils.out_text_dict_append_list(text_consistent, file_mrs_name, '[MR.ReceivedIPower]:<surplus rip [cellid:%d]-[%s]>'%(cell_id_ret_list[1],subfram_id))
                                subfram_dict[file_mrs_name][str(cell_id_ret_list[1])]['consistent'] = 1
                    break
            for cell_id_entity in subfram_dict[file_mrs_name]:
                for subfram_entity in gl.MR_CONF['SubFrameNum'].strip(',').split(','):
                    if subfram_dict[file_mrs_name][cell_id_entity][subfram_entity] != 1 and mr_utils.is_mr_item_need_exist('MR.ReceivedIPower') == True:
                        mr_utils.out_text_dict_append_list(text_consistent, file_mrs_name, '[MR.ReceivedIPower]:<confusion rip [cellid:%s]-[%s]>'%(cell_id_entity,subfram_entity))
                        subfram_dict[file_mrs_name][cell_id_entity]['consistent'] = 1
        except Exception as result:
            raise Exception('[%s]   %s |  %s'%(__name__ ,str(result.__traceback__.tb_lineno) if not gl.IS_PY2  else '',result))

    #==========================================================================
    #   write output 
    #==========================================================================
    gl.OUT_STR_LIST.append(date_time + " | ")
    gl.OUT_STR_LIST.append(gl.TEST_CONF['test_total_time'] + " | \n")
    for mrs_file in subfram_dict:
        gl.OUT_STR_LIST.append('======>%s:'%(mrs_file))
        for cell_id_entity in subfram_dict[mrs_file]:
            gl.OUT_STR_LIST.append(' | cellid:[%s]-ripnum:[%d] | %s | \n'%(cell_id_entity, len(subfram_dict[mrs_file][cell_id_entity]) - 1,
                    str(subfram_dict[mrs_file][cell_id_entity]['consistent'] == 0)))
        if text_consistent.__contains__(mrs_file) == True and len(text_consistent[mrs_file]) != 0:
            gl.OUT_STR_LIST.append('====>error:\n')
            for i in range(len(text_consistent[mrs_file])):
                gl.OUT_STR_LIST.append(text_consistent[mrs_file][i] + '\n')
                

#========================================================================================#
#             test 54
# check the number of prb  <==>  actual prb number
#========================================================================================#               
                
def test54_file_integrity():
    # mr_utils.test_out_data_item_header("test_54")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    # get prb_list and rip_list
    prbnum_list = mr_utils.get_prb_list()
    ripnum_list = mr_utils.get_split_value( gl.MR_CONF['SubFrameNum'], ',', [',', '{', '}'] )

    ripprb_num = len( ripnum_list ) * len(prbnum_list)
    
    is_consistence = 0
    text_consistence = {}
    mrs_ripprb_dict = {}

    for time_entity in gl.MR_DICT:
        try:
            # get file_name as the key
            file_name = mr_utils.get_file_name(time_entity, 'MRS')
            if file_name == '':
                continue
            mrs_ripprb_dict[file_name] = {}
            
            mrs_dom = mr_utils.xml_dom(time_entity, 'MRS')
            mrs_root = mr_utils.xml_doc(mrs_dom)

            for measurement_entity in mr_utils.xml_element(mrs_root, 'measurement'):
                if mr_utils.xml_attr(measurement_entity, 'mrName') == 'MR.RIPPRB':
                    for object_entity in mr_utils.xml_element(measurement_entity, 'object'):
                        id_str_list =  mr_utils.get_split_value( mr_utils.xml_attr(object_entity, 'id'), ':', [' ', ':'] )
                        eci_id  =   int(id_str_list[0])
                        rip     =   id_str_list[2]
                        prb     =   id_str_list[3]
                        #TODO: eciid
                        cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)

                        # judge the prb num with rip
                        if mrs_ripprb_dict[file_name].__contains__(str(cell_id_ret_list[1])) == False:
                            mrs_ripprb_dict[file_name][str(cell_id_ret_list[1])] = []
                        if rip in ripnum_list and prb in prbnum_list:
                            mrs_ripprb_dict[file_name][str(cell_id_ret_list[1])].append(rip + ':' + prb)
                        else:
                            mr_utils.out_text_dict_append_list(text_consistence, file_name, '<ripprb not match>:[%s:%s]'%(rip,prb))
                    #just check ripprb
                    break 
        except Exception as result:
            raise Exception('[%s]   %s |  %s'%(__name__ ,str(result.__traceback__.tb_lineno) if not gl.IS_PY2  else '',result))

    #==========================================================================
    #   write output 
    #==========================================================================
    gl.OUT_STR_LIST.append(date_time + " | ")
    gl.OUT_STR_LIST.append(gl.TEST_CONF['test_total_time'] + " | \n")
    for file_name in mrs_ripprb_dict:
        gl.OUT_STR_LIST.append('====>%s: \n'%(file_name))
        for cell_id_str in mrs_ripprb_dict[file_name]:
            ripnum = len(mrs_ripprb_dict[file_name][cell_id_str])
            gl.OUT_STR_LIST.append('\t\t<[cell:%s]:[%d] | %s>\n'%(cell_id_str, ripnum, ripnum == ripprb_num ))
        if text_consistence.__contains__(file_name) == True:
            gl.OUT_STR_LIST.append('\terror:\n')
            for i in range(len(text_consistence[file_name])):
                gl.OUT_STR_LIST.append(text_consistence[file_name][i] + '\n')
                

#========================================================================================#
#             test 55
# check the number of rsrp,rsrq,phr  ,  is MRO == MRS
#========================================================================================#   

def test55_file_integrity():
    # mr_utils.test_out_data_item_header("test_55")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)
    
    # MRO & MRS item 
    out_list_mro = ['MR.LteScRSRP', 'MR.LteScRSRQ', 'MR.LteScPHR']
    out_list_mrs = ['MR.RSRP', 'MR.RSRQ', 'MR.PowerHeadRoom']

    for time_entity in gl.MR_DICT:
        # every period: MRO count & MRS count
        mro_count_dict = {'MR.LteScRSRP':{}, 'MR.LteScRSRQ':{}, 'MR.LteScPHR':{}}
        mrs_count_dict = {'MR.RSRP':{}, 'MR.RSRQ':{}, 'MR.PowerHeadRoom':{}}
        
        # MRO : the value_list position:  <v> ...2 3 4 ... </v> point to rsrp, rsrq,phr
        # pos_dict = {'MR.LteScRSRP': {'pos':2}, 'MR.LteScRSRQ': {'pos':3}, 'MR.LteScPHR':{'pos':4}}
        pos_dict = {}
        try:
            
            mro_measurement_list = []
            mrs_measurement_list = []
            mro_file_name = mr_utils.get_file_name(time_entity, 'MRO')
            mrs_file_name = mr_utils.get_file_name(time_entity, 'MRS')

            if 'MRO' in gl.MR_CONF['MeasureType'] and mro_file_name != '' :
                mro_dom     = mr_utils.xml_dom(time_entity, 'MRO')
                mro_root    = mr_utils.xml_doc(mro_dom)
                mro_measurement_list = mr_utils.xml_element(mro_root, 'measurement')

            if 'MRS' in gl.MR_CONF['MeasureType'] and mrs_file_name != '':
                mrs_dom     = mr_utils.xml_dom(time_entity, 'MRS')
                mrs_root    = mr_utils.xml_doc(mrs_dom)
                mrs_measurement_list = mr_utils.xml_element(mrs_root, 'measurement')

            # if no mro file, not excecute this
            for mro_measurement_entity in mro_measurement_list:
                smr_entity = mr_utils.xml_element(mro_measurement_entity, 'smr')[0]
                smr_str = mr_utils.xml_text(smr_entity)
                
                # if pos_dict's position is wrong , the function is useful, or same result
                mr_utils.get_mr_item_pos(pos_dict, smr_str)
                
                for object_entity in mr_utils.xml_element(mro_measurement_entity, 'object'):
                    eci_id = int( mr_utils.xml_attr(object_entity, 'id') )
                    #TODO: eciid
                    cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)
                    cell_id = str(cell_id_ret_list[1])
                    # use cellid + ueid  as key
                    # MmeUeS1apId = mr_utils.xml_attr(object_entity, 'MmeUeS1apId')
                    cell_mmeUe = cell_id # + '_'  + MmeUeS1apId

                    value_str = mr_utils.xml_text(mr_utils.xml_element(object_entity, 'v')[0])
                    value_list = mr_utils.get_split_value(value_str, ' ')
                    
                    for i in range(len(value_list)):
                        mro_mr_Name = pos_dict[smr_str.strip(" ")][i]
                        if mro_mr_Name in mro_count_dict:
                            if cell_mmeUe not in mro_count_dict[mro_mr_Name]:
                                mro_count_dict[mro_mr_Name][cell_mmeUe] = 0
                            if str(value_list[i]).isdigit() == True: 
                                mro_count_dict[mro_mr_Name][cell_mmeUe] += 1
                # rsrp, rsrq, phr is UE data, don't need to judge rip
                break
            for mrs_measurement_entity in mrs_measurement_list:
                for mrs_mr_Name in mrs_count_dict:
                    if mr_utils.xml_attr(mrs_measurement_entity, 'mrName') == mrs_mr_Name:
                        for object_entity in mr_utils.xml_element(mrs_measurement_entity, 'object'):
                            #TODO: eciid
                            eci_id = int(mr_utils.xml_attr(object_entity, 'id'))
                            cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)
                            cell_id = str(cell_id_ret_list[1])
                            temp_count = 0
                            for value_entity in object_entity.getElementsByTagName('v'):
                                temp_count += mr_utils.add_digital_string(mr_utils.xml_text(value_entity))

                            mrs_count_dict[mrs_mr_Name][cell_id] = temp_count

            #==========================================================================
            #   write output 
            #==========================================================================
            gl.OUT_STR_LIST.append(time_entity + ' : \n')
            gl.OUT_STR_LIST.append(date_time + " | ")
            gl.OUT_STR_LIST.append(gl.TEST_CONF['test_total_time'] + " | ")
            
            gl.OUT_STR_LIST.append('[MRO文件上报采样点数]' + " | ")
            for mr_name in out_list_mro:
                gl.OUT_STR_LIST.append('{')
                for cell_id in mro_count_dict[mr_name]:
                    gl.OUT_STR_LIST.append(' [%s]=%s '%(cell_id, str(mro_count_dict[mr_name][cell_id])) )
                gl.OUT_STR_LIST.append('} | ')
            gl.OUT_STR_LIST.append('\n' + " " * (len(date_time) + len(gl.TEST_CONF['test_total_time']) + 6))

            gl.OUT_STR_LIST.append('[MRS文件统计采样点数]' + " | ")
            for mr_name in out_list_mrs:
                gl.OUT_STR_LIST.append('{')
                for cell_id in mrs_count_dict[mr_name]:
                    gl.OUT_STR_LIST.append(' [%s]=%s '%(cell_id, str(mrs_count_dict[mr_name][cell_id])))
                gl.OUT_STR_LIST.append('} | ')
            gl.OUT_STR_LIST.append('\n' + " " * (len(date_time) + len(gl.TEST_CONF['test_total_time']) + 6))

            gl.OUT_STR_LIST.append('[统计采样点数是否完整]' + " | ")
            gl.OUT_STR_LIST.append(str(mro_count_dict['MR.LteScRSRP'] == mrs_count_dict['MR.RSRP'] or \
                                      mr_utils.is_mr_item_need_exist('MR.LteScRSRP') == False or \
                                      mr_utils.is_mr_item_need_exist('MR.RSRP') == False) + ' | ')
            gl.OUT_STR_LIST.append(str(mro_count_dict['MR.LteScRSRQ'] == mrs_count_dict['MR.RSRQ'] or \
                                      mr_utils.is_mr_item_need_exist('MR.LteScRSRQ') == False or \
                                      mr_utils.is_mr_item_need_exist('MR.RSRQ') == False) + ' | ')
            gl.OUT_STR_LIST.append(str(mro_count_dict['MR.LteScPHR'] == mrs_count_dict['MR.PowerHeadRoom'] or \
                                      mr_utils.is_mr_item_need_exist('MR.LteScPHR') == False or \
                                      mr_utils.is_mr_item_need_exist('MR.PowerHeadRoom') == False) + ' | \n')

        except Exception as result:
            raise Exception('[%s]   %s |  %s'%(__name__ ,str(result.__traceback__.tb_lineno) if not gl.IS_PY2  else '',result))



#========================================================================================#
#             test 56
# check the number of rip, sinrul  ,  is MRO == MRS
#========================================================================================#   

def test56_file_integrity():
    # mr_utils.test_out_data_item_header("test_56")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    out_mro_list = ['MR.LteScRIP', 'MR.LteScSinrUL']
    out_mrs_list = ['MR.ReceivedIPower', 'MR.SinrUL']
    
    for time_entity in gl.MR_DICT:
        try:
            mro_count_dict = {'MR.LteScRIP':{}, 'MR.LteScSinrUL':{}}
            mrs_count_dict = {'MR.ReceivedIPower':{}, 'MR.SinrUL':{}}
            
            sinrul_pos_dict = {'MR.LteScSinrUL':{'pos':5}}
            
            mro_measurement_list = []
            mrs_measurement_list = []
            
            mro_file_name = mr_utils.get_file_name(time_entity, 'MRO')
            mrs_file_name = mr_utils.get_file_name(time_entity, 'MRS')
            
            if 'MRO' in gl.MR_CONF['MeasureType'] and mro_file_name != '' :
                mro_dom     = mr_utils.xml_dom(time_entity, 'MRO')
                mro_root    = mr_utils.xml_doc(mro_dom)
                mro_measurement_list = mr_utils.xml_element(mro_root, 'measurement')
            if 'MRS' in  gl.MR_CONF['MeasureType'] and mrs_file_name != '':
                mrs_dom     = mr_utils.xml_dom(time_entity, 'MRS')
                mrs_root    = mr_utils.xml_doc(mrs_dom) 
                mrs_measurement_list = mr_utils.xml_element(mrs_root, 'measurement')

            for mr_name in out_mro_list:
                for mro_measurement_entity in mro_measurement_list:
                    for smr_entity in mr_utils.xml_element(mro_measurement_entity, 'smr'):                 
                        if mr_name in mr_utils.get_split_value( mr_utils.xml_text(smr_entity) ):
                            for object_entity in mr_utils.xml_element(mro_measurement_entity, 'object'):
                                # get object attribute: id 
                                object_value = mr_utils.xml_attr(object_entity, 'id')
                                object_value_list = mr_utils.get_split_value(object_value, ':', [":"])
                                #TODO: eciid
                                eci_id = int(  object_value_list[0] )
                                cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)
                                object_id = str(cell_id_ret_list[1])
                                
                                object_prb = ""
                                # rip num -> key distinction
                                if mr_name == 'MR.LteScRIP':
                                    object_prb = object_value_list[2]   

                                if mro_count_dict[mr_name].__contains__(object_id) == False :
                                    if mr_name == 'MR.LteScRIP':
                                        mro_count_dict[mr_name][object_id] = {}
                                    else:
                                        mro_count_dict[mr_name][object_id] = 0

                                if mr_name == 'MR.LteScRIP' and mro_count_dict[mr_name][object_id].__contains__(object_prb) == False:
                                    mro_count_dict[mr_name][object_id][object_prb] = 0
                                
                                for value_entity in object_entity.getElementsByTagName('v'):
                                    if mr_name == 'MR.LteScRIP' and mr_utils.xml_text(value_entity).isdigit() == True:
                                        mro_count_dict[mr_name][object_id][object_prb] += 1
                                    elif mr_name == 'MR.LteScSinrUL' :
                                        pos = mr_utils.get_mro_pos_by_smr('MR.LteScSinrUL', mr_utils.xml_text(smr_entity))
                                        if  mr_utils.get_split_value(mr_utils.xml_text(value_entity))[pos].isdigit() == True:
                                            mro_count_dict[mr_name][object_id] += 1
                
            for mr_name in out_mrs_list:
                for measurement_entity in mrs_measurement_list:
                    if mr_name == mr_utils.xml_attr(measurement_entity, 'mrName'):
                        for object_entity in mr_utils.xml_element(measurement_entity, 'object'):
                            object_value_list = mr_utils.get_split_value(mr_utils.xml_attr(object_entity, 'id'), ':', [':'])
                            eci_id = int( object_value_list[0])
                            cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)
                            object_id = str(cell_id_ret_list[1])
                            object_prb = ''
                            
                            if mr_name == 'MR.ReceivedIPower':
                                object_prb = object_value_list[2]

                            temp_num = 0
                            for value_entity in mr_utils.xml_element(object_entity, 'v'):
                                temp_num += mr_utils.add_digital_string(mr_utils.xml_text(value_entity))

                            if mr_name == 'MR.ReceivedIPower':
                                if mrs_count_dict[mr_name].__contains__(object_id) == False:
                                    mrs_count_dict[mr_name][object_id] = {}
                                if mrs_count_dict[mr_name][object_id].__contains__(object_prb) == False:
                                    mrs_count_dict[mr_name][object_id][object_prb] = 0
                                mrs_count_dict[mr_name][object_id][object_prb] = temp_num
                            elif mr_name == 'MR.SinrUL':
                                if mrs_count_dict[mr_name].__contains__(object_id) == False:
                                    mrs_count_dict[mr_name][object_id] = 0
                                mrs_count_dict[mr_name][object_id] = temp_num

            #==========================================================================
            #   write output 
            #==========================================================================
            gl.OUT_STR_LIST.append(time_entity + ' : \n')
            gl.OUT_STR_LIST.append(date_time + " | ")
            gl.OUT_STR_LIST.append(gl.TEST_CONF['test_total_time'] + " | ")

            gl.OUT_STR_LIST.append('[MRO文件上报采样点数]' + " | ")
            for mr_name in out_mro_list:
                for object_id in mro_count_dict[mr_name]:
                    if mr_name == 'MR.LteScRIP':
                        gl.OUT_STR_LIST.append('[' + object_id + ']={' )
                        for object_prb in mro_count_dict[mr_name][object_id]:
                            gl.OUT_STR_LIST.append('[' + object_prb + ']=' + str(mro_count_dict[mr_name][object_id][object_prb]) + ' ')
                        gl.OUT_STR_LIST.append('} | ')
                    if mr_name == 'MR.LteScSinrUL':
                        gl.OUT_STR_LIST.append('[' + object_id + ']=' + str(mro_count_dict[mr_name][object_id]) + ' | ')

            gl.OUT_STR_LIST.append('\n' + " " * (len(date_time) + len(gl.TEST_CONF['test_total_time']) + 6))

            gl.OUT_STR_LIST.append('[MRS文件统计采样点数]' + " | ")
            for mr_name in out_mrs_list:
                for object_id in mrs_count_dict[mr_name]:
                    if mr_name == 'MR.ReceivedIPower':
                        gl.OUT_STR_LIST.append('[' + object_id + ']={' )
                        for object_prb in mrs_count_dict[mr_name][object_id]:
                            gl.OUT_STR_LIST.append('[' + object_prb + ']=' + str(mrs_count_dict[mr_name][object_id][object_prb]) + ' ')
                        gl.OUT_STR_LIST.append('} | ')
                    if mr_name == 'MR.SinrUL':
                        gl.OUT_STR_LIST.append('[' + object_id + ']=' + str(mrs_count_dict[mr_name][object_id]) + ' | ')

            gl.OUT_STR_LIST.append('\n' + " " * (len(date_time) + len(gl.TEST_CONF['test_total_time']) + 6))

            test_flag = 0
            for object_id in mrs_count_dict['MR.ReceivedIPower']:
                for object_prb in mrs_count_dict['MR.ReceivedIPower'][object_id]:
                    if mro_count_dict['MR.LteScRIP'].__contains__(object_id) == False:
                        continue
                    if mro_count_dict['MR.LteScRIP'][object_id].__contains__(object_prb) == True and mro_count_dict['MR.LteScRIP'][object_id][object_prb] != mrs_count_dict['MR.ReceivedIPower'][object_id][object_prb]:
                        test_flag = 1
            gl.OUT_STR_LIST.append('[统计采样点数是否完整]' + " | ")
            gl.OUT_STR_LIST.append(str(0 == test_flag) + ' | ')
            gl.OUT_STR_LIST.append(str(mro_count_dict['MR.LteScSinrUL'] == mrs_count_dict['MR.SinrUL'] or\
                                      mr_utils.is_mr_item_need_exist('MR.LteScSinrUL') == False or \
                                      mr_utils.is_mr_item_need_exist('MR.SinrUL') == False) + ' | \n')
        except Exception as result:
            raise Exception('[%s]   %s |  %s'%(__name__ ,str(result.__traceback__.tb_lineno) if not gl.IS_PY2  else '',result))

#========================================================================================#
#             test 57
# check the normativity of MRS
#========================================================================================#   

def test57_file_integrity():
    # mr_utils.test_out_data_item_header("test_57")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    #the excel row
    sheet_row_dict = {'MR.RSRP':[3,49], 'MR.RSRQ':[51,67],'MR.PowerHeadRoom':[114,176], 'MR.ReceivedIPower':[178,229],'MR.SinrUL':[807,842], 'MR.RIPPRB':[844, 895]}

    #the standard string from doc
    rsrp_smr = 'MR.RSRP.00 MR.RSRP.01 MR.RSRP.02 MR.RSRP.03 MR.RSRP.04 MR.RSRP.05 MR.RSRP.06 MR.RSRP.07 MR.RSRP.08 MR.RSRP.09 MR.RSRP.10 MR.RSRP.11 MR.RSRP.12 MR.RSRP.13 MR.RSRP.14 MR.RSRP.15 MR.RSRP.16 MR.RSRP.17 MR.RSRP.18 MR.RSRP.19 MR.RSRP.20 MR.RSRP.21 MR.RSRP.22 MR.RSRP.23 MR.RSRP.24 MR.RSRP.25 MR.RSRP.26 MR.RSRP.27 MR.RSRP.28 MR.RSRP.29 MR.RSRP.30 MR.RSRP.31 MR.RSRP.32 MR.RSRP.33 MR.RSRP.34 MR.RSRP.35 MR.RSRP.36 MR.RSRP.37 MR.RSRP.38 MR.RSRP.39 MR.RSRP.40 MR.RSRP.41 MR.RSRP.42 MR.RSRP.43 MR.RSRP.44 MR.RSRP.45 MR.RSRP.46 MR.RSRP.47 '
    rsrq_smr = 'MR.RSRQ.00 MR.RSRQ.01 MR.RSRQ.02 MR.RSRQ.03 MR.RSRQ.04 MR.RSRQ.05 MR.RSRQ.06 MR.RSRQ.07 MR.RSRQ.08 MR.RSRQ.09 MR.RSRQ.10 MR.RSRQ.11 MR.RSRQ.12 MR.RSRQ.13 MR.RSRQ.14 MR.RSRQ.15 MR.RSRQ.16 MR.RSRQ.17 '
    rip_smr = 'MR.ReceivedIPower.00 MR.ReceivedIPower.01 MR.ReceivedIPower.02 MR.ReceivedIPower.03 MR.ReceivedIPower.04 MR.ReceivedIPower.05 MR.ReceivedIPower.06 MR.ReceivedIPower.07 MR.ReceivedIPower.08 MR.ReceivedIPower.09 MR.ReceivedIPower.10 MR.ReceivedIPower.11 MR.ReceivedIPower.12 MR.ReceivedIPower.13 MR.ReceivedIPower.14 MR.ReceivedIPower.15 MR.ReceivedIPower.16 MR.ReceivedIPower.17 MR.ReceivedIPower.18 MR.ReceivedIPower.19 MR.ReceivedIPower.20 MR.ReceivedIPower.21 MR.ReceivedIPower.22 MR.ReceivedIPower.23 MR.ReceivedIPower.24 MR.ReceivedIPower.25 MR.ReceivedIPower.26 MR.ReceivedIPower.27 MR.ReceivedIPower.28 MR.ReceivedIPower.29 MR.ReceivedIPower.30 MR.ReceivedIPower.31 MR.ReceivedIPower.32 MR.ReceivedIPower.33 MR.ReceivedIPower.34 MR.ReceivedIPower.35 MR.ReceivedIPower.36 MR.ReceivedIPower.37 MR.ReceivedIPower.38 MR.ReceivedIPower.39 MR.ReceivedIPower.40 MR.ReceivedIPower.41 MR.ReceivedIPower.42 MR.ReceivedIPower.43 MR.ReceivedIPower.44 MR.ReceivedIPower.45 MR.ReceivedIPower.46 MR.ReceivedIPower.47 MR.ReceivedIPower.48 MR.ReceivedIPower.49 MR.ReceivedIPower.50 MR.ReceivedIPower.51 MR.ReceivedIPower.52 '
    ripprb_smr = 'MR.RIPPRB.00 MR.RIPPRB.01 MR.RIPPRB.02 MR.RIPPRB.03 MR.RIPPRB.04 MR.RIPPRB.05 MR.RIPPRB.06 MR.RIPPRB.07 MR.RIPPRB.08 MR.RIPPRB.09 MR.RIPPRB.10 MR.RIPPRB.11 MR.RIPPRB.12 MR.RIPPRB.13 MR.RIPPRB.14 MR.RIPPRB.15 MR.RIPPRB.16 MR.RIPPRB.17 MR.RIPPRB.18 MR.RIPPRB.19 MR.RIPPRB.20 MR.RIPPRB.21 MR.RIPPRB.22 MR.RIPPRB.23 MR.RIPPRB.24 MR.RIPPRB.25 MR.RIPPRB.26 MR.RIPPRB.27 MR.RIPPRB.28 MR.RIPPRB.29 MR.RIPPRB.30 MR.RIPPRB.31 MR.RIPPRB.32 MR.RIPPRB.33 MR.RIPPRB.34 MR.RIPPRB.35 MR.RIPPRB.36 MR.RIPPRB.37 MR.RIPPRB.38 MR.RIPPRB.39 MR.RIPPRB.40 MR.RIPPRB.41 MR.RIPPRB.42 MR.RIPPRB.43 MR.RIPPRB.44 MR.RIPPRB.45 MR.RIPPRB.46 MR.RIPPRB.47 MR.RIPPRB.48 MR.RIPPRB.49 MR.RIPPRB.50 MR.RIPPRB.51 MR.RIPPRB.52 '
    phr_smr = 'MR.PowerHeadRoom.00 MR.PowerHeadRoom.01 MR.PowerHeadRoom.02 MR.PowerHeadRoom.03 MR.PowerHeadRoom.04 MR.PowerHeadRoom.05 MR.PowerHeadRoom.06 MR.PowerHeadRoom.07 MR.PowerHeadRoom.08 MR.PowerHeadRoom.09 MR.PowerHeadRoom.10 MR.PowerHeadRoom.11 MR.PowerHeadRoom.12 MR.PowerHeadRoom.13 MR.PowerHeadRoom.14 MR.PowerHeadRoom.15 MR.PowerHeadRoom.16 MR.PowerHeadRoom.17 MR.PowerHeadRoom.18 MR.PowerHeadRoom.19 MR.PowerHeadRoom.20 MR.PowerHeadRoom.21 MR.PowerHeadRoom.22 MR.PowerHeadRoom.23 MR.PowerHeadRoom.24 MR.PowerHeadRoom.25 MR.PowerHeadRoom.26 MR.PowerHeadRoom.27 MR.PowerHeadRoom.28 MR.PowerHeadRoom.29 MR.PowerHeadRoom.30 MR.PowerHeadRoom.31 MR.PowerHeadRoom.32 MR.PowerHeadRoom.33 MR.PowerHeadRoom.34 MR.PowerHeadRoom.35 MR.PowerHeadRoom.36 MR.PowerHeadRoom.37 MR.PowerHeadRoom.38 MR.PowerHeadRoom.39 MR.PowerHeadRoom.40 MR.PowerHeadRoom.41 MR.PowerHeadRoom.42 MR.PowerHeadRoom.43 MR.PowerHeadRoom.44 MR.PowerHeadRoom.45 MR.PowerHeadRoom.46 MR.PowerHeadRoom.47 MR.PowerHeadRoom.48 MR.PowerHeadRoom.49 MR.PowerHeadRoom.50 MR.PowerHeadRoom.51 MR.PowerHeadRoom.52 MR.PowerHeadRoom.53 MR.PowerHeadRoom.54 MR.PowerHeadRoom.55 MR.PowerHeadRoom.56 MR.PowerHeadRoom.57 MR.PowerHeadRoom.58 MR.PowerHeadRoom.59 MR.PowerHeadRoom.60 MR.PowerHeadRoom.61 MR.PowerHeadRoom.62 MR.PowerHeadRoom.63 '
    sinrul_smr = 'MR.SinrUL.00 MR.SinrUL.01 MR.SinrUL.02 MR.SinrUL.03 MR.SinrUL.04 MR.SinrUL.05 MR.SinrUL.06 MR.SinrUL.07 MR.SinrUL.08 MR.SinrUL.09 MR.SinrUL.10 MR.SinrUL.11 MR.SinrUL.12 MR.SinrUL.13 MR.SinrUL.14 MR.SinrUL.15 MR.SinrUL.16 MR.SinrUL.17 MR.SinrUL.18 MR.SinrUL.19 MR.SinrUL.20 MR.SinrUL.21 MR.SinrUL.22 MR.SinrUL.23 MR.SinrUL.24 MR.SinrUL.25 MR.SinrUL.26 MR.SinrUL.27 MR.SinrUL.28 MR.SinrUL.29 MR.SinrUL.30 MR.SinrUL.31 MR.SinrUL.32 MR.SinrUL.33 MR.SinrUL.34 MR.SinrUL.35 MR.SinrUL.36'

    format_standard_dict = {'MR.RSRP':{'smr':rsrp_smr, 'v':48}, 'MR.RSRQ':{'smr':rsrq_smr, 'v':18}, 'MR.ReceivedIPower':{'smr':rip_smr, 'v':53}, 'MR.RIPPRB':{'smr':ripprb_smr,'v':53}, 'MR.SinrUL':{'smr':sinrul_smr, 'v':37}, 'MR.PowerHeadRoom':{'smr':phr_smr, 'v':64}}
    out_text_list = {}
    
    try:
        
        out_mrs_flag_dict = {'MR.RSRP':3, 'MR.RSRQ':3, 'MR.ReceivedIPower':3, 'MR.RIPPRB':3, 'MR.SinrUL':3, 'MR.PowerHeadRoom':3 }

        for time_entity in gl.MR_DICT:
            # flag to judge accuracy
            out_mrs_flag_dict = {'MR.RSRP':3, 'MR.RSRQ':3, 'MR.ReceivedIPower':3, 'MR.RIPPRB':3, 'MR.SinrUL':3, 'MR.PowerHeadRoom':3 }
            file_name = mr_utils.get_file_name(time_entity, 'MRS')
            if file_name == '':
                continue
            
            mrs_dom = mr_utils.xml_dom(time_entity, 'MRS')
            mrs_root = mr_utils.xml_doc(mrs_dom)
            
            for measurement_entity in mr_utils.xml_element(mrs_root, 'measurement'):
                for standard_dict_key in format_standard_dict:
                    if standard_dict_key == mr_utils.xml_attr(measurement_entity, 'mrName'):
                        #mrName匹配成功
                        out_mrs_flag_dict[standard_dict_key] -= 1
                        
                        #smr数据正确是否
                        smr_list = mr_utils.xml_element(measurement_entity, 'smr')
                        if len(smr_list) == 1 and mr_utils.xml_text(smr_list[0]).strip() == format_standard_dict[standard_dict_key]['smr'].strip():
                            out_mrs_flag_dict[standard_dict_key] -= 1
                            
                        #value对应的个数正确是否
                        test_value_num_flag = 0
                        for object_entity in mr_utils.xml_element(measurement_entity, 'object'):
                            for value_entity in mr_utils.xml_element(object_entity, 'v'):
                                if format_standard_dict[standard_dict_key]['v'] != len( mr_utils.get_split_value(mr_utils.xml_text(value_entity)) ):
                                    test_value_num_flag = 1
                        if test_value_num_flag == 0:
                            out_mrs_flag_dict[standard_dict_key] -= 1
            
            #判断mrs文件输出是否有问题
            for mrs_dict_key in out_mrs_flag_dict:
                if out_mrs_flag_dict[mrs_dict_key] != 0 and mr_utils.is_mr_item_need_exist(mrs_dict_key) == True:
                    mr_utils.out_text_dict_append_list(out_text_list, file_name, '-[%s not match]' % (mrs_dict_key))
                elif out_mrs_flag_dict[mrs_dict_key] != 3 and mr_utils.is_mr_item_need_exist(mrs_dict_key) == False:
                    mr_utils.out_text_dict_append_list(out_text_list, file_name, '-[%s not match]' % (mrs_dict_key))
            if out_text_list.__contains__(file_name) == True:
                break


        full_name = os.path.join(gl.SOURCE_PATH, gl.XLS_NAME)
        if int(gl.TEST_CONF['is_57_out_excel']) == 1 and os.path.exists(full_name):
            excel_workbook = openpyxl.load_workbook(full_name)
            work_sheet = excel_workbook.worksheets[0]
            # sheet_rows = work_sheet.max_row

            for mrs_dict_key in out_mrs_flag_dict:
                if out_mrs_flag_dict[mrs_dict_key] != 0:
                    for i in range(sheet_row_dict[mrs_dict_key][0], sheet_row_dict[mrs_dict_key][1]+1):
                        work_sheet.cell(i, 20, 'N')
                else:
                    for i in range(sheet_row_dict[mrs_dict_key][0], sheet_row_dict[mrs_dict_key][1]+1):
                        work_sheet.cell(i, 20, 'Y')
            excel_workbook.save(filename=os.path.join(gl.OUTPUT_PATH, gl.XLS_NAME))

        #==========================================================================
        #   write output 
        #==========================================================================
        gl.OUT_STR_LIST.append(date_time + " | ")
        gl.OUT_STR_LIST.append(gl.TEST_CONF['test_total_time'] + " | " )
        

        if len(out_text_list) != 0:
            gl.OUT_STR_LIST.append('\nerror:\n')
            for file_name in out_text_list:
                gl.OUT_STR_LIST.append('======> %s:'%(file_name))
                for i in range(len(out_text_list[file_name])):
                    gl.OUT_STR_LIST.append('\t%s\n' % (out_text_list[file_name][i]))
        else:
            gl.OUT_STR_LIST.append('successful\n')
        if int(gl.TEST_CONF['is_57_out_excel']) == 1 and os.path.exists(full_name): 
            gl.OUT_STR_LIST.append('结果已写入->'  )
            gl.OUT_STR_LIST.append(gl.XLS_NAME)
        gl.OUT_STR_LIST.append('\n')
    except Exception as result:
        raise Exception('[%s]   %s |  %s'%(__name__ ,str(result.__traceback__.tb_lineno) if not gl.IS_PY2  else '',result))
    
    
    
#========================================================================================#
#             test 58
# check the capacity of sample of MRO
#========================================================================================#     
    
def test58_file_integrity():
    # mr_utils.test_out_data_item_header("test_58")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)


    cellid = int(gl.TEST_CONF['cellid'])
    ideal_sample_num = int(int(gl.MR_CONF['UploadPeriod']) * 60 * 1000 / int(gl.MR_CONF['SamplePeriod']))

    # test_cell_id = int(gl.TEST_CONF['cellid']) | int(gl.TEST_CONF['cellid'])  << 8
    out_mro_list = ['MR.LteScRSRP','MR.LteScRSRQ','MR.LteScPHR','MR.LteScSinrUL']
    out_text_list = {}
    # excel position
    xml_mro_item_dict = {"MR.LteScRSRP":2,'MR.LteNcRSRP':3,'MR.LteScRSRQ':4,'MR.LteNcRSRQ':5,'MR.LteScPHR':6, 'MR.LteScRIP':7, 'MR.LteScSinrUL':8 }
    temp_loop_flag = 0
    is_object_empty = 0

    # first add 
    gl.OUT_STR_LIST.append(date_time + " | ")
    gl.OUT_STR_LIST.append(gl.TEST_CONF['test_total_time'] + " | " )

    for time_entity in gl.MR_DICT:
        try:
            temp_test_flag_dict = {'MR.LteScRSRP':0,'MR.LteScRSRQ':0,'MR.LteScPHR':0,'MR.LteScSinrUL':0,'MR.LteScRIP':0, 'MR.LteNcRSRP':0, 'MR.LteNcRSRQ':0}
            mro_file = mr_utils.get_file_name(time_entity, 'MRO')
            if mro_file == '':
                continue
            mro_dom = mr_utils.xml_dom(time_entity, 'MRO')
            mro_root = mr_utils.xml_doc(mro_dom)
            mro_item_dict = {}

            for enb_entity in mr_utils.xml_element(mro_root, 'eNB'):
                # if mr_utils.is_enb_id_exist(enb_entity.getAttribute('id')) == True:
                for measurement_entity in mr_utils.xml_element(enb_entity, 'measurement'):
                    if len(mr_utils.xml_element(measurement_entity, 'object')) == 0:
                        is_object_empty = 1
                        smr_value = mr_utils.xml_text(mr_utils.xml_element(measurement_entity, 'smr')[0])
                        if 'MR.LteScRIP' not in  smr_value and\
                                    mr_utils.is_mr_item_need_exist('MR.LteScRSRP') == True and\
                                    mr_utils.is_mr_item_need_exist('MR.LteScRSRQ') == True and\
                                    mr_utils.is_mr_item_need_exist('MR.LteScPHR') == True and\
                                    mr_utils.is_mr_item_need_exist('MR.LteScSinrUL') == True :
                            mr_utils.out_text_dict_append_list(out_text_list, mro_file, 'L3 data None')
                        if 'MR.LteScRIP' in smr_value  and  mr_utils.is_mr_item_need_exist('MR.LteScRIP') == True:
                            mr_utils.out_text_dict_append_list(out_text_list, mro_file, 'L2 data None')
                        break
                    for smr_entity in mr_utils.xml_element(measurement_entity, 'smr'):
                        smr_value_str = mr_utils.xml_text(smr_entity)
                        for object_entity in mr_utils.xml_element(measurement_entity, 'object'):
                            eci_id = int(mr_utils.get_split_value(mr_utils.xml_attr(object_entity, 'id'), ':', [':', ' '])[0])
                            #TODO: eciid
                            cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)
                            
                            object_id = str(cell_id_ret_list[1])
                            object_ue_id = object_id + "_" + mr_utils.xml_attr(object_entity, 'MmeUeS1apId')
                            if cell_id_ret_list[0] == False:
                                continue
                            if mro_item_dict.__contains__(object_ue_id) == False  :
                                mro_item_dict[object_ue_id] = {
                                    'MR.LteScRSRP':     {'pos':2, 'TimeStamp':0, 'range':[0, 97],   'flag':3,   'num':0, 'item_num':{}          }, 
                                    'MR.LteScRSRQ':     {'pos':3, 'TimeStamp':0, 'range':[0,34],    'flag':3,   'num':0, 'item_num':{}          },
                                    'MR.LteScPHR':      {'pos':4, 'TimeStamp':0, 'range':[0,63],    'flag':3,   'num':0, 'item_num':{}          }, 
                                    'MR.LteScSinrUL':   {'pos':5, 'TimeStamp':0, 'range':[0,36],    'flag':3,   'num':0, 'item_num':{}          },
                                    'MR.LteScRIP':      {'pos':0, 'TimeStamp':0, 'range':[0,511],   'flag':3,   'num':0, 'prbnum':{'item_num':{}}}, 
                                    'MR.LteNcRSRP':     {'pos':8, 'TimeStamp':0, 'range':[0,97],    'flag':3,   'num':0, 'item_num':{}          },
                                    'MR.LteNcRSRQ':     {'pos':9, 'TimeStamp':0, 'range':[0,34],    'flag':3,   'num':0, 'item_num':{}          } }
                            # phr ,sinrul, rip as the one measurement, so check for every smr
                            # now phr,sinrul add to ue data, can simplify this code 
                            mr_utils.get_mr_item_pos_dict(mro_item_dict[object_ue_id], smr_value_str)
                            for mr_name_entity in mro_item_dict[object_ue_id]:
                                #mr_Name匹配上, flag-1
                                if mr_name_entity not in smr_value_str:
                                    continue
                                mr_item_pos = mro_item_dict[object_ue_id][mr_name_entity]['pos']
                                if mr_name_entity == mr_utils.get_split_value(smr_value_str)[mr_item_pos]:
                                    mro_item_dict[object_ue_id][mr_name_entity]['flag'] -= 1
                                    if mr_name_entity != 'MR.LteScRIP':
                                        #判断 时间戳是否满足 MR测量的 SamplePeriod
                                        time_stamp_str = mr_utils.xml_attr(object_entity, 'TimeStamp')
                                        if mro_item_dict[object_ue_id][mr_name_entity]['TimeStamp'] == 0:
                                            mro_item_dict[object_ue_id][mr_name_entity]['TimeStamp'] = mr_utils.get_timestamp_by_str_format(time_stamp_str)
                                        else:
                                            time_spec = mr_utils.get_timestamp_by_str_format(time_stamp_str) - mro_item_dict[object_ue_id][mr_name_entity]['TimeStamp']
                                            if time_spec != int(gl.MR_CONF['SamplePeriod']):
                                                temp_test_flag_dict[mr_name_entity] += 1
                                                mr_utils.out_text_dict_append_list(out_text_list, mro_file, '[{0}]= << TimeStamp duplicate >> TimeStamp:[{1}]\n'.format(mr_name_entity, time_stamp_str) \
                                                        if time_spec == 0 else '[{0}]= << TimeStamp gap inaccurate >> TimeStamp:[{1}]\n'.format(mr_name_entity, time_stamp_str))
                                                temp_loop_flag = 1
                                            #当出现时间间隔错误, 继续往下检索, save the last timestamp
                                            mro_item_dict[object_ue_id][mr_name_entity]['TimeStamp'] = mr_utils.get_timestamp_by_str_format(time_stamp_str)

                                    else:
                                        # rip : ripnum in a period 
                                        id_list1 = mr_utils.get_split_value(mr_utils.xml_attr(object_entity, 'id'), ':', [' ', ':'])
                                        #prbnum : cellid + rip
                                        prbnum =  str(mr_utils.is_cell_id_exist(int(id_list1[0]))[1]) + ':' + id_list1[2]
                                        if mro_item_dict[object_ue_id][mr_name_entity]['prbnum'].__contains__(prbnum) == False:
                                            mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum] = {'num':0, 'TimeStamp':0}
                                        mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['num'] += 1
                                        #TimeStamp 判断
                                        time_stamp_str = mr_utils.xml_attr(object_entity, 'TimeStamp')
                                        time_stamp_value = mr_utils.get_timestamp_by_str_format(time_stamp_str)
                                        
                                        if mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['TimeStamp'] == 0:
                                            mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['TimeStamp'] = time_stamp_value
                                        else:
                                            time_spec = time_stamp_value - mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['TimeStamp']
                                            if time_spec != int(gl.MR_CONF['SamplePeriod']):
                                                temp_test_flag_dict[mr_name_entity] += 1
                                                len_of_mrofile = len(out_text_list[mro_file])
                                                if len_of_mrofile == 0 or time_stamp_str not in out_text_list[mro_file][len_of_mrofile - 1]:
                                                    mr_utils.out_text_dict_append_list(out_text_list, mro_file, '[{0}]=<TimeStamp duplicate> TimeStamp:{1}\n'.format(mr_name_entity, time_stamp_str) \
                                                        if time_spec == 0 else '[{0}]=<TimeStamp gap inaccurate> TimeStamp:{1}\n'.format(mr_name_entity, time_stamp_str))
                                                temp_loop_flag = 1
                                        mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['TimeStamp'] = time_stamp_value

                                    if temp_test_flag_dict[mr_name_entity] == 0:
                                        mro_item_dict[object_ue_id][mr_name_entity]['flag'] -= 1
                                    value_temp_test = 0
                                    for value_entity in mr_utils.xml_element(object_entity, 'v'):
                                        value_num = -10
                                        value_list1 = mr_utils.get_split_value(mr_utils.xml_text(value_entity))
                                        if value_list1[mr_item_pos].isdigit() == True:
                                            value_num = int(value_list1[mr_item_pos])
                                        else:
                                            value_temp_test = 1
                                            continue
                                        if value_num < mro_item_dict[object_ue_id][mr_name_entity]['range'][0] or value_num > mro_item_dict[object_ue_id][mr_name_entity]['range'][1]:
                                            temp_test_flag_dict[mr_name_entity] += 1
                                            mr_utils.out_text_dict_append_list(out_text_list, mro_file, '[' + mr_name_entity + ']={' + 'value Comfusion:' + str(value_num) + '=[' + str(mro_item_dict[object_ue_id][mr_name_entity]['range']) + ']'  + '}\n')
                                            temp_loop_flag = 1
                                            value_temp_test = 1
                                        else:
                                            value_temp_test = 0

                                    mro_item_dict[object_ue_id][mr_name_entity]['num'] += 1

                                    if temp_test_flag_dict[mr_name_entity] == 0:
                                        mro_item_dict[object_ue_id][mr_name_entity]['flag'] -= 1

            #==========================================================================
            #   write output 
            #==========================================================================
            gl.OUT_STR_LIST.append('\n [' + mro_file + ']:\n')
            gl.OUT_STR_LIST.append('[' + str(cellid) + ']=[')
            gl.OUT_STR_LIST.append(str(ideal_sample_num) + 'or' + str(ideal_sample_num+1) + '] |')
            test_integrity_pqsh = 0
            for object_id in mro_item_dict:
                if 'NIL' in object_id:
                    continue
                gl.OUT_STR_LIST.append( object_id + ":")
                for mr_name in out_mro_list:
                    gl.OUT_STR_LIST.append(str(mro_item_dict[object_id][mr_name]['num']) + ' | ')
                if mro_item_dict[object_id][mr_name]['num'] != ideal_sample_num and mro_item_dict[object_id][mr_name]['num'] != ideal_sample_num + 1:
                    test_integrity_pqsh = 1
                if temp_test_flag_dict[mr_name] != 0:
                    test_integrity_pqsh = 1
            gl.OUT_STR_LIST.append(str(test_integrity_pqsh == 0) + ' | \n')
            
            idea_rip_num = ideal_sample_num * len(gl.MR_CONF['SubFrameNum'].strip(',').split(','))

            gl.OUT_STR_LIST.append(date_time + " | ")
            gl.OUT_STR_LIST.append(gl.TEST_CONF['test_total_time'] + " | " )
            gl.OUT_STR_LIST.append('ideal rip_num=' + str(idea_rip_num) + 'or' + str((ideal_sample_num+1) * len(gl.MR_CONF['SubFrameNum'].strip(',').split(','))) + ' | ')

            sum_all_prbnum_count = 0
            for object_id in mro_item_dict:
                if re.search(r'NIL', object_id) == None:
                    continue
                gl.OUT_STR_LIST.append('{')
                for prbnum_id in mro_item_dict[object_id]['MR.LteScRIP']['prbnum']:
                    if mro_item_dict[object_id]['MR.LteScRIP']['prbnum'][prbnum_id].__contains__('num') == True:
                        gl.OUT_STR_LIST.append(' [' + prbnum_id + ']=' + str(mro_item_dict[object_id]['MR.LteScRIP']['prbnum'][prbnum_id]['num'])  )
                        sum_all_prbnum_count += mro_item_dict[object_id]['MR.LteScRIP']['prbnum'][prbnum_id]['num']
                gl.OUT_STR_LIST.append('  [total]=' + str(mro_item_dict[object_id]['MR.LteScRIP']['num']) + ' } | ')

            result_div = sum_all_prbnum_count*1.0 / (idea_rip_num) * 1.0

            rip_accuracy_index = result_div if result_div == 1.0 else sum_all_prbnum_count*1.0 / ((ideal_sample_num + 1) * len(gl.MR_CONF['SubFrameNum'].strip(',').split(','))) * 1.0
            gl.OUT_STR_LIST.append(str( rip_accuracy_index) + ' | \n')
        except  Exception as result:
            raise Exception('[%s]   %s |  %s'%(__name__ ,str(result.__traceback__.tb_lineno) if not gl.IS_PY2  else '',result))

    full_name = os.path.join(gl.SOURCE_PATH, gl.XLS_NAME)
    save_name = os.path.join(gl.OUTPUT_PATH, gl.XLS_NAME)
    if int(gl.TEST_CONF['is_58_out_excel']) == 1 and os.path.exists(full_name):
        gl.OUT_STR_LIST.append('\n结果已写入 successful ->'  )
        gl.OUT_STR_LIST.append(save_name)

    if temp_loop_flag == 1:
        gl.OUT_STR_LIST.append('\nerror:\n')

        for mr_file in out_text_list:
            gl.OUT_STR_LIST.append('=====>' + mr_file + ':\n')
            for i in range(len(out_text_list[mr_file])):
                gl.OUT_STR_LIST.append('\t' + out_text_list[mr_file][i])
    gl.OUT_STR_LIST.append('\n')
    
    if int(gl.TEST_CONF['is_58_out_excel']) == 1 and os.path.exists(full_name):
        excel_workbook = openpyxl.load_workbook(full_name)
        work_sheet = excel_workbook.worksheets[2]

        for mrs_dict_key in temp_test_flag_dict:
            if temp_test_flag_dict[mrs_dict_key] == 0:
                work_sheet.cell(xml_mro_item_dict[mrs_dict_key], 17, 'Y')
            else:
                work_sheet.cell(xml_mro_item_dict[mrs_dict_key], 17, 'N')
        excel_workbook.save(filename=save_name)


#========================================================================================#
#             test 59
# check the capacity of sample of MRE A1-A6, B1,B2
#========================================================================================#    

def test59_file_integrity():
    # mr_utils.test_out_data_item_header("test_59")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    
    # the standard scale of mr_item 
    range_list = [[0,97], [0,97], [0,34], [0,34], [0, 41589], [0, 503], [0, 41589],  [0, 503], [0,1023], [0,63], [0,7], [0,7]]

    out_mre_list = ['A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'B1', 'B2']

    mre_smr_head = 'MR.LteScRSRP MR.LteNcRSRP MR.LteScRSRQ MR.LteNcRSRQ MR.LteScEarfcn MR.LteScPci MR.LteNcEarfcn MR.LteNcPci MR.GsmNcellBcch MR.GsmNcellCarrierRSSI MR.GsmNcellNcc MR.GsmNcellBcc'
    mre_smr_list = ['MR.LteScRSRP' ,'MR.LteNcRSRP' ,'MR.LteScRSRQ' ,'MR.LteNcRSRQ' ,'MR.LteScEarfcn' ,'MR.LteScPci' ,'MR.LteNcEarfcn' ,'MR.LteNcPci' ,'MR.GsmNcellBcch' ,'MR.GsmNcellCarrierRSSI' ,'MR.GsmNcellNcc' ,'MR.GsmNcellBcc']

    # excel position
    xml_out_dict = {'A1':{'item':[2,5], 'event':2}, 'A2':{'item':[6,9], 'event':6}, 'A3':{'item':[10,13], 'event':10}, 'A4':{'item':[14,17], 'event':14}, 'A5':{'item':[18,25], 'event':18},
                    'A6':{'item':[26,33], 'event':26}, 'B1':{'item':[34,37], 'event':34}, 'B2':{'item':[38, 45], 'event':38}}

    out_mre_text = {}
    mre_conf_dict = {}
    
    
    for time_entity in gl.MR_DICT:
        try:
            
            event_deal_list = []
            mre_conf_dict = {   'A1':{'flag':0, 'pos':[0,2,4,5],            'range':range_list, 'error_pos_list':[]},
                                'A2':{'flag':0, 'pos':[0,2,4,5],            'range':range_list, 'error_pos_list':[]},
                                'A3':{'flag':0, 'pos':[0,1,2,3,4,5,6,7],    'range':range_list, 'error_pos_list':[]},
                                'A4':{'flag':0, 'pos':[0,1,2,3,4,5,6,7],    'range':range_list, 'error_pos_list':[]},
                                'A5':{'flag':0, 'pos':[0,1,2,3,4,5,6,7],    'range':range_list, 'error_pos_list':[]},
                                'A6':{'flag':0, 'pos':[0,1,2,3,4,5,6,7],    'range':range_list, 'error_pos_list':[]},
                                'B1':{'flag':0, 'pos':[0,2,4,5,8,9,10,11],  'range':range_list, 'error_pos_list':[]},
                                'B2':{'flag':0, 'pos':[0,2,4,5,8,9,10,11],  'range':range_list, 'error_pos_list':[]}}
            for event_type in mre_conf_dict:
                if event_type in gl.TEST_CONF['event']:
                    event_deal_list.append(event_type)
            temp_mre_test_flag_dict = {'A1':0, 'A2':0, 'A3':0, 'A4':0, 'A5':0, 'A6':0, 'B1':0, 'B2':0}

            mre_file_name   =   mr_utils.get_file_name(time_entity, 'MRE')
            if mre_file_name == '' :
                continue
            mre_dom         =   mr_utils.xml_dom(time_entity, 'MRE')
            mre_root = mr_utils.xml_doc(mre_dom)
            smr_list = mr_utils.xml_element(mre_root, 'smr')
            if len(smr_list) == 0:
                continue
            
            smr_str = mr_utils.xml_text(smr_list[0])
            if len(smr_list) == 0 :
                temp_flag = 0
                for i in range(len(mre_smr_list)):
                    if mr_utils.is_mr_item_need_exist(mre_smr_list[i]) == True:
                        temp_flag = 1
                        break
                if temp_flag == 1:
                    for event_type in event_deal_list:
                        temp_mre_test_flag_dict[event_type] = 1
                    mr_utils.out_text_dict_append_list(out_mre_text, mre_file_name, 'smr err:' + 'no smr lable' if len(smr_list) == 0 else smr_list[0].firstChild.data)
            
            if smr_str not in mre_smr_head:
                mr_utils.get_mre_pos_list_by_mapping(mre_conf_dict, smr_str)
                #for event in mre_conf_dict:
                #    print (str(mre_conf_dict[event]['pos']))
            for object_entity in mr_utils.xml_element(mre_root, 'object'):
                test_excess_flag = 0
                for event_type in event_deal_list:
                    if event_type == mr_utils.xml_attr(object_entity, 'EventType'):
                        # get time
                        time_now_str = str(mr_utils.xml_attr(object_entity, 'TimeStamp'))
                        time_now_timestamp = mr_utils.get_timestamp_by_str_format(time_now_str)
                        

                        test_excess_flag = 1
                        mre_conf_dict[event_type]['flag'] = 1
                        for value_entity in mr_utils.xml_element(object_entity, 'v'):
                            value_str = (mr_utils.xml_text(value_entity))
                            value_list = mr_utils.get_split_value(value_str)
                            for pos in range(12):
                                if pos in mre_conf_dict[event_type]['pos']:
                                    value_num = 0
                                    if value_list[pos].isdigit() == True:
                                        value_num = int(value_list[pos])

                                    if value_num < mre_conf_dict[event_type]['range'][pos][0] or value_num > mre_conf_dict[event_type]['range'][pos][1]:
                                        temp_mre_test_flag_dict[event_type] += 1
                                        mr_utils.out_text_dict_append_list(out_mre_text, mre_file_name, 'value confusion: << {0} = {1} ->({2},{3} )>> event-[{4}] TimeStamp:{5}\n'.\
                                            format(mre_smr_list[pos],str(value_num),str(mre_conf_dict[event_type]['range'][pos][0]),
                                            str(mre_conf_dict[event_type]['range'][pos][1]), event_type, time_now_str) )

                                        if pos not in mre_conf_dict[event_type]['error_pos_list']:
                                            mre_conf_dict[event_type]['error_pos_list'].append(pos)
                                else:
                                    if len(value_list) <= pos:
                                        break
                                    if value_list[pos] != 'NIL':
                                        temp_mre_test_flag_dict[event_type] += 1
                                        mr_utils.out_text_dict_append_list(out_mre_text, mre_file_name, 
                                                'event format confusion(not NIL):[{0}] event:[{1}}] TimeStamp:[{2}]'.format(mre_smr_list[pos], event_type, time_now_str))
                                        if pos not in mre_conf_dict[event_type]['error_pos_list']:
                                            mre_conf_dict[event_type]['error_pos_list'].append(pos)

                if test_excess_flag == 0:
                    mr_utils.out_text_dict_append_list(out_mre_text, mre_file_name, 'suplus event:{0} - TimeStamp:{1}\n'.format(mr_utils.xml_attr(object_entity, 'EventType'),  time_now_str))

            #==========================================================================
            #   write output 
            #==========================================================================
            gl.OUT_STR_LIST.append('====>%s:\n'%(mre_file_name))
            gl.OUT_STR_LIST.append(date_time + " | " )
            gl.OUT_STR_LIST.append(gl.TEST_CONF['test_total_time'] + " | " )

            for event_type in out_mre_list:
                gl.OUT_STR_LIST.append('True | ' if event_type in event_deal_list else 'False | ')

            gl.OUT_STR_LIST.append('\n' + date_time + " | ")
            gl.OUT_STR_LIST.append(gl.TEST_CONF['test_total_time'] + " | " )

            for event_type in out_mre_list:
                if temp_mre_test_flag_dict[event_type] == 0:
                    gl.OUT_STR_LIST.append('True | ' if event_type in event_deal_list else 'False | ' )

                else:
                    gl.OUT_STR_LIST.append('[')
                    for pos in mre_conf_dict[event_type]['error_pos_list']:
                        gl.OUT_STR_LIST.append(mre_smr_list[pos] + ' ' )
                    gl.OUT_STR_LIST.append('] | ')

            gl.OUT_STR_LIST.append('\n\n')
        except Exception as result:
            raise Exception('[%s]   %s |  %s'%(__name__ ,str(result.__traceback__.tb_lineno) if not gl.IS_PY2  else '',result))

    full_name = os.path.join(gl.SOURCE_PATH, gl.XLS_NAME)
    save_name = os.path.join(gl.OUTPUT_PATH, gl.XLS_NAME)
    if int(gl.TEST_CONF['is_59_out_excel']) == 1 and os.path.exists(full_name):
        excel_workbook = openpyxl.load_workbook(full_name)
        work_sheet = excel_workbook.worksheets[3]

        for event_type in out_mre_list:
            work_sheet.cell(xml_out_dict[event_type]['event'], 9, 'Y' if event_type in event_deal_list else 'N')
            test_count = 0
            for i in range(xml_out_dict[event_type]['item'][0], xml_out_dict[event_type]['item'][1]+1):
                if mre_conf_dict[event_type]['pos'][test_count] not in mre_conf_dict[event_type]['error_pos_list'] and event_type in event_deal_list :
                    work_sheet.cell(i, 8, 'Y')
                else:
                    work_sheet.cell(i, 8, 'N')
                test_count += 1

        excel_workbook.save(filename=save_name)

        gl.OUT_STR_LIST.append('结果已写入 successful  ->' + gl.XLS_NAME + '\n')

    for mre_file_name in out_mre_text:
        gl.OUT_STR_LIST.append('\n======>' + mre_file_name + ':\n')
        for list_entity in out_mre_text[mre_file_name]:
            gl.OUT_STR_LIST.append('\t' + list_entity)
            
    gl.OUT_STR_LIST.append('\n')
    
#========================================================================================#
#             test 61
# check the accuracy of sample of MRO rsrp, rsrq, phr, compare with UE log information. single UE 
#========================================================================================# 
  
def test61_file_accuracy():
    # mr_utils.test_out_data_item_header("test_61")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)
    gl.OUT_STR_LIST.append(date_time + " | ")
    gl.OUT_STR_LIST.append(gl.TEST_CONF['test_total_time'] + " | \n" )
    
    try:
        for time_entity in gl.MR_DICT:
            mrs_file_name = mr_utils.get_file_name(time_entity, 'MRS')
            mro_file_name = mr_utils.get_file_name(time_entity, 'MRO')
            measurement_mro_list = []
            measurement_mrs_list = []
            object_num = 0
            count_cnt = 0
            ue_info_dict = {
                'MR.RSRP' : [ 0 for x in range(0, 48) ],
                'MR.RSRQ' : [ 0 for x in range(0, 18) ],
                'MR.PowerHeadRoom' : [ 0 for x in range(0, 64) ],
            }
            
            MRS_info_dict = {'MR.RSRP':[], 'MR.RSRQ':[], 'MR.PowerHeadRoom':[]}
            
            if mrs_file_name != '':
                mrs_dom = mr_utils.xml_dom(time_entity, 'MRS')
                mrs_root = mr_utils.xml_doc(mrs_dom)
                measurement_mrs_list = mr_utils.xml_element(mrs_root, 'measurement')
            mrs_start_time_str , _, mrs_end_time_str, _ = mr_utils.get_start_end_timestr_from_mr_file(time_entity)
                
            if mro_file_name != '':
                mro_dom = mr_utils.xml_dom(time_entity, 'MRO')
                mro_root = mr_utils.xml_doc(mro_dom)
                measurement_mro_list = mr_utils.xml_element(mro_root, 'measurement')

            # read mrs_file get rsrp,rsrq,phr  statistical data  
            for measurement_entity in measurement_mrs_list:
                mr_name = mr_utils.xml_attr(measurement_entity, 'mrName') 
                if mr_name  in MRS_info_dict:
                    # single UE , just consider the one object
                    object_node = mr_utils.xml_element(measurement_entity, "object")[0]
                    value_str = str(  mr_utils.xml_text(mr_utils.xml_element(object_node, "v")[0]) )
                    MRS_info_dict[mr_name] = mr_utils.get_split_value(value_str)
            # read mro_file: get the first sample time and end sample time 
            for measurement_entity in measurement_mro_list:
                smr_name = mr_utils.xml_text(mr_utils.xml_element(measurement_entity, 'smr')[0])
                if 'MR.LteScEarfcn' in smr_name:
                    object_list = mr_utils.xml_element(measurement_entity, 'object')
                    mrs_start_time_str = mr_utils.xml_attr(object_list[0], 'TimeStamp')
                    mrs_end_time_str = mr_utils.xml_attr(object_list[-1], 'TimeStamp') 
                    object_num = len(object_list)       
            
            ue_list = mr_utils.read_ue_log_filter(gl.SOURCE_PATH + gl.TEST_CONF['ue_log_file_name'], mrs_start_time_str, mrs_end_time_str)  

               
            phr_list = mr_utils.get_phr_data(gl.SOURCE_PATH + gl.TEST_CONF['phr_file_name'], mrs_start_time_str, mrs_end_time_str)

            for i in range(len(ue_list)):
                if count_cnt >= object_num:
                    break
                if i != 0 and abs(mr_utils.get_timestamp_by_str_format(ue_list[i]['timestamp'])  - mr_utils.get_timestamp_by_str_format(ue_list[i - 1]['timestamp'])) < 2000: 
                #ue_info store the ueLog information , include the scrsrp, rsrq, and nc rsrp,rsrq,pci . so judge the timestamp for the one sample_time_information then count rsrp,rsrq 
                    continue
                mr_utils.rsrp_count(ue_list[i]['MR.LteScRSRP'], ue_info_dict['MR.RSRP'])
        
                mr_utils.rsrq_count(ue_list[i]['MR.LteScRSRQ'], ue_info_dict['MR.RSRQ'])

                mr_utils.phr_count(phr_list[count_cnt]['phr'], ue_info_dict['MR.PowerHeadRoom'])
                count_cnt += 1
            mr_utils.create_61_out_excel(time_entity, ue_info_dict, MRS_info_dict)
        
            gl.OUT_STR_LIST.append(mrs_file_name + "    =========>  successful write result into [" + gl.OUTPUT_PATH + time_entity + '.xlsx' + ']\n')
    except Exception as result:
        raise Exception('[%s]   %s |  %s'%(__name__ ,str(result.__traceback__.tb_lineno) if not gl.IS_PY2  else '',result))
        


#========================================================================================#
#             test 62
# check the accuracy of prb, compare  MRS/MRO' prb  and  PM_file ULMean
#========================================================================================#         

def test62_file_accuracy():
    # mr_utils.test_out_data_item_header("test_62")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)
    gl.OUT_STR_LIST.append(date_time + " | ")
    gl.OUT_STR_LIST.append(gl.TEST_CONF['test_total_time'] + " | \n" )
    try:
        for time_entity in gl.MR_DICT:
            pm_value_list = [ str('-126.0000') for x in range(341,441) ]
            map_prb = {}
            out_list = []
            if time_entity in gl.PM_DICT:
                pm_dom = gl.PM_DICT[time_entity]
                pm_root = mr_utils.xml_doc(pm_dom)
                measData_node = mr_utils.xml_element(pm_root, "measData")[0]
                measValue_node = mr_utils.xml_element(measData_node, "measValue")[3]
                pm_value_list = [ mr_utils.xml_text(x) for x in mr_utils.xml_element(measValue_node, "r")[341:441] ]
                
            mrs_file_name = mr_utils.get_file_name(time_entity, 'MRS')
            
            if mrs_file_name == '':
                continue
            mrs_dom = mr_utils.xml_dom(time_entity, 'MRS')
            mrs_root = mr_utils.xml_doc(mrs_dom)
            measurement_list = mr_utils.xml_element(mrs_root, 'measurement')
            ripprb_node = measurement_list[4]
            
            for object_entity in mr_utils.xml_element(ripprb_node, "object"):
                id_list = mr_utils.get_split_value(str(mr_utils.xml_attr(object_entity, "id")), ':', [':', ' '])
                prbnum = int(id_list[3])
                ripnum = int(id_list[2])
                value_list = mr_utils.get_split_value(   mr_utils.xml_text(   mr_utils.xml_element(object_entity, "v")[0]   )  )
                # map_prb[ripprb_idx] = value_list
                
                if prbnum not in map_prb:
                    map_prb[prbnum] = [int(x) for x in value_list]
                else:
                    for i in range(len(value_list)):
                        map_prb[prbnum][i] += int(value_list[i])
                
            for prb_entity in map_prb:
                value1 = 0
                for i in range(len(map_prb[prb_entity])):
                    item_value = int(map_prb[prb_entity][i])
                    if item_value != 0:
                        value1 += (10 ** (gl.k_list[i])) * item_value
                
                value2 = sum(map_prb[prb_entity])         

                temp_out = []

                out = 10 * math.log10(float(value1 / float(value2))) if value2 != 0 else float(-126.000000)
                dup = float(pm_value_list[prb_entity]) - out
                
                temp_out.append(pm_value_list[prb_entity])
                temp_out.append("%.6f"%(out))
                temp_out.append("%03.6f"%(dup))
                
                out_list.append(temp_out)
                
            gl.OUT_STR_LIST.append('\n' + mrs_file_name + "  : \n")
            out_item_list = ['网管  值', 'MR平均值', '偏差均值' ]
            for i in range(3):
                line = " " * 8 + "|" + out_item_list[i] + " " * 5 + '|' + "|".join([str(x[i]) + " "*(11-len(str(x[i]))) for x in out_list])
                gl.OUT_STR_LIST.append(line+ '\n')


    except Exception as result:
        raise Exception('[%s]   %s |  %s'%(__name__ ,str(result.__traceback__.tb_lineno) if not gl.IS_PY2  else '',result))
    
    
#========================================================================================#
#             test 63
# check the accuracy of neighbor cell, compare MRO nc_cell information and UE information
#========================================================================================#   
    
def test63_file_accuracy():
    # mr_utils.test_out_data_item_header("test_63")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)
    gl.OUT_STR_LIST.append(date_time + " | ")
    gl.OUT_STR_LIST.append(gl.TEST_CONF['test_total_time'] + " | " )
    output_list = ["MR.LteNcRSRP", "MR.LteNcRSRQ", "MR.LteNcEarfcn", "MR.LteNcPci"]
    output_head_str = ['信令文件输出采样点数目', 'MR采样点正确匹配数目  ', '偏离程度              ']
    item_pos = {}
    mro_error_output_dist = {}
    try:
        for time_entity in gl.MR_DICT:
            mro_file_name = mr_utils.get_file_name(time_entity, 'MRO')
            count_cell_ue = set()
            count_cell_mro = set()
            # ue matched number
            # count_list = [ 0, 0, 0, 0]
            # mro ncell number matched 
            mro_value_cnt_list  = [0, 0, 0, 0]
            # ue nccell number
            ue_nc_count_list = [0, 0, 0, 0]
            
            mro_start_time_str , _, mro_end_time_str, _ = mr_utils.get_start_end_timestr_from_mr_file(time_entity)

            if mro_file_name == '':
                continue
            mro_dom = mr_utils.xml_dom(time_entity, 'MRO')
            mro_root = mr_utils.xml_doc(mro_dom)
            measurement_node = mr_utils.xml_element(mro_root, 'measurement')[0]
            smr_str = mr_utils.xml_text( mr_utils.xml_element(measurement_node, 'smr')[0] )
            mr_utils.get_mr_item_pos_single_smr(item_pos, smr_str)
            object_list = mr_utils.xml_element(measurement_node, 'object')
            
            mro_start_time_str = mr_utils.xml_attr(object_list[0], 'TimeStamp')
            mro_end_time_str = mr_utils.xml_attr(object_list[-1], 'TimeStamp') 
            
            
            ue_list = mr_utils.read_ue_log_filter(gl.SOURCE_PATH + gl.TEST_CONF['ue_log_file_name'], mro_start_time_str, mro_end_time_str)  
            map_earfcn = mr_utils.get_map_pci_earfcn()
            temp_count_ue_cnt = 0

            for object_entity in object_list:
                time_str = mr_utils.xml_attr(object_entity, 'TimeStamp')
                # print (time_str)
                mro_time_stamp = mr_utils.get_timestamp_by_str_format(time_str)
                ue_time_stamp = mr_utils.get_timestamp_by_str_format(ue_list[temp_count_ue_cnt]['timestamp'])
                # print (time_str)
                # print (ue_list[temp_count_ue_cnt]['timestamp'])
                if mro_time_stamp - ue_time_stamp > 5120:
                    for i in range(temp_count_ue_cnt, len(ue_list)):
                        ue_time_stamp = mr_utils.get_timestamp_by_str_format(ue_list[i]['timestamp'])
                        if mro_time_stamp - ue_time_stamp < 5120 :
                            temp_count_ue_cnt = i
                            break
                if ue_time_stamp - mro_time_stamp > 5120:
                    continue  
                # print (len(ue_list[temp_count_ue_cnt]['nc_info']))
                for value_entity in mr_utils.xml_element(object_entity, 'v'):

                    value_list = mr_utils.get_split_value(mr_utils.xml_text(value_entity))

                    UE_dict_temp = {
                        'MR.LteNcRSRP':     str(ue_list[temp_count_ue_cnt]['nc_info'][0]['MR.LteNcRSRP'])            if len(ue_list[temp_count_ue_cnt]['nc_info']) != 0      else 'NIL',
                        'MR.LteNcRSRQ':     str(ue_list[temp_count_ue_cnt]['nc_info'][0]['MR.LteNcRSRQ'])            if len(ue_list[temp_count_ue_cnt]['nc_info']) != 0      else 'NIL',
                        'MR.LteNcEarfcn':   str(map_earfcn[str(ue_list[temp_count_ue_cnt]['nc_info'][0]['MR.LteNcPci'])]) if len(ue_list[temp_count_ue_cnt]['nc_info']) != 0      else 'NIL',
                        'MR.LteNcPci':      str(ue_list[temp_count_ue_cnt]['nc_info'][0]['MR.LteNcPci'])             if len(ue_list[temp_count_ue_cnt]['nc_info']) != 0      else 'NIL'
                    }
                    
                    for i in range(len(output_list)):
                        item = output_list[i]
                        item_value = value_list[item_pos[item]]
                        # get ue nc_cell number
                        ue_nc_count_list[i]     += 1    if UE_dict_temp[item] != 'NIL'  or item_value == UE_dict_temp[item] == 'NIL'    else 0
                        # add cell by nc-pci 
                        if item == 'MR.LteNcPci':
                            count_cell_ue.add(UE_dict_temp['MR.LteNcPci'] if UE_dict_temp['MR.LteNcPci'] != 'NIL' else 'NIL')
                            count_cell_mro.add(item_value if item_value != 'NIL' else 'NIL' )
                        
                        if UE_dict_temp[item] == item_value:
                            # if item_value != 'NIL' and UE_dict_temp[item] != 'NIL': 
                            mro_value_cnt_list[i] += 1 
                        else:
                            mr_utils.out_text_dict_append_list(mro_error_output_dist, mro_file_name, 
                                "\n====>\t<UE timestamp : %s >   |%s| => \t [%s] != [%s] \t <= |%s| |  <MRO timestamp: %s >  "%(
                                    ue_list[temp_count_ue_cnt]['timestamp'], item, UE_dict_temp[item], item_value , item, time_str  ))
                    
                    if len(ue_list[temp_count_ue_cnt]['nc_info']) != 0:
                        ue_list[temp_count_ue_cnt]['nc_info'].pop(0)
                    if len(ue_list[temp_count_ue_cnt]['nc_info']) == 0:
                        temp_count_ue_cnt += 1
            
            gl.OUT_STR_LIST.append('\n' + mro_file_name + ':\n')
            out_list_temp = [ue_nc_count_list, mro_value_cnt_list, count_cell_ue, count_cell_mro]
            for i in range(len(output_head_str) ):
                gl.OUT_STR_LIST.append (output_head_str[i] + ':  ')
                # print (ue_nc_count_list)
                if i != 2:
                    gl.OUT_STR_LIST.append ( ' | '.join([str(x) for x in out_list_temp[i]]) + '   |  ' )
                    gl.OUT_STR_LIST.append( str(len(out_list_temp[i + 2])) + ' |  cell_list:' + str(out_list_temp[i + 2]) + '\n')
                else:
                    gl.OUT_STR_LIST.append ( ' | '.join([ str((ue_nc_count_list[x]) == (mro_value_cnt_list[x])) for x in range(len(output_list)) ]) + '   |  ' )
                    gl.OUT_STR_LIST.append ( str(count_cell_ue == count_cell_mro) +  ' |  \n' )
        for mro_file_name in mro_error_output_dist:
            gl.OUT_STR_LIST.append('\n======>' + mro_file_name + ':\n')
            for list_entity in mro_error_output_dist[mro_file_name]:
                gl.OUT_STR_LIST.append('\t' + list_entity)
        gl.OUT_STR_LIST.append('\n')
    except Exception as result:
        raise Exception('[%s]   %s |  %s'%(__name__ ,str(result.__traceback__.tb_lineno) if not gl.IS_PY2  else '',result))
        
#========================================================================================#
#             test 71
# check the accuracy of file, MRO,MRE,MRS   filename, standard
#========================================================================================#   

def test71_file_accuracy():
    # mr_utils.test_out_data_item_header("test_71")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)
    
    gl.OUT_STR_LIST.append(date_time + " | ")
    gl.OUT_STR_LIST.append(gl.TEST_CONF['test_total_time'] + " | " )
    
    time_sub = 0 if gl.IS_SSH_CONN == 0 else mr_utils.get_two_linux_time_sub()
    test_flag = 0
    out_text_list = {}
    file_idx_list = ['MRO', 'MRE', 'MRS']
    check_function_list = [lambda root: mr_utils.is_mro_correct(root), lambda root: mr_utils.is_mre_correct(root), lambda root: mr_utils.is_mrs_correct(root)]
    try:
        for time_entity in gl.MR_DICT:
            for i in range(len(file_idx_list)):
                test_flag = 0
                file_name = mr_utils.get_file_name(time_entity, file_idx_list[i])
                if file_name == '':
                    continue
                xml_dom = mr_utils.xml_dom(time_entity, file_idx_list[i])
                
                # check file_name 
                file_dict = mr_utils.get_file_dict(time_entity, file_idx_list[i])
                ret_file_check = mr_utils.MR_xml_file_name_check(file_dict)
                if  ret_file_check[0] == False :
                    test_flag |= (0x1 << 1)
                    mr_utils.out_text_dict_append_list(out_text_list, file_name, 'file_name format err: ==> %s   [%s]\n' %( file_name, ret_file_check[1] ))
                
                mr_file_root = mr_utils.xml_doc(xml_dom)
                # start_report_time is mili
                start_report_Time = mr_utils.get_timestamp_by_str_format(  mr_utils.xml_attr(mr_utils.xml_element(mr_file_root, 'fileHeader')[0], 'reportTime')  )
                start_report_Time /= 1000
                file_create_time = 0
                if gl.IS_LINUX == 0:
                    #TODO: should get two linux time gap -> time_sub when process in windows, need to get the the actual reporttime. when in the linux server machine,no need 
                    start_report_Time -= time_sub
                    
                    # if gl.MR_REMOTE_FILE_TIME_DIST.__contains__(file_name) == False:
                    #     file_create_time = time.mktime(time.gmtime(os.path.getmtime()))
                    # else:
                    #     file_create_time = gl.MR_REMOTE_FILE_TIME_DIST[filename]
                else:
                    file_create_time = time.mktime(time.gmtime(os.path.getmtime( gl.MR_TEST_PATH + file_name)))
                if file_create_time - start_report_Time > int(gl.TEST_CONF['file_delay_time'])*60:
                    test_flag |= (0x1 << 4)
                    mr_utils.out_text_dict_append_list(out_text_list, file_name,
                      'file create time:[%s]-[%s] \n' % (str(time.strftime( '%Y-%m-%dT%H:%M:%S',time.localtime(file_create_time))), str(time.strftime( '%Y-%m-%dT%H:%M:%S',time.localtime(start_report_Time)))  ) )
                
                #TODO: check mr file, can modify the function by self
                mr_ret_list = check_function_list[i](mr_file_root)

                if mr_ret_list[0] == False:
                    test_flag |= (0x1 << 2)
                    mr_utils.out_text_dict_append_list(out_text_list,file_name, 'file format  -> %s\n' % ( mr_ret_list[1]))

                gl.OUT_STR_LIST.append('====>' + file_name + ':   \n')
                gl.OUT_STR_LIST.append('N | ' if test_flag & (0x1 << 1) != 0 else 'Y | ')

                gl.OUT_STR_LIST.append('N | ' if test_flag & (0x1 << 2) != 0 else 'Y | ')
                #TODO: check the compressed file, haven't make, so write [] 
                gl.OUT_STR_LIST.append('[] | N | \n' if test_flag & (0x1 << 4) != 0 else '[] | Y | \n')

        if len(out_text_list) != 0:
            gl.OUT_STR_LIST.append('error:\n')
        for file_name in out_text_list:
            gl.OUT_STR_LIST.append('======>' + file_name + ' :\n')
            for i in range(len(out_text_list[file_name])):
                gl.OUT_STR_LIST.append('\t' + out_text_list[file_name][i])
    except Exception as result:
        raise Exception('[%s]   %s |  %s'%(__name__ ,str(result.__traceback__.tb_lineno) if not gl.IS_PY2  else '',result))
    
    
#========================================================================================#
#             test 72
# check the accuracy of file, MRO,MRE,MRS   format , check the file_structure,attribute_tag,name
#========================================================================================#   
#TODO: when error, test72 will output lots of information, can optimize it 
    
def test72_file_accuracy():
    # mr_utils.test_out_data_item_header("test_72")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)
    gl.OUT_STR_LIST.append(date_time + " | ")
    gl.OUT_STR_LIST.append(gl.TEST_CONF['test_total_time'] + " | \n" )
    schema_dict = {"MRO":gl.SOURCE_PATH+"mro_schema.xsd", "MRE":gl.SOURCE_PATH+'mre_schema.xsd', "MRS":gl.SOURCE_PATH+'mrs_schema.xsd'}
    
    out_text_dict = {}
    file_header_list_name = ['reportTime', 'startTime', 'endTime']

    for time_entity in gl.MR_DICT:
        try:
            test_flag = 0
            measureItem_list = {}
            mr_utils.get_measureItem_list(measureItem_list)
            for mr_type in gl.MR_DICT[time_entity][0]:
                
                # check file by schema
                schema_doc  = etree.parse(schema_dict[mr_type])
                schema_ret = etree.XMLSchema(schema_doc)
                mr_file_full_name = mr_utils.get_file_name(time_entity, mr_type)
                # when file is lost 
                if mr_file_full_name == ''  :
                    if mr_utils.is_mr_item_need_exist(mr_type) == True:
                        # raise Exception('test72 : %s lost %s'%(time_entity, mr_type))
                        print ('test72 : %s lost %s'%(time_entity, mr_type))
                    continue

                data = etree.parse(gl.MR_TEST_PATH +  mr_file_full_name)
                if schema_ret.validate(data) == False:
                    test_flag |= (0x1 << 1)
                    mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name,schema_ret.error_log )
                    
                # check file by self
                #TODO: can modify there, when test need
                
                mr_doc = mr_utils.xml_dom(time_entity, mr_type)
                mr_root = mr_utils.xml_doc(mr_doc)
                    
                # MRO & MRE 
                if mr_type != 'MRS':
                    
                    file_header_list = mr_utils.xml_element(mr_root, 'fileHeader')
                    enb_id = int( mr_utils.xml_attr(mr_utils.xml_element(mr_root, 'eNB')[0], 'id'))

                    for list_item_name in file_header_list_name:
                        if mr_utils.is_str_format_time( mr_utils.xml_attr(file_header_list[0], list_item_name), gl.TIME_FORMAT) == False:
                            mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, list_item_name + ' format error')
                            test_flag |= (0x1 << 2)
                    temp_period = mr_utils.xml_attr(file_header_list[0], 'period')
                    if  temp_period != '0':
                        mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, ' MRO upload peroiod error :[0] !=  ' + temp_period)
                        test_flag |= (0x1 << 2)
                    jobid = int( mr_utils.xml_attr(file_header_list[0], 'jobid'))
                    if  jobid < 0 or jobid > 4294967295 :
                        mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, 'jobid :[0-4294967295] !=  ' + str(jobid))
                        test_flag |= (0x1 << 2)

                    for measurement_entity in mr_utils.xml_element(mr_root, 'measurement'):
                        smr_value = mr_utils.xml_text(mr_utils.xml_element(measurement_entity, 'smr')[0])
                        smr_value_list = mr_utils.get_split_value(smr_value)

                        for i in range(len(smr_value_list)):
                            temp_flag = 0
                            for mr_item in measureItem_list:
                                if smr_value_list[i] == mr_item:
                                    temp_flag = 1
                                    measureItem_list[mr_item] += 1
                                    break
                            if temp_flag == 0:
                                mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, 'surplus the mr_item : %s'%(smr_value_list[i]))
                        for object_entity in mr_utils.xml_element(measurement_entity, 'object'):
                            # check eci cell id
                            id_list_temp = mr_utils.get_split_value(mr_utils.xml_attr(object_entity, 'id'), ':', [':'])
                            eci_id = int( id_list_temp[0]  ) 
                            #TODO: enbid have some trouble, eci = enb << 8 | cellid ,  eci = enb * 256 + cellid.  
                            enb_id = eci_id >> 8 & 0xff
                            cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)
                            
                            if cell_id_ret_list[0] == False or mr_utils.is_enb_id_exist(enb_id) == False:
                                mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '[%s] <cell_id(%d) or enb_id(%d) not exist>'%(
                                    smr_value, cell_id_ret_list[1], enb_id))
                                test_flag |= (0x1 << 3)
                            
                            time_str_temp = mr_utils.xml_attr(object_entity, 'TimeStamp')
                            if mr_utils.is_str_format_time(time_str_temp, gl.TIME_FORMAT) == False:
                                mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '[%s] <timestamp format error> %s'%( smr_value,  time_str_temp  ))
                                test_flag |= (0x1 << 3)
                            # rip is special
                            if smr_value_list[0] == 'MR.LteScRIP':
                                if  mr_utils.is_eci_correct(int(id_list_temp[0])) == False          or       id_list_temp[2] not in gl.MR_CONF['SubFrameNum'] or\
                                    mr_utils.xml_attr(object_entity, 'MmeUeS1apId') != 'NIL' or  mr_utils.xml_attr(object_entity, 'MmeGroupId') != 'NIL' or \
                                    mr_utils.xml_attr(object_entity, 'MmeCode') != 'NIL' :
                                        test_flag |= (0x1 << 4)
                                        mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, 'MR.LteScRIP:<rip object_format error> %s ' %  time_str_temp )
                            else:
                                if mr_utils.is_eci_correct( eci_id ) == False :
                                    mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '[L3] <eci_id not match> %s'% ( time_str_temp ))
                                    test_flag |= (0x1 << 3)
                            temp_value_flag = 0

                            for value_entity in mr_utils.xml_element(object_entity, 'v'):
                                value_str = mr_utils.xml_text(value_entity)
                                if mr_utils.is_mr_value_correct(smr_value, value_str) == False:
                                    mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, 'value not match')
                                    test_flag |= (0x1 << 4)
                                    temp_value_flag = 1
                            if temp_value_flag == 0:
                                for mr_entity in measureItem_list:
                                    if mr_entity != 'MR.RSRP' or mr_entity != 'MR.RSRQ' or mr_entity != 'MR.ReceivedIPower' or mr_entity != 'MR.RIPPRB' or mr_entity != 'MR.PowerHeadRoom':
                                        measureItem_list[mr_entity] += 1

                else:
                    file_header_list = mr_utils.xml_element(mr_root, 'fileHeader')
                    #得到MRO的earfcn，判断MRS文件中的earfcn是否一致
                    mro_doc = mr_utils.xml_dom(time_entity, 'MRO')
                    mro_root = mr_utils.xml_doc(mro_doc)
                    measurement_list = mr_utils.xml_element(mro_root, 'measurement')
                    object_list_m0 = mr_utils.xml_element(   measurement_list[0], 'object'  )
                    value_list_o0 = mr_utils.xml_element(  object_list_m0[0] , 'v'  )
                    value_str_0 = mr_utils.xml_text( value_list_o0[0] )
                    # MRS earfcn value is same , get just one 
                    earfcn_value = mr_utils.get_split_value( value_str_0 ) [0] if len( measurement_list ) != 0 and len( object_list_m0 ) != 0 else '-1'

                    #获得子帧分帧
                    prbnum_list = mr_utils.get_prb_list()
                    subfram_list = mr_utils.get_split_value( gl.MR_CONF['SubFrameNum'], ',', [',', '{', '}'] )

                    for list_item_name in file_header_list_name:
                        if mr_utils.is_str_format_time( mr_utils.xml_attr(file_header_list[0], list_item_name), gl.TIME_FORMAT) == False:
                            mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '%s format error'%(list_item_name))
                            test_flag |= (0x1 << 2)
                    period_temp = mr_utils.xml_attr(file_header_list[0], 'period')
                    if  period_temp != gl.MR_CONF['UploadPeriod']:
                        mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, 'upload peroiod error :[0] !=  ' + period_temp )
                        test_flag |= (0x1 << 2)
                    jobid_temp  = int( mr_utils.xml_attr(file_header_list[0], 'jobid'))
                    if  jobid_temp < 0 or jobid_temp > 4294967295 :
                        mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, 'jobid :[0-4294967295] !=  ' + jobid_temp)
                        test_flag |= (0x1 << 2)

                    for measurement_entity in mr_utils.xml_element(mr_root, 'measurement'):
                        mr_name = mr_utils.xml_attr(measurement_entity, 'mrName')
                        if measureItem_list.__contains__(mr_name) == True:
                            measureItem_list[mr_name] += 1
                        else:
                            mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '<surplus mrItem> %s'%(mr_name))
                        smr_value = mr_utils.xml_text(    mr_utils.xml_element(measurement_entity, 'smr') [0]   )
                        object_list = mr_utils.xml_element(measurement_entity, 'object')
                        # RIP and prb special, deal the object_eciid
                        if mr_name == "MR.ReceivedIPower":
                            if len(subfram_list)*len(mr_utils.get_split_value(gl.TEST_CONF['cellid'], ',', [',', ' '])) != len(object_list):
                                mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, "MR.ReceivedIPower:object format error:[%s]" % (str(subfram_list[0:len(subfram_list)])))
                                test_flag |= (0x1 << 2)
                            for object_entity in object_list:
                                id_str = mr_utils.xml_attr(object_entity, 'id')
                                id_list = mr_utils.get_split_value(id_str, ':', [':', ' '])
                                if mr_utils.is_eci_correct(int(id_list[0])) == False or (id_list[1] != earfcn_value and earfcn_value != '-1') or id_list[2] not in subfram_list:
                                    mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name,"MR.ReceivedIPower:object format error:[%s]"%(id_str))
                                    test_flag |= (0x1 << 4)

                        elif mr_name == 'MR.RIPPRB':
                            if len(subfram_list)*len(prbnum_list)*len(gl.TEST_CONF['cellid'].strip(',').split(',')) != len(object_list):
                                mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, "MR.RIPPRB:object format error:[%s * %s]" % (str(subfram_list[0:len(subfram_list)]) , str(prbnum_list[0:len(prbnum_list)])))
                                test_flag |= (0x1 << 4)
                            for object_entity in object_list:
                                id_str = mr_utils.xml_attr(object_entity, 'id')
                                id_list = mr_utils.get_split_value(id_str, ':', [':' , ' '])
                                if mr_utils.is_eci_correct(int(id_list[0])) == False or (id_list[1] != earfcn_value and earfcn_value != '-1') or id_list[2] not in subfram_list or id_list[2] not in prbnum_list:
                                    mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, "MR.RIPPRB:object format error:[%s]"%(id_str))
                                    test_flag |= (0x1 << 4)
                        else:
                            if len(object_list) != len(mr_utils.get_split_value(gl.TEST_CONF['cellid'], ',', [',', ' '])):
                                mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '%s: object format error [num not match] - [%d]'%(mr_name,len(object_list)))
                                test_flag |= (0x1 << 4)
                            for object_entity in object_list:
                                id_str = mr_utils.xml_attr(object_entity, 'id')
                                id_list = mr_utils.get_split_value(id_str, ':', [':', ' '])
                                if mr_utils.is_eci_correct(int(id_list[0])) == False:
                                    mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '%s <object id not match > %s'%(mr_name, id_str))
                                    test_flag |= (0x1 << 4)
                        temp_value_flag = 0

                        for object_entity in object_list:
                            for value_entity in mr_utils.xml_element(object_entity, 'v'):
                                value_str = mr_utils.xml_text(value_entity)
                                if mr_utils.is_mrs_measurement_smr_value_correct(mr_name, smr_value, value_str) == False:
                                    mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '<smr or value not match> \n--->smr:%s \n---> value:%s'%(smr_value, value_str))
                                    test_flag |= (0x1 << 4)
                                    temp_value_flag = 1
                        if temp_value_flag == 0:
                            measureItem_list[mr_name] += 1
                
                gl.OUT_STR_LIST.append('\n======> ' + mr_file_full_name + ':  \n\t')
                gl.OUT_STR_LIST.append('N | N | ' if test_flag & (0x1 << 1) != 0 else 'Y | Y | ')
                gl.OUT_STR_LIST.append('N | N | ' if test_flag & (0x1 << 2) != 0 else 'Y | Y | ' )
                gl.OUT_STR_LIST.append('N | ' if test_flag & (0x1 << 3) != 0 else 'Y | ')
                gl.OUT_STR_LIST.append('N | \n' if test_flag & (0x1 << 4) != 0 else 'Y | \n')
            
        except Exception as result:
            raise Exception('[%s]   %s |  %s'%(__name__ ,str(result.__traceback__.tb_lineno) if not gl.IS_PY2  else '',result))
            # raise Exception('-%s- <%s> MRO:%s MRE:%s MRS:%s'%(str(result.__traceback__.tb_lineno),result, gl.MR_DICT[time_entity][0]['MRO'], gl.MR_DICT[time_entity][0]['MRE'], gl.MR_DICT[time_entity][0]['MRS']))
    if test_flag != 0:
        for file_name in out_text_dict:
            gl.OUT_STR_LIST.append('=====>' + file_name + ' :\n')
            for i in range(len(out_text_dict[file_name])):
                gl.OUT_STR_LIST.append('\t' + str(out_text_dict[file_name][i]) + '\n')
         
         
         
#========================================================================================#
#             test 73
# check the accuracy of MRO sample number,1.the number 2.time ascending ...
#========================================================================================#        
                
def test73_file_accuracy():
    # mr_utils.test_out_data_item_header("test_73")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)
    gl.OUT_STR_LIST.append(date_time + " | ")
    gl.OUT_STR_LIST.append(gl.TEST_CONF['test_total_time'] + " | \n" )
    max_mro_object_num = math.ceil( (int(gl.MR_CONF['UploadPeriod']) * 60 * 1000) / float(int(gl.MR_CONF['SamplePeriod'])))
    
    out_text_dict = {}
    
    
    for time_entity in gl.MR_DICT:
        try:
            mro_file_flag_dict = {}
            mro_file = mr_utils.get_file_name(time_entity, 'MRO')
            if mro_file == '':
                continue
            mro_dom = mr_utils.xml_dom(time_entity, 'MRO')
            mro_root = mr_utils.xml_doc(mro_dom)
            time_stamp_temp = 0

            for measurement_entity in mr_utils.xml_element(mro_root, 'measurement'):
                smr_list = mr_utils.xml_element(measurement_entity, 'smr')
                smr_str_s0 = mr_utils.xml_text(smr_list[0]) 
                smr_list_s0 = mr_utils.get_split_value(smr_str_s0)
                # ignore L2 data
                if smr_list_s0[0] == 'MR.LteScRIP' :
                    continue
                for object_entity in mr_utils.xml_element(measurement_entity, 'object'):
                    MmeUeS1apId = mr_utils.xml_attr(object_entity, 'MmeUeS1apId')
                    MmeGroupId = mr_utils.xml_attr(object_entity, 'MmeGroupId')
                    MmeCode = mr_utils.xml_attr(object_entity, 'MmeCode')
                    time_str = mr_utils.xml_attr(object_entity, 'TimeStamp')
                    mro_current_time_stamp = mr_utils.get_timestamp_by_str_format(time_str)
                    ue_mme_name = MmeUeS1apId  + '|' + MmeGroupId + '|' + MmeCode
                    if ue_mme_name not in mro_file_flag_dict:
                        # mro_object_num: count the object number,     is_ascend: if the time is ascending 
                        mro_file_flag_dict[ue_mme_name] = {'mro_object_num':0, 'is_ascend':0}
                    if time_stamp_temp != 0:
                        time_spec = mro_current_time_stamp - time_stamp_temp
                        if time_spec != int(gl.MR_CONF['SamplePeriod']):
                            mro_file_flag_dict[ue_mme_name]['is_ascend'] = 1
                            mr_utils.out_text_dict_append_list(out_text_dict, mro_file, "time spec: [%s]-[%s] != %s "%(time_str, mr_utils.get_time_format_by_timestamp(time_stamp_temp), gl.MR_CONF['SamplePeriod'] ))
                    # check the next timestamp
                    time_stamp_temp = mro_current_time_stamp
                    mro_file_flag_dict[ue_mme_name]['mro_object_num'] += 1
            
            gl.OUT_STR_LIST.append('======>' + mro_file + ": \n\t")
            gl.OUT_STR_LIST.append(str(max_mro_object_num) + ' | ')
            if len(mro_file_flag_dict) == 0:
                gl.OUT_STR_LIST.append("no L3 information\n")
                continue
            for ue_mme_name in mro_file_flag_dict:
                gl.OUT_STR_LIST.append( "[" + ue_mme_name + "]: [num = " + str(mro_file_flag_dict[ue_mme_name]['mro_object_num']) + " ]-[is_ascend: "+ str(mro_file_flag_dict[ue_mme_name]['is_ascend'] == 0) + " ]  ")
                if mro_file_flag_dict[ue_mme_name]['mro_object_num'] > max_mro_object_num or mro_file_flag_dict[ue_mme_name]['is_ascend'] != 0:
                    gl.OUT_STR_LIST.append(' | N |\n')
                else:
                    gl.OUT_STR_LIST.append(' | Y |\n')
        except Exception as result:
            raise Exception('[%s]   %s |  %s'%(__name__ ,str(result.__traceback__.tb_lineno) if not gl.IS_PY2  else '',result))

    for file_name in out_text_dict:
        gl.OUT_STR_LIST.append('=====>' + file_name + ' :\n')
        for i in range(len(out_text_dict[file_name])):
            gl.OUT_STR_LIST.append('\t' + str(out_text_dict[file_name][i]) + '\n')

        